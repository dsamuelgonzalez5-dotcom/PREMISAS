import streamlit as st
import pandas as pd
import re
import os
import sqlite3
import unicodedata
from datetime import datetime, timedelta
from io import BytesIO
import openpyxl 
import xml.etree.ElementTree as ET
from xml.dom import minidom
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Side, Font
import base64
import requests
import anthropic
import json

# ── API Key desde config.txt ──────────────────────────────────────────────────
def leer_api_key() -> str:
    config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.txt")
    try:
        with open(config_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#"):
                    return line
    except FileNotFoundError:
        pass
    return ""

_api_key_libranzas = leer_api_key()

# ── Carga de pasos de transformadores desde Excel ───────────────────────────
@st.cache_data
def cargar_pasos_transformadores() -> dict:
    ruta = (lambda f: f if os.path.exists(f) else os.path.join(os.getcwd(), "pasos_transformadores.xlsx"))(os.path.join(os.path.dirname(os.path.abspath(__file__)), "pasos_transformadores.xlsx"))
    try:
        wb = openpyxl.load_workbook(ruta, data_only=True)
        resultado = {}
        for sname in wb.sheetnames:
            if sname == "INDICE" or "NDICE" in sname:
                continue
            ws = wb[sname]
            inicio, fin = [], []
            for row in ws.iter_rows(min_row=3, values_only=True):
                if not row or not row[0]:
                    continue
                tipo_m = str(row[0]).strip().upper()
                desc   = str(row[2] or "").strip()
                ejec   = str(row[3] or "").strip()
                obs    = str(row[4] or "").strip()
                if not desc:
                    continue
                entry = {"paso": row[1], "descripcion": desc, "ejecutor": ejec, "observacion": obs}
                if "INICIO" in tipo_m:
                    inicio.append(entry)
                elif "FINALIZ" in tipo_m:
                    fin.append(entry)
            titulo = str(ws.cell(1,1).value or "")
            resultado[sname] = {"sheet": sname, "titulo": titulo, "inicio": inicio, "finalizacion": fin}
        return resultado
    except Exception:
        return {}


def buscar_pasos_transformador(transformador: str, subestacion: str):
    def norm(t):
        return unicodedata.normalize("NFKD", str(t)).encode("ascii","ignore").decode().upper().strip()
    todos = cargar_pasos_transformadores()
    if not todos:
        return None
    t_norm = norm(transformador)
    se_norm = norm(subestacion)
    mejor, mejor_score = None, 0
    for sname, data in todos.items():
        titulo = norm(data["titulo"])
        score = 0
        if t_norm in titulo:
            score += 3
        for parte in se_norm.split():
            if len(parte) > 3 and parte in titulo:
                score += 2
        if score > mejor_score:
            mejor_score = score
            mejor = data
    return mejor if mejor_score >= 3 else None


# ── Carga de pasos de líneas desde Excel ─────────────────────────────────
@st.cache_data
def cargar_pasos_lineas() -> dict:
    ruta = (lambda f: f if os.path.exists(f) else os.path.join(os.getcwd(), "pasos_lineas.xlsx"))(os.path.join(os.path.dirname(os.path.abspath(__file__)), "pasos_lineas.xlsx"))
    try:
        wb = openpyxl.load_workbook(ruta, data_only=True)
        resultado = {}
        for sname in wb.sheetnames:
            if "NDICE" in sname.upper() or "PLANTILLA" in sname.upper():
                continue
            ws = wb[sname]
            inicio, fin = [], []
            for row in ws.iter_rows(min_row=3, values_only=True):
                if not row or not row[0]:
                    continue
                tipo_m = str(row[0]).strip().upper()
                desc   = str(row[2] or "").strip()
                ejec   = str(row[3] or "").strip()
                obs    = str(row[4] or "").strip()
                if not desc:
                    continue
                entry = {"paso": row[1], "descripcion": desc, "ejecutor": ejec, "observacion": obs}
                if "INICIO" in tipo_m:
                    inicio.append(entry)
                elif "FINALIZ" in tipo_m:
                    fin.append(entry)
            titulo = str(ws.cell(1, 1).value or "")
            resultado[sname] = {"sheet": sname, "titulo": titulo, "inicio": inicio, "finalizacion": fin}
        return resultado
    except Exception:
        return {}


def buscar_pasos_linea(linea: str, subestacion: str = ""):
    """Busca los pasos de una línea en pasos_lineas.xlsx.

    El código de línea (ej. '230-3B') se busca en el nombre del sheet y en
    el título. Si hay varias hojas para la misma línea (variantes por
    subestación o número de libranza) se elige la de mayor score:
      +4 si el código de línea está en el nombre del sheet
      +3 si está en el título
      +2 por cada palabra de la subestación (>3 chars) que aparezca en el título
    Devuelve None si no hay coincidencia suficiente (score < 4).
    """
    def norm(t):
        return unicodedata.normalize("NFKD", str(t)).encode("ascii", "ignore").decode().upper().strip()

    todos = cargar_pasos_lineas()
    if not todos:
        return None

    linea_norm = norm(linea)
    se_norm    = norm(subestacion) if subestacion else ""
    mejor, mejor_score = None, 0

    for sname, data in todos.items():
        sname_norm = norm(sname)
        titulo_norm = norm(data["titulo"])
        score = 0
        if linea_norm in sname_norm:
            score += 4
        if linea_norm in titulo_norm:
            score += 3
        if se_norm:
            for parte in se_norm.split():
                if len(parte) > 3 and parte in titulo_norm:
                    score += 2
        if score > mejor_score:
            mejor_score = score
            mejor = data

    return mejor if mejor_score >= 4 else None


def formatear_pasos_para_prompt(pasos_data: dict) -> str:
    """Formatea los pasos de una hoja (transformador o línea) como texto para el prompt."""
    if not pasos_data:
        return ""
    lines = [f"PASOS REGISTRADOS PARA: {pasos_data['titulo']}", "",
             "MANIOBRAS DE INICIO:"]
    for p in pasos_data["inicio"]:
        obs = f" [{p['observacion']}]" if p["observacion"] else ""
        lines.append(f"  P{p['paso']}: {p['descripcion']} — Por: {p['ejecutor']}{obs}")
    lines += ["", "MANIOBRAS DE FINALIZACION:"]
    for p in pasos_data["finalizacion"]:
        obs = f" [{p['observacion']}]" if p["observacion"] else ""
        lines.append(f"  P{p['paso']}: {p['descripcion']} — Por: {p['ejecutor']}{obs}")
    return "\n".join(lines)


# ── Carga de equipos desde Excel ─────────────────────────────────────────────
@st.cache_data
def cargar_equipos_libranzas() -> list:
    """Carga la lista maestra de equipos desde equipos_libranzas.xlsx.
    Columnas: B=Codigo, D=Agente, E=Instalacion, G=Tipo de Equipo, H=Agentes Extra, I=Status."""
    ruta = (lambda f: f if os.path.exists(f) else os.path.join(os.getcwd(), "equipos_libranzas.xlsx"))(os.path.join(os.path.dirname(os.path.abspath(__file__)), "equipos_libranzas.xlsx"))
    try:
        wb = openpyxl.load_workbook(ruta, data_only=True)
        ws = wb.active
        equipos = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            codigo      = str(row[1] or "").strip().upper()   # col B = Código
            agente      = str(row[2] or "").strip().upper()   # col C = Agente
            instalacion = str(row[3] or "").strip().upper()   # col D = Instalación
            tipo        = str(row[4] or "").strip().upper()   # col E = Tipo de Equipo
            agentes_ext = str(row[5] or "").strip().upper()   # col F = Agentes Extra
            if not codigo:
                continue
            if "->" in instalacion:
                partes    = instalacion.split("->")
                categoria = partes[0].strip()
                nombre    = partes[1].strip()
            else:
                categoria = ""
                nombre    = instalacion
            # Parsear agentes extra (puede venir como " ETESA, ENSA")
            extras = [a.strip() for a in agentes_ext.replace(";", ",").split(",") if a.strip() and a.strip() != "NONE"]
            equipos.append({
                "codigo":        codigo,
                "instalacion":   nombre,
                "categoria":     categoria,
                "tipo":          tipo,
                "agente":        agente,
                "agentes_extra": extras,
            })
        return equipos
    except FileNotFoundError:
        return []
    except Exception:
        return []


@st.cache_data
def cargar_responsables_campo() -> list:
    """Carga la lista de responsables de campo desde responsables-campo.xlsx.
    Columnas: A=Id, B=Nombre, C=Apellido, D=Celular, E=Agente, F=Cargo.
    Permite identificar que 'CUBILLA' o 'NICOLAY SALADO' son personal de ETESA."""
    ruta = (lambda f: f if os.path.exists(f) else os.path.join(os.getcwd(), "responsables-campo.xlsx"))(os.path.join(os.path.dirname(os.path.abspath(__file__)), "responsables-campo.xlsx"))
    try:
        wb = openpyxl.load_workbook(ruta, data_only=True)
        ws = wb.active
        responsables = []
        for row in ws.iter_rows(min_row=3, values_only=True):   # fila 1=título, fila 2=cabecera
            if not row or not row[2]:
                continue
            nombre   = str(row[1] or "").strip().upper()
            apellido = str(row[2] or "").strip().upper()
            agente   = str(row[4] or "").strip().upper()
            if not apellido or not agente:
                continue
            responsables.append({
                "nombre":          nombre,
                "apellido":        apellido,
                "agente":          agente,
                "nombre_completo": f"{nombre} {apellido}".strip(),
            })
        return responsables
    except FileNotFoundError:
        return []
    except Exception:
        return []


def validar_equipo_en_lista(equipo_principal: str, subestacion: str, tipo_equipo: str) -> dict:
    """Valida que el equipo de la libranza exista en la lista maestra y pertenezca a la S/E correcta."""
    equipos = cargar_equipos_libranzas()
    if not equipos:
        return {"estado": "ADVERTENCIA", "detalle": "No se encontró equipos_libranzas.xlsx. Colócalo en la misma carpeta que el app."}

    def _extraer_codigo(texto: str) -> str:
        """
        Intenta extraer el código puro del equipo desde el texto que entrega el AI.
        El AI puede devolver el nombre completo como "LINEA DE CONEXION 230-7"
        o solo el código "230-7" / "23B22".
        Estrategia: buscar el último token que parezca un código eléctrico.
        """
        texto = texto.strip().upper()
        # Patrón de código: NNN-N, 23M32, 115C1, etc.
        m = re.findall(r'\b(\d{2,3}(?:[A-Z]{1,4}\d{1,3}[A-Z]?|-\d{1,3}[A-Z]?))\b', texto)
        return m[-1] if m else texto   # último código encontrado, o el texto completo si no hay

    # Para líneas dobles puede venir "230-3C / 230-4C" — validar cada una
    partes = [e.strip() for e in equipo_principal.replace("/", ",").split(",")]
    resultados = []

    for raw in partes:
        eq  = _extraer_codigo(raw)          # código limpio
        se  = subestacion.strip().upper()

        # Buscar coincidencia exacta de código
        encontrados = [e for e in equipos if e["codigo"] == eq]

        if not encontrados:
            # Fallback: búsqueda flexible por similitud
            sugerencias = list({e["codigo"] for e in equipos if eq[:4] in e["codigo"]})[:3]
            sug_txt = f" ¿Quisiste decir: {', '.join(sugerencias)}?" if sugerencias else ""
            resultados.append({"estado": "ERROR", "msg": f"'{eq}' NO existe en la lista maestra.{sug_txt}"})
            continue

        tipo_upper = tipo_equipo.upper()
        if "LINEA" in tipo_upper or encontrados[0].get("categoria","").upper() == "LINEA":
            # Para líneas solo verificar que exista el código
            resultados.append({"estado": "OK", "msg": f"'{eq}' verificado ({encontrados[0]['tipo']})."})
        else:
            def normalizar(t):
                return unicodedata.normalize("NFKD", t).encode("ascii","ignore").decode().upper().strip()

            se_norm  = normalizar(se)
            match_se = [e for e in encontrados if normalizar(e["instalacion"]) in se_norm or se_norm in normalizar(e["instalacion"])]

            if not match_se:
                se_correctas = list({e["instalacion"] for e in encontrados})
                resultados.append({"estado": "ERROR",
                    "msg": f"'{eq}' existe pero pertenece a: {', '.join(se_correctas)}. La libranza indica: '{subestacion}'."})
            else:
                resultados.append({"estado": "OK",
                    "msg": f"'{eq}' verificado en {match_se[0]['instalacion']} ({match_se[0]['tipo']})."})

    hay_error    = any(r["estado"] == "ERROR" for r in resultados)
    estado_final = "ERROR" if hay_error else "OK"
    detalle_final = " | ".join(r["msg"] for r in resultados)
    return {"estado": estado_final, "detalle": detalle_final}


# ── Validación de equipos en pasos y agentes contra la DB ────────────────────

def _norm_db(t: str) -> str:
    """Normaliza texto para comparaciones: sin tildes, mayúsculas, sin espacios extra."""
    return unicodedata.normalize("NFKD", str(t)).encode("ascii", "ignore").decode().upper().strip()


def _ejecutor_es_agente(ejecutor: str, agente_db: str, tipo_equipo: str,
                         responsables: list,
                         es_accion_campo: bool = False,
                         agentes_extra: list = None) -> tuple[bool, str]:
    """
    Verifica si el ejecutor de un paso es valido para el equipo dado.
    Retorna (es_valido: bool, motivo_error: str).
    """
    if not ejecutor or not agente_db:
        return True, ""

    ej_n  = _norm_db(ejecutor)
    ag_n  = _norm_db(agente_db)
    tip_n = _norm_db(tipo_equipo)
    extras_n = [_norm_db(a) for a in (agentes_extra or [])]

    # ── Agente no ETESA ───────────────────────────────────────────────────────
    if ag_n != "ETESA":
        if ag_n in ej_n or ej_n in ag_n:
            return True, ""
        # Verificar si algun agente extra puede operar el equipo
        for ex_n in extras_n:
            if ex_n in ej_n or ej_n in ex_n:
                return True, ""
            # Si ETESA es agente extra, aceptar CND y responsables de campo ETESA
            if ex_n == "ETESA":
                if ej_n == "CND":
                    return True, ""
                es_resp_etesa = bool(responsables) and any(
                    _norm_db(r["apellido"]) in ej_n or ej_n in _norm_db(r["nombre_completo"])
                    for r in responsables if _norm_db(r.get("agente","")) == "ETESA"
                )
                if es_resp_etesa:
                    return True, ""
        return False, f"Debe operarlo {agente_db} (DB), el paso dice: {ejecutor}"

    # ── Equipo ETESA ──────────────────────────────────────────────────────────
    es_interruptor = "INTERRUPTOR" in tip_n
    es_motorizada  = "MOTORIZ"     in tip_n

    # Interruptores y cuchillas motorizadas: solo CND, EXCEPTO si es
    # una acción de campo (verificar posición, entregar equipo).
    if (es_interruptor or es_motorizada) and not es_accion_campo:
        if ej_n == "CND":
            return True, ""
        es_resp = bool(responsables) and any(
            _norm_db(r["apellido"]) in ej_n or ej_n in _norm_db(r["nombre_completo"])
            for r in responsables if _norm_db(r["agente"]) == "ETESA"
        )
        tipo_label = "interruptores" if es_interruptor else "cuchillas motorizadas"
        if es_resp:
            return False, (f"La operación de {tipo_label} ETESA es exclusiva del CND. "
                           f"'{ejecutor}' es responsable de campo, no puede hacerlo")
        return False, f"Debe operarlo CND (DB), el paso dice: {ejecutor}"

    # Para todo lo demás (cuchillas de tierra, otros tipos, verificar, entregar):
    # CND siempre es válido
    if ej_n == "CND":
        return True, ""

    # Si no hay lista de responsables cargada, no podemos validar → skip sin error
    if not responsables:
        return True, ""

    # Buscar apellido o nombre completo en la lista de responsables de campo ETESA
    for r in responsables:
        if _norm_db(r["agente"]) != "ETESA":
            continue
        if _norm_db(r["apellido"]) in ej_n or ej_n in _norm_db(r["nombre_completo"]):
            return True, ""

    return False, f"'{ejecutor}' no está registrado como responsable de campo de ETESA"


def _extraer_codigos_de_texto(texto: str) -> list:
    """
    Extrae posibles códigos de equipos eléctricos de un texto de paso.
    Reconoce cuchillas (23LA20), interruptores (23B22), líneas (230-3A), bancos (115 C1).
    """
    texto_up = re.sub(r'\s+', ' ', texto.upper())
    patrones = [
        r'\b(\d{2,3}[A-Z]{1,4}\d{1,3}[A-Z]?)\b',    # 23LA20, 11B15, 23LB20, 23A22
        r'\b(\d{2,3}-\d{1,3}[A-Z]?)\b',               # 230-3A, 115-10
        r'\b(\d{2,3}\s+[A-Z]\d{1,2})\b',              # 115 C1, 115 C2
    ]
    codigos = set()
    for p in patrones:
        for m in re.finditer(p, texto_up):
            cod = re.sub(r'\s+', '', m.group(1))
            if len(cod) >= 4:
                codigos.add(cod)
    return list(codigos)


def _extraer_se_de_texto(texto: str) -> str:
    """Extrae el nombre de S/E mencionado en el texto del paso."""
    texto_up = texto.upper()
    for patron in [
        # Requiere barra literal "S/E" para no confundir con el pronombre español "SE"
        r'\bS/E\.?\s+([A-Z][A-Z0-9\s]{2,30})(?=\s+(?:LA|EL|LOS|DE|[A-Z]{1,2}\d|\+|\.|,|$))',
        r'\bS/E\.?\s+([A-Z][A-Z0-9\s]{2,30})',
        r'SUBESTACION\s+([A-Z][A-Z0-9\s]{2,30})',
    ]:
        m = re.search(patron, texto_up)
        if m:
            nombre = m.group(1).strip()
            # Quitar sufijos tipo "AL CND", "A CND", "AL AREA"
            nombre = re.sub(r'\s+AL?\s+(?:CND|AREA)$', '', nombre).strip()
            return _norm_db(nombre)
    return ""

def validar_paso_vs_db(descripcion: str, ejecutor: str,
                        subestacion_libranza: str, equipos_db: list) -> list:
    """
    Valida TODOS los códigos de equipo PRINCIPALES encontrados en la descripción.

    Reglas especiales:
      • Pasos de ENTREGA (contienen 'ENTREGAR'): no se valida contra DB.
        El prompt ya define quién entrega/recibe — no hay nada que cruzar.
      • Código de LÍNEA (ej: 230-7) junto a otros códigos de S/E (23N7, 23M32…):
        la línea aparece solo como contexto ("DE LA LÍNEA 230-7 EN LA S/E…"),
        no se valida. Solo se valida el código de línea cuando es el ÚNICO código
        encontrado en el paso.
      • LÍNEA como único código: solo verificar existencia (sin chequeo de S/E ni agente).
      • VERIFICAR steps: es_accion_campo=True → responsable de campo válido.
    """
    if not equipos_db or not descripcion:
        return []

    desc_up = descripcion.upper()

    # ── Pasos que no requieren validación DB ─────────────────────────────────
    # ENTREGAR: el prompt define quién entrega/recibe — no hay nada que cruzar.
    # RECIERRE:  verificar recierre lo hace CND según el prompt — sin chequeo DB.
    if "ENTREGAR" in desc_up or "ENTREGA" in desc_up or "RECIERRE" in desc_up:
        return []

    # Índice multi-candidato: código → lista de equipos
    db_idx: dict[str, list] = {}
    for eq in equipos_db:
        clave = re.sub(r'\s+', '', eq['codigo'])
        db_idx.setdefault(clave, []).append(eq)

    codigos_todos   = _extraer_codigos_de_texto(descripcion)
    se_en_paso      = _extraer_se_de_texto(descripcion)
    # Limpiar sufijos tipo "AL CND", "AL AREA", "A CND" del nombre de SE
    se_en_paso = re.sub(r'\s+(AL?\s+CND|AL\s+AREA|A\s+CND)\s*$', '', se_en_paso.strip()).strip() if se_en_paso else se_en_paso
    se_libranza     = _norm_db(subestacion_libranza)
    ejecutor_n      = _norm_db(ejecutor) if ejecutor else ""
    responsables    = cargar_responsables_campo()
    es_accion_campo = "VERIFICAR" in desc_up

    # ── Separar códigos de línea (patrón NNN-N) de códigos de S/E ────────────
    patron_linea = re.compile(r'^\d{2,3}-\d{1,3}[A-Z]?$')
    codigos_se   = [c for c in codigos_todos if not patron_linea.match(c)]
    codigos_lin  = [c for c in codigos_todos if     patron_linea.match(c)]

    # Si hay equipos de S/E, los códigos de línea son solo contexto → ignorarlos
    codigos = codigos_se if codigos_se else codigos_lin

    resultados = []

    for cod in codigos:
        candidatos = db_idx.get(cod)
        if not candidatos:
            continue

        eq_ref = candidatos[0]

        # ── LÍNEA como único código: solo verificar existencia ────────────────
        if _norm_db(eq_ref.get("categoria", "")) == "LINEA":
            resultados.append({
                "codigo":         cod,
                "tipo":           eq_ref["tipo"],
                "agente_db":      eq_ref["agente"],
                "instalacion_db": eq_ref["instalacion"],
                "ejecutor":       ejecutor,
                "estado":         "OK",
                "errores":        [],
            })
            continue

        # ── Equipos de S/E: buscar candidato que coincida con la S/E ──────────
        se_ref = se_en_paso if se_en_paso else se_libranza

        eq_match = None
        if se_ref:
            eq_match = next(
                (c for c in candidatos
                 if se_ref in _norm_db(c['instalacion']) or _norm_db(c['instalacion']) in se_ref),
                None
            )

        inst_ok = eq_match is not None
        eq      = eq_match if eq_match else candidatos[0]

        agente_ok, motivo_agente = _ejecutor_es_agente(
            ejecutor, eq['agente'], eq['tipo'], responsables,
            es_accion_campo=es_accion_campo,
            agentes_extra=eq.get('agentes_extra', [])
        )

        errores = []
        if not inst_ok and se_ref:
            se_encontradas = ", ".join(sorted({c['instalacion'] for c in candidatos}))
            errores.append(f"'{cod}' no existe en {se_ref}. En la DB está en: {se_encontradas}")
        if not agente_ok and ejecutor_n:
            errores.append(motivo_agente)

        resultados.append({
            "codigo":         cod,
            "tipo":           eq['tipo'],
            "agente_db":      eq['agente'],
            "instalacion_db": eq['instalacion'],
            "ejecutor":       ejecutor,
            "estado":         "ERROR" if errores else "OK",
            "errores":        errores,
        })

    return resultados


def validar_agentes_vs_db(equipo_principal: str, datos_generales: list,
                           equipos_db: list) -> dict:
    """
    Compara los agentes extra registrados en la DB para el equipo principal
    contra los agentes involucrados/informados extraídos de la libranza.
    """
    if not equipos_db or not equipo_principal:
        return {"estado": "SIN_DATOS", "detalle": "", "agentes_db": []}

    eq_n = _norm_db(re.sub(r'\s+', '', equipo_principal))
    encontrado = next(
        (eq for eq in equipos_db if re.sub(r'\s+', '', eq['codigo']) == eq_n), None
    )

    if not encontrado:
        return {"estado": "SIN_DATOS", "detalle": f"Equipo {equipo_principal} no hallado en DB.", "agentes_db": []}

    extras = encontrado.get("agentes_extra", [])
    if not extras:
        return {"estado": "OK", "detalle": "No hay agentes extra en la DB para este equipo.", "agentes_db": []}

    # Texto combinado de los campos de agentes de datos_generales
    texto_agentes = " ".join(
        _norm_db(dg.get("detalle", ""))
        for dg in datos_generales
        if any(k in _norm_db(dg.get("campo", "")) for k in ["INVOLUCRADO", "INFORMADO"])
    )

    faltantes  = [ag for ag in extras if _norm_db(ag) not in texto_agentes]
    presentes  = [ag for ag in extras if _norm_db(ag) in texto_agentes]

    if faltantes:
        return {
            "estado": "ADVERTENCIA",
            "detalle": f"Agentes de la DB no encontrados en la libranza: {', '.join(faltantes)}",
            "agentes_db": extras,
            "presentes":  presentes,
            "faltantes":  faltantes,
        }

    return {
        "estado": "OK",
        "detalle": f"Todos los agentes de la DB están en la libranza: {', '.join(extras)}",
        "agentes_db": extras,
        "presentes":  extras,
        "faltantes":  [],
    }


def validar_consistencia_responsables(todos_pasos: list) -> None:
    """
    Verifica que cada responsable de campo ETESA opere en una sola S/E.
    Modifica in-place las validaciones_db afectadas agregando advertencias.
    """
    from collections import defaultdict
    responsables = cargar_responsables_campo()

    ejecutor_ses: dict = defaultdict(set)
    for paso in todos_pasos:
        for vdb in paso.get("validaciones_db", []):
            if _norm_db(vdb.get("agente_db", "")) != "ETESA":
                continue
            ej = _norm_db(vdb.get("ejecutor", ""))
            if not ej or ej == "CND":
                continue
            es_conocido = any(
                _norm_db(r["apellido"]) in ej or ej in _norm_db(r["nombre_completo"])
                for r in responsables if _norm_db(r["agente"]) == "ETESA"
            )
            if not es_conocido:
                continue
            desc = paso.get("descripcion_encontrada") or paso.get("descripcion_esperada") or ""
            se = _extraer_se_de_texto(desc)
            if se:
                ejecutor_ses[ej].add(se)

    conflictos = {ej: ses for ej, ses in ejecutor_ses.items() if len(ses) > 1}
    if not conflictos:
        return

    for paso in todos_pasos:
        for vdb in paso.get("validaciones_db", []):
            if _norm_db(vdb.get("agente_db", "")) != "ETESA":
                continue
            ej = _norm_db(vdb.get("ejecutor", ""))
            if ej not in conflictos:
                continue
            desc = paso.get("descripcion_encontrada") or paso.get("descripcion_esperada") or ""
            se = _extraer_se_de_texto(desc)
            if not se:
                continue
            ses_txt = " y ".join(sorted(conflictos[ej]))
            msg = (f"Conflicto de presencia: '{ej}' aparece operando en {ses_txt} "
                   f"— una persona no puede estar en dos S/Es simultáneamente")
            if msg not in vdb.get("errores", []):
                vdb.setdefault("errores", []).append(msg)
                if vdb["estado"] == "OK":
                    vdb["estado"] = "ADVERTENCIA"


def enriquecer_pasos_con_db(maniobras: list, subestacion: str, equipos_db: list) -> list:
    """Agrega el campo 'validaciones_db' (lista) a cada paso con todos los equipos validados."""
    resultado = []
    for paso in maniobras:
        p = dict(paso)
        desc = p.get("descripcion_encontrada") or p.get("descripcion_esperada") or ""
        ejec = p.get("ejecutor_encontrado") or p.get("ejecutor_esperado") or ""
        p["validaciones_db"] = validar_paso_vs_db(desc, ejec, subestacion, equipos_db)
        resultado.append(p)
    return resultado


# ── GitHub: URL base del repositorio (raw, rama main) ────────────────────────
GITHUB_RAW = "https://raw.githubusercontent.com/LilisophiG24/Reportes-del-EOR/main"

# ── Función para cargar imagen como base64 ───────────────────────────────────
def get_base64_image(image_path):
    """Carga una imagen local y la convierte a base64 para incrustarla en HTML."""
    try:
        with open(image_path, "rb") as f:
            data = f.read()
        return base64.b64encode(data).decode("utf-8")
    except Exception:
        return None

# --- 1. CONFIGURACIÓN INICIAL ---
# PRIMERA LÍNEA  EN EJECUTARSE
st.set_page_config(layout="wide", initial_sidebar_state="expanded", page_title="Gestor De Reportes")

# Estilos CSS inyectados

st.markdown("""
    <style>

    /* === TEMA OSCURO — elementos especificos ===

    .stApp, [data-testid="stAppViewContainer"] {
        background-color: #111418 !important;
    }
    .block-container, .stMainBlockContainer {
        background-color: #111418 !important;
    }
    section[data-testid="stSidebar"] > div:first-child {
        background-color: #0e1117 !important;
    }
    [data-testid="stTextInput"] input,
    [data-testid="stNumberInput"] input,
    [data-testid="stTextArea"] textarea,
    [data-testid="stDateInput"] input {
        background-color: #1a1f27 !important;
        color: #e8e4dc !important;
        border-color: #2e3540 !important;
    }
    [data-testid="stSelectbox"] > div > div,
    [data-testid="stMultiSelect"] > div > div {
        background-color: #1a1f27 !important;
        color: #e8e4dc !important;
        border-color: #2e3540 !important;
    }
    [data-baseweb="popover"] ul, [data-baseweb="menu"] {
        background-color: #1a1f27 !important;
    }
    [role="option"] { background-color: #1a1f27 !important; color: #e8e4dc !important; }
    [role="option"]:hover { background-color: #2e3540 !important; }
    [data-testid="stExpander"] {
        background-color: #1a1f27 !important;
        border-color: #2e3540 !important;
    }
    [data-testid="stExpander"] summary { color: #e8e4dc !important; }
    [data-testid="stFileUploader"] {
        background-color: #1a1f27 !important;
        border-color: #2e3540 !important;
    }
    [data-testid="stFileUploader"] span,
    [data-testid="stFileUploader"] p { color: #e8e4dc !important; }
    [data-testid="stTabs"] [role="tab"] { color: #8a9099 !important; }
    [data-testid="stTabs"] [role="tab"][aria-selected="true"] {
        color: #4cde8f !important;
        border-bottom-color: #4cde8f !important;
    }
    .stMarkdown p, .stMarkdown li, .stMarkdown h1,
    .stMarkdown h2, .stMarkdown h3, .stMarkdown h4 {
        color: #e8e4dc !important;
    }
    [data-testid="stRadio"] label span,
    [data-testid="stCheckbox"] label span { color: #e8e4dc !important; }
    ::-webkit-scrollbar { width: 6px; height: 6px; }
    ::-webkit-scrollbar-track { background: #111418; }
    ::-webkit-scrollbar-thumb { background: #2e3540; border-radius: 3px; }
    ::-webkit-scrollbar-thumb:hover { background: #4cde8f; }

    /* === HOVER EFFECTS === */

    /* Sidebar: opciones del radio button */
    [data-testid="stSidebar"] [data-testid="stRadio"] label {
        display: block;
        padding: 6px 10px;
        border-radius: 6px;
        transition: background 0.18s ease;
        cursor: pointer;
    }
    [data-testid="stSidebar"] [data-testid="stRadio"] label:hover {
        background-color: rgba(100, 110, 130, 0.18) !important;
    }

    /* Botones primarios */
    .stButton > button {
        transition: filter 0.18s ease, transform 0.12s ease !important;
    }
    .stButton > button:hover {
        filter: brightness(1.15) !important;
        transform: translateY(-1px) !important;
    }
    .stButton > button:active {
        transform: translateY(0px) !important;
        filter: brightness(0.95) !important;
    }

    /* Download buttons */
    [data-testid="stDownloadButton"] > button {
        transition: filter 0.18s ease, transform 0.12s ease !important;
    }
    [data-testid="stDownloadButton"] > button:hover {
        filter: brightness(1.15) !important;
        transform: translateY(-1px) !important;
    }

   /* === SIDEBAR FIJO — SIEMPRE VISIBLE ===
    section[data-testid="stSidebar"] {
        display: flex !important;
        visibility: visible !important;
        transform: translateX(0) !important;
        width: 3rem !important;
        min-width: 1rem !important;
        position: fixed !important;
        top: 0 !important;
        left: 0 !important;
        height: 100vh !important;
        z-index: 9999 !important;
        overflow-y: auto !important;
    }
    /* Ocultar botón de colapsar (ya no hace falta) */
    [data-testid="stSidebarCollapseButton"],
    [data-testid="collapsedControl"] {
        display: none !important;
    }
    /* Header y footer de Streamlit: ocultos */
    header[data-testid="stHeader"],
    footer { display: none !important; }
    #MainMenu, .stDeployButton { visibility: hidden !important; }
    /* Contenido principal: desplazado para no quedar tapado por el sidebar */
    .stApp { overflow: visible !important; }
    .stMain > .block-container,
    .stMainBlockContainer {
        margin-left: 2rem !important;
        padding: 0 !important;
        max-width: calc(100% - 6.5rem) !important;
    }

/* Eliminar padding del contenedor principal */
.block-container {
    padding: 0 !important;
    margin: 0 !important;
    max-width: 100% !important;
    }

/* Eliminar padding del app root */
.stMainBlockContainer {
    padding: 0 !important;
    }

/* Forzar el banner a salir de cualquier margen residual */
.hero-banner {
    margin-left: -1rem !important;
    margin-right: -1rem !important;
    margin-top: -1rem !important;
    border-radius: 0 !important;   /* opcional: quitar bordes redondeados si quieres full edge */
    }
    /* Hero banner */
    .hero-banner {
        width: 101%;
        height: 42vh;
        min-height: 220px;
        max-height: 420px;
        background-size: cover;
        background-position: center 70%;
        background-repeat: no-repeat;
        border-radius: 0 0 18px 18px;
        position: relative;
        margin-bottom: 1.5rem;
        box-shadow: 0 6px 32px rgba(0,0,0,0.28);
        overflow: hidden;
    }
    /* Gradient overlay: bottom fade so content below reads cleanly */
    .hero-banner::after {
        content: "";
        position: absolute;
        inset: 0;
        background: linear-gradient(
            to bottom,
            rgba(0,0,0,0.08) 0%,
            rgba(0,0,0,0.15) 55%,
            rgba(14,17,23,0.72) 100%
        );
        border-radius: inherit;
    }
    /* Text overlay inside banner */
    .hero-text {
        position: absolute;
        top: 10%;        /* ← ajusta este % para subir/bajar */
        left: 32px;
        z-index: 2;
        color: #ffffff;
        text-shadow: 0 2px 12px rgba(0,0,0,0.7);
    }
    </style>
""", unsafe_allow_html=True)

# ── Hero banner con imagen ────────────────────────────────────────────────────
_img_b64 = get_base64_image("fondo_pagina.jpg")
if _img_b64:
    st.markdown(f"""
        <div class="hero-banner"
             style="background-image: url('data:image/jpeg;base64,{_img_b64}');">
            
        </div>
    """, unsafe_allow_html=True)


# --- 2. CONSTANTES Y VARIABLES GLOBALES ---
COLUMNAS_RTR = [
    'L', 'Tipo de Elemento', 'Tipo de indisponibilidad', 'Tipo de Causa', 
    'Nodo I', 'Nodo J', 'Nodo K', 'CKT', 'F. Indisponibilidad (dd-mmm-aaaa)', 
    'Hora Inicio CA+1', 'Hora Fin CA+1', 'Hora de duración', 
    'Energía no servida (MWh)', 'Descripción'
]

# --- 3. FUNCIONES DE BASE DE DATOS ---

def init_databases():
    """Inicializa todas las tablas necesarias al arrancar la app."""
    # Tabla de Unidades
    with sqlite3.connect('unidades_reporte.db') as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS reporte_unidades (
                unidad TEXT PRIMARY KEY, 
                salida TEXT, 
                causa TEXT, 
                libranza TEXT, 
                entrada TEXT, 
                potencia REAL
            )
        """)
    
    # Tabla de Líneas RTR (Unificada en un solo archivo)
    with sqlite3.connect('rtr_seguimiento.db') as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS lineas_abiertas (
                linea TEXT PRIMARY KEY, 
                descripcion_manual TEXT
            )
        """)

    # Tabla de Centrales en Vertimiento
    with sqlite3.connect('vertimiento.db') as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS reporte_vertimiento (
                central TEXT PRIMARY KEY,
                fecha_inicio TEXT,
                condicion TEXT
            )
        """)

    # Tabla de Eventos ASEP
    with sqlite3.connect('eventos_asep.db') as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS eventos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                numero INTEGER,
                fecha TEXT,
                elaborado_por TEXT,
                datos_json TEXT
            )
        """)

init_databases()

def limpiar_tabla_unidades():
    try:
        with sqlite3.connect('unidades_reporte.db') as conn:
            conn.execute("DELETE FROM reporte_unidades")
        st.success("Tabla de unidades vaciada correctamente.")
    except Exception as e:
        st.error(f"Error al limpiar: {e}")

# --- 4. FUNCIONES DE CARGA Y PROCESAMIENTO (CACHED) ---

@st.cache_data
def cargar_potencias():
    """Carga el archivo de potencias y lo guarda en caché."""
    try:
        df = pd.read_excel("potencia unidades.xlsx")
        df.columns = [str(c).strip() for c in df.columns]
        diccionario = {}
        for _, fila in df.iterrows():
            prefijo = str(fila.get('Prefijo', '')).strip().upper()
            try:
                valor = float(fila.get('Potencia por unidad(MW)', 0))
            except:
                valor = 0.0
            if prefijo and prefijo != 'NAN':
                diccionario[prefijo] = valor
        return diccionario
    except Exception as e:
        st.error(f"Error cargando excel de potencias: {e}")
        return {}

@st.cache_data
def cargar_equipos():
    """Carga el archivo de equipos y lo guarda en cache.
    Soporta .xls y .xlsx. Usa multiples engines para compatibilidad con Streamlit Cloud."""
    try:
        for nombre in ["equipos.xls", "equipos.xlsx"]:
            try:
                df = None
                engines = [None, "xlrd", "openpyxl"] if nombre.endswith(".xls") else [None, "openpyxl"]
                for eng in engines:
                    try:
                        kw = {"engine": eng} if eng else {}
                        df = pd.read_excel(nombre, header=0, **kw)
                        break
                    except Exception:
                        continue
                if df is None:
                    continue
                if not df.empty:
                    primera = str(df.iloc[0].values).upper()
                    if "NODO" in primera or "CKT" in primera or "BUS" in primera:
                        df = df.iloc[1:].reset_index(drop=True)
                return df
            except FileNotFoundError:
                continue
        return pd.DataFrame()
    except Exception:
        return pd.DataFrame()

def calcular_duracion(h_inicio, h_fin):
    """Calcula la diferencia entre dos horas en formato HH:MM."""
    try:
        if not h_inicio or not h_fin or ":" not in str(h_inicio) or ":" not in str(h_fin):
            return ""
        fmt = '%H:%M'
        inicio = datetime.strptime(str(h_inicio), fmt)
        fin = datetime.strptime(str(h_fin), fmt)
        if fin < inicio:
            duracion = (fin + timedelta(days=1)) - inicio
        else:
            duracion = fin - inicio
        total_segundos = int(duracion.total_seconds())
        horas = total_segundos // 3600
        minutos = (total_segundos % 3600) // 60
        return f"{horas:02d}:{minutos:02d}"
    except:
        return ""

def sin_tildes(texto):
    """Elimina tildes y caracteres especiales para comparaciones robustas."""
    return unicodedata.normalize('NFKD', str(texto)).encode('ascii', 'ignore').decode('ascii').upper()

def limpiar_y_formatear_desc(texto):
    """Limpia el texto de la bitácora para dejar solo lo relevante."""
    texto = str(texto).upper()
    match_sol = re.search(r'(SOLMANT\s*[\d\-\s]+)', texto)
    solmant_part = match_sol.group(1).strip() if match_sol else ""
    resumen = re.sub(r'^\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2}\s*', '', texto)
    if solmant_part:
        resumen_limpio = resumen.replace(solmant_part, "").strip(" :,-")
        return f"{solmant_part}: {resumen_limpio}"
    return resumen.strip()

# --- 5. LÓGICA DE NEGOCIO ---

def procesar_bitacora_unidades(df_bitacora):
    """Extrae eventos de indisponibilidad de unidades generadoras desde la bitácora.

    Mejoras sobre la versión anterior:
    - Detección de columnas: identifica automáticamente la columna de hora, fecha y
      descripción en lugar de concatenar todo ciegamente.
    - Normalización unicode: maneja acentos (SINCRONIZÓ, FINALIZÓ, etc.).
    - Libranza: patrón preciso ETESA/CNE-NNN-YYYY (igual que módulo RTR).
    - Causa: busca "POR <motivo>" antes de caer al fallback de split.
    - Skip de líneas/barras: solo descarta la fila cuando el tema principal NO es
      una unidad generadora; si el texto contiene un código XxxG# válido, sigue.
    - Múltiples salidas por unidad: se registran todas (no solo la última).
    - Keywords ampliados para ENTRADA y SALIDA.
    """
    try:
        dict_potencias = cargar_potencias()

        defaults = {
            'EAL': 24.0, 'CH1G1': 100, 'CH1G2': 100, 'EST': 60.0, 'LYE': 2.33, 'BAY': 86.67,
            'ALG': 4.93, 'BAI': 43.8, 'BDT': 6.3, 'BBL': 14.42, 'BON': 10.6, 'MLI': 17.21,
            'BU1': 1.71, 'BU2': 1.95, 'COC': 7.75, 'CON': 5, 'DOL': 1.04,
            'EFR': 2.22, 'FOR': 100, 'GUA': 12.65, 'LES': 23.6, 'LAP': 10.02,
            'LCR': 6.27, 'LPS': 5, 'LPN': 5, 'LOR': 17.5, 'LP1': 1.58,
            'LP2': 4.445, 'LVA': 27.4, 'PDO': 16.3, 'PAA': 3, 'PE1': 6.63,
            'PE2': 6.26, 'PRU': 29.345, 'RP4': 7.15, 'SAL': 9.317, 'SAA': 5,
            'SLO': 4.06, 'CAT': 8.3, 'MIR': 40.81, 'PAC': 17.84, 'PAM': 16,
            'SPK': 4, 'TCO': 50, 'TRO': 1.68, 'EALG3': 0.98,
            'CNOG1': 73.66, 'CNOG2': 159, 'CNOG3': 76.62, 'CNOG4': 77.56,
            'GTUG1': 208, 'GTUG2': 208, 'GTUG3': 213
        }
        for k, v in defaults.items():
            if k not in dict_potencias:
                dict_potencias[k] = v

        # ── Normalización unicode (maneja acentos) ────────────────────────────
        def _norm(t):
            return unicodedata.normalize("NFKD", str(t)).encode("ascii", "ignore").decode().upper().strip()

        # ── Detección de columnas por contenido ───────────────────────────────
        # Se revisan las primeras 15 filas no vacías para identificar qué columna
        # tiene HH:MM (hora) y cuál tiene DD/MM/YYYY (fecha).
        col_hora_idx  = None
        col_fecha_idx = None
        muestra = df_bitacora.head(15)
        for j in range(len(df_bitacora.columns)):
            col_vals = muestra.iloc[:, j].fillna("").astype(str).str.strip()
            tiene_hora  = col_vals.str.match(r'^\d{1,2}:\d{2}(:\d{2})?$').any()
            tiene_fecha = col_vals.str.match(r'^\d{2}/\d{2}/\d{4}$').any()
            if tiene_hora  and col_hora_idx  is None: col_hora_idx  = j
            if tiene_fecha and col_fecha_idx is None: col_fecha_idx = j

        # ── Extracción de fecha y hora ────────────────────────────────────────
        def _extraer_fecha_hora(texto_norm, fila_raw):
            # Intento 1: fecha+hora juntos en el texto (DD/MM/YYYY HH:MM)
            m = re.search(r'(\d{2}/\d{2}/\d{4})\s*(\d{2}:\d{2})', texto_norm)
            if m:
                return f"{m.group(1)} {m.group(2)}"
            # Intento 2: columnas dedicadas
            fecha = hora = ""
            if col_fecha_idx is not None:
                v = str(fila_raw.iloc[col_fecha_idx]).strip()
                if re.match(r'\d{2}/\d{2}/\d{4}', v):
                    fecha = v[:10]
            if col_hora_idx is not None:
                v = str(fila_raw.iloc[col_hora_idx]).strip()
                m2 = re.match(r'(\d{1,2}:\d{2})', v)
                if m2:
                    hora = m2.group(1)
            if fecha and hora:
                return f"{fecha} {hora}"
            if fecha:
                return fecha
            # Intento 3: solo fecha en el texto
            m3 = re.search(r'(\d{2}/\d{2}/\d{4})', texto_norm)
            return m3.group(1) if m3 else "Ver Bitácora"

        # ── Extracción de número de libranza ─────────────────────────────────
        # Patrón preciso (igual al módulo RTR): ETESA-NNN-YYYY o CNE-NNN-YYYY
        _RE_LIBRANZA = re.compile(
            r'\b(?:ETESA|CNE)\s*-\s*\d{1,4}\s*-\s*\d{4}\b', re.IGNORECASE
        )
        def _extraer_libranza(texto):
            m = _RE_LIBRANZA.search(texto)
            if m:
                return re.sub(r'\s*-\s*', '-', m.group(0)).upper()
            # Fallback genérico: XXX-NNN-YYYY
            m2 = re.search(r'\b[A-Z]{2,8}-\d{2,4}-\d{4}\b', texto)
            return m2.group(0) if m2 else "S/L"

        # ── Extracción de causa ───────────────────────────────────────────────
        def _extraer_causa(texto, unidad, libranza):
            # Buscar "POR <motivo>" (patrón más común en bitácoras panameñas)
            m_por = re.search(r'\bPOR\b\s*(.{5,120}?)(?=\s{2,}|\s*[A-Z]{5,}\s*:|\s*\d{2}/|\s*$)', texto)
            if m_por:
                causa = m_por.group(1).strip(" :-.,")
            else:
                # Fallback: texto después de la primera aparición del código de unidad
                idx = texto.find(unidad)
                causa = texto[idx + len(unidad):].strip(" :-") if idx != -1 else "SIN DETALLE"
            # Limpiar número de libranza de la causa
            if libranza != "S/L":
                causa = causa.replace(libranza, "").strip(" :-")
            # Limpiar timestamps que hayan quedado dentro de la causa
            causa = re.sub(r'\d{2}/\d{2}/\d{4}\s*\d{2}:\d{2}', '', causa).strip()
            return causa[:150] if causa else "SIN DETALLE"

        # ── Keywords de entrada (retorno al servicio) ─────────────────────────
        KW_ENTRADA = [
            "EN LINEA", "SINCRONIZ", "NORMALIZ",
            "VIENE DE SALIDA FORZADA DISPONIBLE",
            "EN LINEA VIENE DE DISPARO DISPONIBLE",
            "EN SERVICIO", "RETORNO AL SERVICIO", "RETORNA AL",
            "REGRESA AL SERVICIO", "DISPONIBLE Y EN LINEA",
            "FINALIZO LIBRANZA",   # alias frecuente
        ]
        # Keywords de salida (fuera de servicio)
        KW_SALIDA = [
            "DISPARO", "SALIDA FORZADA", "DECLARADA INDISPONIBLE",
            "DECLARADO INDISPONIBLE", "FUERA DE SERVICIO",
            "SALIDA DE EMERGENCIA", "FORZADA FUERA",
            "INICIO LIBRANZA",     # alias frecuente
            "INICIO(A) LIBRANZA",
        ]

        # ── Procesamiento fila a fila ─────────────────────────────────────────
        # salidas_unidades: dict de LISTAS para capturar múltiples eventos por unidad
        salidas_unidades: dict = {}    # { unidad: [ {datos_salida}, ... ] }
        entradas_unidades: set = set() # unidades que regresaron en esta bitácora

        for i in range(len(df_bitacora)):
            fila_raw  = df_bitacora.iloc[i].fillna("").astype(str)
            texto_raw = " ".join(fila_raw)
            texto     = _norm(texto_raw)

            # ── Filtro de tema principal ──────────────────────────────────────
            # Solo descartar si el tema es línea/barra/interruptor/transformador
            # y NO aparece ningún código de unidad generadora en el texto.
            es_tema_red = any(r in texto for r in [
                "LINEA 230", "LINEA 115", "INTERRUPTOR", "BARRA ", "TRANSFORMADOR"
            ])
            contiene_unidad_gen = bool(re.search(r'\b[A-Z][A-Z0-9]{1,4}G\d+\b', texto))
            if es_tema_red and not contiene_unidad_gen:
                continue

            # Filtro de estados parciales (no son indisponibilidades completas)
            if any(bas in texto for bas in ["LIMITADO", "LIMITA", "RESTRICCION", "PRUEBA", "POR SEGURIDAD"]):
                continue

            match = re.search(r'\b([A-Z][A-Z0-9]{1,4})G(\d+)\b', texto)
            if not match:
                continue

            unidad  = match.group(0)         # ej. "GTUG1"
            prefijo = match.group(1)         # ej. "GTU"
            potencia_val = dict_potencias.get(unidad, dict_potencias.get(prefijo, 0.0))

            # ── Clasificación del evento ──────────────────────────────────────
            es_entrada = (
                any(e in texto for e in KW_ENTRADA) or
                ("DISPONIBLE" in texto
                 and "NO DISPONIBLE" not in texto
                 and "INDISPONIBLE" not in texto) or
                ("FINALIZO" in texto and "LIBRANZA" in texto)
            )
            es_salida = (
                any(s in texto for s in KW_SALIDA) or
                ("INDISPONIBLE" in texto and "NO INDISPONIBLE" not in texto) or
                ("INICIO" in texto and "LIBRANZA" in texto)
            )

            # Cuando ambas flags están activas (ej: "FINALIZO LIBRANZA … INDISPONIBLE"),
            # la entrada tiene prioridad (la unidad quedó disponible).
            if es_entrada:
                entradas_unidades.add(unidad)
                continue

            if es_salida:
                libranza_id = _extraer_libranza(texto)
                fecha_hora  = _extraer_fecha_hora(texto, fila_raw)
                causa       = _extraer_causa(texto, unidad, libranza_id)

                evento = {
                    "salida":    fecha_hora,
                    "causa":     causa,
                    "libranza":  libranza_id,
                    "potencia":  potencia_val,
                    "repetido":  unidad in salidas_unidades,  # True si ya había una salida previa
                }
                salidas_unidades.setdefault(unidad, []).append(evento)

        # ── Construcción del resultado ────────────────────────────────────────
        filas_detectadas = []
        with sqlite3.connect('unidades_reporte.db') as conn:
            # Actualizar fecha de entrada para unidades que regresaron
            hoy_str = datetime.now().strftime("%d/%m/%Y %H:%M")
            for unidad in entradas_unidades:
                conn.execute("""
                    UPDATE reporte_unidades
                    SET entrada = ?
                    WHERE unidad = ? AND (entrada = 'SIN FECHA DE SINCRONIZACIÓN' OR entrada = 'PENDIENTE')
                """, (hoy_str, unidad))

            # Volcar todas las salidas detectadas (incluyendo múltiples por unidad).
            # Si la unidad también tuvo ENTRADA en esta misma bitácora (ciclo completo),
            # no se agrega como candidato nuevo (el ciclo ya cerró en la propia bitácora).
            for unidad, lista_salidas in salidas_unidades.items():
                if unidad in entradas_unidades:
                    continue  # ciclo completo en la misma bitácora
                for data in lista_salidas:
                    obs = "⚠️ Repetido en bitácora" if data["repetido"] else ""
                    filas_detectadas.append({
                        "UNIDAD":               unidad,
                        "FECHA DE SALIDA":      data["salida"],
                        "CAUSA":                data["causa"],
                        "LIBRANZA":             data["libranza"],
                        "FECHA DE ENTRADA":     "SIN FECHA DE SINCRONIZACIÓN",
                        "POTENCIA INDISPONIBLE": data["potencia"],
                        "OBSERVACIONES":        obs,
                        "SELECCIONAR":          False,
                    })
            conn.commit()

        columnas_ordenadas = [
            "SELECCIONAR", "UNIDAD", "FECHA DE SALIDA", "CAUSA",
            "LIBRANZA", "FECHA DE ENTRADA", "POTENCIA INDISPONIBLE", "OBSERVACIONES"
        ]
        df_resultado = pd.DataFrame(filas_detectadas)
        if not df_resultado.empty:
            df_resultado = df_resultado[columnas_ordenadas]
            df_resultado['fecha_temp'] = pd.to_datetime(
                df_resultado['FECHA DE SALIDA'], errors='coerce', dayfirst=True
            )
            df_resultado = df_resultado.sort_values(
                by='fecha_temp', ascending=False, na_position='last'
            ).drop(columns=['fecha_temp']).reset_index(drop=True)

        return len(df_resultado), df_resultado

    except Exception as e:
        st.error(f"Error en el filtrado: {e}")
        return 0, pd.DataFrame(columns=[
            "SELECCIONAR", "UNIDAD", "FECHA DE SALIDA", "CAUSA",
            "LIBRANZA", "FECHA DE ENTRADA", "POTENCIA INDISPONIBLE", "OBSERVACIONES"
        ])
    

# --- FUNCIONES AUXILIARES RTR MEJORADAS ---

def crear_fila_rtr(info, hora_fin, col_ni, col_nj, col_ckt):
    db = info['row_info']
    h_ini = info['inicio']
    duracion = calcular_duracion(h_ini, hora_fin)
    
    return {
        'L': '', 
        'Tipo de Elemento': 'LINEA',
        'Tipo de indisponibilidad': info['tipo'],
        'Tipo de Causa': info.get('causa', 'DESCONOCIDA'),
        'Nodo I': str(db.get(col_ni, '')).replace('.0', '').strip(),
        'Nodo J': str(db.get(col_nj, '')).replace('.0', '').strip(),
        'Nodo K': '', 
        'CKT': str(db.get(col_ckt, '')).replace('.0', '').strip(),
        'F. Indisponibilidad (dd-mmm-aaaa)': info['fecha'],
        'Hora Inicio CA+1': h_ini,
        'Hora Fin CA+1': hora_fin,
        'Hora de duración': duracion,           
        'Energía no servida (MWh)': '',   
        'Descripción': limpiar_y_formatear_desc(info['desc'])
    }

def es_evento_rtr_valido(texto, equipo):
    """
    Valida que el texto corresponde a un evento REAL de línea RTR.

    Patrón válido requerido:
        ACTION_KEYWORD  [SUBESTACION 1-3 palabras]  LINEA  [230|115|66]-ID

    Ejemplos válidos:
        LIBRANZA INICIO(A) CHORRERA LINEA 230-3A
        ABIERTO(A) EL HIGO LINEA 230-4B
        CERRADO(A) PANAMA LINEA 230-3A
        SOLMANT207: ABIERTO(A) CHORRERA LINEA 230-3B LIBRANZA ETESA...

    Ejemplos INVÁLIDOS (mención incidental de la línea):
        LIBRANZA FINALIZO(A) MIRAFLORES UNIDAD DE GENERACION MIRG... LINEA 115-5
        LIBRANZA INICIO(A) PROGRESO INTERRUPTOR 23B22 ETESA-294... LINEA 230-10
    """
    equipo_re = equipo.replace('-', r'\s*-\s*')

    # 1. El equipo debe aparecer precedido de la palabra "LINEA"
    patron_linea = rf'\bLINEA\s+{equipo_re}\b'
    match_linea = re.search(patron_linea, texto)
    if not match_linea:
        return False  # La línea no aparece como "LINEA 230-XX"

    # 2. Revisar el texto ANTES de "LINEA 230-XX":
    #    si hay palabras que indican que se trata de otro elemento (unidad, interruptor,
    #    transformador, barra), la mención de la línea es solo contextual → descartar.
    texto_previo = texto[:match_linea.start()]
    exclusiones = ["UNIDAD DE GENERACION", "INTERRUPTOR", "TRANSFORMADOR", "BARRA "]
    if any(exc in texto_previo for exc in exclusiones):
        return False

    # 3. Debe existir alguna acción relevante de línea ANTES de "LINEA 230-XX"
    acciones = [
        "LIBRANZA INICIO", "INICIO(A)", "LIBRANZA FINALIZO", "FINALIZO(A)",
        "ABIERTO(A)", "DISPARO", "DESENERGIZADO(A)",
        "CERRADO(A)", "CERRADA", "ENERGIZADO(A)", "NORMALIZADO(A)"
    ]
    if not any(acc in texto_previo for acc in acciones):
        return False

    return True


# ── Helpers para lógica RTR mejorada ────────────────────────────────────────

_RE_NUM_LIBRANZA = re.compile(r'\bETESA\s*-\s*\d+\s*-\s*\d+\b', re.IGNORECASE)

def _extraer_num_libranza(texto: str) -> str:
    """Extrae el número de libranza (p.ej. ETESA-360-2026) del texto.
    Devuelve cadena vacía si no lo encuentra."""
    m = _RE_NUM_LIBRANZA.search(texto)
    if not m:
        return ""
    return re.sub(r'\s*-\s*', '-', m.group(0)).upper()


def _es_portico(texto: str) -> bool:
    """Indica si la descripción hace referencia a un pórtico de línea.
    Tolera acentos (PORTICO / PÓRTICO)."""
    t = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode().upper()
    return "PORTICO" in t


def _hora_a_min(h: str) -> int:
    """Convierte 'HH:MM' a minutos desde 00:00. -1 si no es válido."""
    try:
        hh, mm = h.split(":")
        return int(hh) * 60 + int(mm)
    except Exception:
        return -1


def procesar_logica_rtr(df_bit, df_eq):
    # Identificar columnas
    col_id = next((c for c in df_eq.columns if "IDENTIFICACION" in str(c).upper()), None)
    col_ni = next((c for c in df_eq.columns if "FROM" in str(c).upper() or "NODO I" in str(c).upper()), None)
    col_nj = next((c for c in df_eq.columns if "TO" in str(c).upper() or "NODO J" in str(c).upper()), None)
    col_ckt = next((c for c in df_eq.columns if "CKT" in str(c).upper() or "EOR" in str(c).upper()), None)
    
    if col_id is None:
        return pd.DataFrame()
 
    validos = [str(n).strip().upper() for n in df_eq[col_id].unique() if str(n).startswith(('230', '115', '66'))]
    
    # pendientes ahora es dict de LISTAS para soportar múltiples ciclos por equipo
    # Formato: { 'equipo': [ {info_1}, {info_2}, ... ] }
    pendientes = {}
    eventos_finales = []
 
    # Evitar duplicados exactos de cierre huérfano: { 'equipo': {'17:00', '18:00'} }
    ultimos_cierres = {}
 
    # Equipos cerrados físicamente (cerrado/normalizado) cuyo fin_libranza posterior debe descartarse
    recien_cerrados_fisico = set()
 
    # ── Pre-parse: convertir filas a eventos estructurados para permitir look-ahead ──
    # Esto nos permite detectar patrones como "FINALIZO seguido de INICIO ≤ 2 min"
    # sin perder la información de las filas siguientes.
    def _parsear_fila(texto: str):
        """Detecta equipo, hora y tipo de evento en una fila. None si no es evento RTR."""
        equipo = None
        for e in validos:
            patron = r'\b' + e.replace('-', r'\s*-\s*') + r'\b'
            if re.search(patron, texto):
                equipo = e
                break
        if not equipo:
            return None

        # Si el primer equipo no pasa la validación, buscar alternativa
        # (caso: "LINEA 230-4B ... LT 230-3B Y 230-4B" donde 230-3B es incidental)
        if not es_evento_rtr_valido(texto, equipo):
            alt = None
            for e in validos:
                if e == equipo:
                    continue
                patron = r'\b' + e.replace('-', r'\s*-\s*') + r'\b'
                if re.search(patron, texto) and es_evento_rtr_valido(texto, e):
                    alt = e
                    break
            if alt:
                equipo = alt
            else:
                return None

        m = re.search(r'(\d{2}:\d{2})', texto)
        hora = m.group(1) if m else "00:00"

        fecha_m = re.search(r'(\d{2}/\d{2}/\d{4})', texto)
        fecha = fecha_m.group(1) if fecha_m else datetime.now().strftime('%d/%m/%Y')

        # Si el FINALIZO indica explícitamente que la línea sigue abierta (cambio
        # administrativo de número de libranza sin cierre físico), no se debe
        # interpretar como un cierre de indisponibilidad.
        _linea_sigue_abierta = any(p in texto for p in [
            "LINEA CONTINUA ABIERTA", "LÍNEA CONTINUA ABIERTA",
            "CONTINUA ABIERTA", "LÍNEA ABIERTA", "LINEA ABIERTA",
            "SIGUE ABIERTA", "PERMANECE ABIERTA",
        ])

        return {
            'equipo': equipo,
            'hora': hora,
            'fecha': fecha,
            'texto': texto,
            'es_cierre_fisico': ("INICIO(A)" not in texto or "LIBRANZA" not in texto)
                and any(p in texto for p in ["CERRADO(A)", "CERRADA", "ENERGIZADO(A)", "NORMALIZADO(A)"]),
            # Un FINALIZO/INICIO de pórtico no cierra ni abre una indisponibilidad de línea.
            # Se filtra aquí para que no llegue a ninguna rama de la lógica de cierre/apertura.
            # Tampoco se marca como fin si la anotación dice que la línea continúa abierta
            # (ej: "LINEA CONTINUA ABIERTA") — es solo un cambio administrativo de libranza.
            'es_fin_libranza': "FINALIZO" in texto and "LIBRANZA" in texto
                and not _es_portico(texto) and not _linea_sigue_abierta,
            'es_disparo': "DISPARO" in texto,
            'es_apertura_fisica': any(p in texto for p in ["ABIERTO(A)", "DESENERGIZADO(A)"]),
            'es_inicio_libranza': "INICIO(A)" in texto and "LIBRANZA" in texto,
            'num_libranza': _extraer_num_libranza(texto),
            'es_portico': _es_portico(texto),
        }

    eventos = []
    for _, row in df_bit.iterrows():
        texto = " ".join(str(x) for x in row.values if pd.notnull(x)).upper()
        eventos.append(_parsear_fila(texto))

    # ── Pre-procesamiento: descartar primer fin_libranza cuando el segundo dice LINEA CERRADA ──
    # Si una línea tiene dos fin_libranza y el segundo contiene "LINEA CERRADA",
    # el primero es incorrecto: el equipo no puede cerrarse dos veces sin un inicio en medio.
    # El segundo fin_libranza (con LINEA CERRADA) es el cierre real.
    _indices_omitir = set()
    _equipos_en_eventos = {ev['equipo'] for ev in eventos if ev}
    for _eq in _equipos_en_eventos:
        _idx_fin = [i for i, ev in enumerate(eventos)
                    if ev and ev['equipo'] == _eq and ev['es_fin_libranza']]
        if len(_idx_fin) < 2:
            continue
        for _k in range(len(_idx_fin) - 1):
            _i1, _i2 = _idx_fin[_k], _idx_fin[_k + 1]
            # Si el segundo fin_libranza dice LINEA CERRADA → el primero es incorrecto
            _ev2 = eventos[_i2]
            if _ev2 and any(p in _ev2['texto'] for p in
                            ['LINEA CERRADA', 'CERRADO(A)', 'ENERGIZADO(A)', 'NORMALIZADO(A)']):
                _indices_omitir.add(_i1)  # descartar el primer fin_libranza

    # ── Loop principal sobre eventos pre-parseados ─────────────────────────
    for i, ev in enumerate(eventos):
        if ev is None:
            continue
        if i in _indices_omitir:
            continue  # primer fin_libranza descartado (hay cierre físico antes del segundo)

        equipo = ev['equipo']
        hora = ev['hora']
        fecha = ev['fecha']
        texto = ev['texto']
        es_cierre_fisico = ev['es_cierre_fisico']
        es_fin_libranza = ev['es_fin_libranza']
        es_disparo = ev['es_disparo']
        es_apertura_fisica = ev['es_apertura_fisica']
        es_inicio_libranza = ev['es_inicio_libranza']

        # --- LÓGICA DE CIERRE ---
        if es_cierre_fisico or es_fin_libranza:

            lista = pendientes.get(equipo, [])

            # fin_libranza se descarta si ya hubo un cierre físico (el cerrado físico es
            # más preciso; fin_libranza es solo el trámite administrativo posterior).
            if es_fin_libranza and equipo in recien_cerrados_fisico:
                continue

            # ── NUEVAS REGLAS PARA FIN DE LIBRANZA ────────────────────────
            if es_fin_libranza and lista:
                pendiente_actual = lista[0]

                num_pend = pendiente_actual.get('num_libranza', '')
                num_fin  = ev['num_libranza']

                # REGLA 3: detectar cambio de libranza (FINALIZO → INICIO en ≤ 2 min
                # para el mismo equipo). La línea sigue abierta, solo cambió el número.
                # En lugar de bloquear todos los FINALIZOs futuros con un flag, actualizamos
                # num_libranza al de la nueva libranza. Así el próximo FINALIZO (cuando
                # la línea cierre de verdad) será reconocido normalmente por la Regla 2.
                min_fin = _hora_a_min(hora)
                if min_fin >= 0:
                    nuevo_num = None
                    for j in range(i + 1, len(eventos)):
                        ev2 = eventos[j]
                        if ev2 is None:
                            continue
                        min_j = _hora_a_min(ev2['hora'])
                        if min_j < 0:
                            continue
                        if (min_j - min_fin) > 5:
                            break  # demasiado adelante en tiempo → no es cambio inmediato
                        if ev2['equipo'] != equipo:
                            continue
                        if not ev2['es_inicio_libranza']:
                            continue
                        if 0 <= (min_j - min_fin) <= 2:
                            nuevo_num = ev2['num_libranza']  # capturar el num de la nueva libranza
                            break
                    if nuevo_num is not None:
                        # Cambio de libranza: actualizar el pendiente con el nuevo número.
                        # El próximo FINALIZO de la nueva libranza cerrará normalmente.
                        pendiente_actual['num_libranza'] = nuevo_num
                        continue

                # REGLA 2: matching por número de libranza. Si el pendiente tiene un num
                # conocido y este FINALIZO es de otra libranza, saltar — PERO solo si
                # hay otro FINALIZO adelante (para no dejar el evento abierto si el
                # match nunca llega).
                if num_pend and num_fin and num_pend != num_fin:
                    hay_otro_finalizo = False
                    for j in range(i + 1, len(eventos)):
                        ev2 = eventos[j]
                        if ev2 is None or ev2['equipo'] != equipo:
                            continue
                        if ev2['es_cierre_fisico']:
                            break  # habrá un cierre físico → no es ambiguo
                        if ev2['es_fin_libranza']:
                            hay_otro_finalizo = True
                            break
                    if hay_otro_finalizo:
                        continue
                    # Si no hay otro FINALIZO por venir, caer al cierre original como fallback.

                # REGLA: si viene un cierre físico del mismo equipo dentro de los próximos
                # 10 minutos, ceder el cierre a ese evento (el cierre físico es más preciso).
                min_fin_actual = _hora_a_min(hora)
                if min_fin_actual >= 0:
                    hay_cierre_fisico_pronto = False
                    for j in range(i + 1, len(eventos)):
                        ev2 = eventos[j]
                        if ev2 is None:
                            continue
                        min_j = _hora_a_min(ev2['hora'])
                        if min_j < 0:
                            continue
                        if (min_j - min_fin_actual) > 10:
                            break
                        if ev2['equipo'] == equipo and ev2['es_cierre_fisico']:
                            hay_cierre_fisico_pronto = True
                            break
                    if hay_cierre_fisico_pronto:
                        continue  # el cierre físico próximo cerrará con su hora real

            # ── LÓGICA DE CIERRE ORIGINAL ─────────────────────────────────
            if lista:
                info = lista.pop(0)
                if not lista:
                    pendientes.pop(equipo, None)
                eventos_finales.append(crear_fila_rtr(info, hora, col_ni, col_nj, col_ckt))
                ultimos_cierres.setdefault(equipo, set()).add(hora)
                if es_cierre_fisico:
                    # Marcar que este equipo ya tuvo cierre físico,
                    # para descartar el fin_libranza que pueda venir después
                    recien_cerrados_fisico.add(equipo)

            else:
                # Cierre huérfano: ignorar si ya se procesó un cierre a la misma hora
                if hora in ultimos_cierres.get(equipo, set()):
                    continue
                # Cierre huérfano: ignorar si el equipo ya tuvo un cierre físico
                # (cubre el segundo extremo cuando cierra a hora diferente, ej: 18:35 y 18:38)
                if equipo in recien_cerrados_fisico:
                    if es_cierre_fisico:
                        ultimos_cierres.setdefault(equipo, set()).add(hora)
                    continue

                # ── NUEVA REGLA: huérfano con cambio de libranza inmediato ──
                # Si es un FINALIZO huérfano (sin pendiente) y dentro de ≤ 2 min
                # viene un INICIO del mismo equipo, la línea venía abierta del día
                # anterior y solo está cambiando el número de libranza.
                # Crear pendiente virtual desde 00:00 con el num de la nueva libranza
                # para que el próximo FINALIZO lo cierre correctamente.
                if es_fin_libranza:
                    min_fin = _hora_a_min(hora)
                    if min_fin >= 0:
                        nuevo_num = None
                        for j in range(i + 1, len(eventos)):
                            ev2 = eventos[j]
                            if ev2 is None:
                                continue
                            min_j = _hora_a_min(ev2['hora'])
                            if min_j < 0:
                                continue
                            if (min_j - min_fin) > 5:
                                break
                            if ev2['equipo'] != equipo:
                                continue
                            if not ev2['es_inicio_libranza']:
                                continue
                            if 0 <= (min_j - min_fin) <= 2:
                                nuevo_num = ev2['num_libranza']
                                break
                        if nuevo_num is not None:
                            _df_eq_match = df_eq[df_eq[col_id].astype(str).str.strip().str.upper() == equipo]
                            if _df_eq_match.empty:
                                continue
                            eq_row = _df_eq_match.iloc[0]
                            pendientes.setdefault(equipo, []).append({
                                'inicio': '00:00',
                                'fecha': fecha,
                                'desc': texto + " (LÍNEA ABIERTA DEL DÍA ANTERIOR)",
                                'tipo': 'PROGRAMADA',
                                'causa': 'MANTENIMIENTO',
                                'row_info': eq_row,
                                'prioridad': 1,
                                'fin_libranza': None,
                                'num_libranza': nuevo_num,  # ← cierra con la nueva libranza
                                'es_portico': False,
                            })
                            continue  # no crear huérfano cerrado

                # 2do FINALIZO de portico multi-extremo: ya cerrado a la misma hora
                if hora in ultimos_cierres.get(equipo, set()):
                    continue
                _df_eq_huerfano = df_eq[df_eq[col_id].astype(str).str.strip().str.upper() == equipo]
                if _df_eq_huerfano.empty:
                    continue
                eq_info = _df_eq_huerfano.iloc[0]
                info_huerfana = {
                    'inicio': '00:00',
                    'fecha': fecha,
                    'desc': texto + " (APERTURA NO REGISTRADA EN BITÁCORA)",
                    'tipo': 'PROGRAMADA',
                    'causa': 'MANTENIMIENTO',
                    'row_info': eq_info
                }
                eventos_finales.append(crear_fila_rtr(info_huerfana, hora, col_ni, col_nj, col_ckt))
                ultimos_cierres.setdefault(equipo, set()).add(hora)
                if es_cierre_fisico:
                    recien_cerrados_fisico.add(equipo)

        # --- LÓGICA DE APERTURA ---
        elif es_disparo or es_apertura_fisica or es_inicio_libranza:

            if es_disparo:
                tipo = "NO PROGRAMADA"
                causa = "DISPARO"
                prioridad = 3
            elif es_apertura_fisica:
                tipo = "PROGRAMADA"
                causa = "MANTENIMIENTO"
                prioridad = 2
            else:  # Inicio libranza
                tipo = "PROGRAMADA"
                causa = "MANTENIMIENTO"
                prioridad = 1

            lista = pendientes.setdefault(equipo, [])

            # Al registrar una nueva apertura, limpiar el marcador de cierre físico
            # (el equipo puede volver a abrirse después de haberse cerrado)
            recien_cerrados_fisico.discard(equipo)

            # ── NUEVA REGLA 1: para INICIO de libranza con pendiente existente ──
            if es_inicio_libranza and lista:
                pendiente = lista[0]

                # Solo consideramos swap si el pendiente también es un inicio de
                # libranza (no queremos sobreescribir un DISPARO o APERTURA_FISICA).
                if pendiente.get('prioridad', 0) == 1:
                    hora_pend = _hora_a_min(pendiente.get('inicio', ''))
                    hora_new  = _hora_a_min(hora)

                    # Caso 1a: misma hora exacta y el pendiente es PÓRTICO pero el
                    # nuevo NO lo es → reemplazar (el pórtico inicia después aunque
                    # tenga la misma hora registrada). Actualiza num_libranza también.
                    if (hora_pend >= 0 and hora_pend == hora_new
                            and pendiente.get('es_portico')
                            and not ev['es_portico']):
                        pendiente['inicio'] = hora
                        pendiente['desc'] = texto
                        pendiente['num_libranza'] = ev['num_libranza']
                        pendiente['es_portico'] = False
                        continue

                # Regla original: ya hay apertura para este equipo → descartar
                # (la línea se abre una sola vez por ciclo; los inicios del segundo
                # extremo o de pórticos son redundantes).
                continue

            # Buscar si existe un evento pendiente de MENOR prioridad (ej. inicio_libranza)
            # para hacer upgrade en vez de crear nuevo ciclo.
            # Regla: abierto/disparo reemplaza a inicio_libranza del mismo ciclo.
            entrada_a_actualizar = next(
                (p for p in lista if p['prioridad'] < prioridad), None
            )

            if entrada_a_actualizar:
                # Upgrade: reemplazamos hora y datos con el evento de mayor prioridad
                entrada_a_actualizar['inicio'] = hora
                entrada_a_actualizar['desc'] = texto
                entrada_a_actualizar['prioridad'] = prioridad
                entrada_a_actualizar['tipo'] = tipo
                entrada_a_actualizar['causa'] = causa

            elif any(p['inicio'] == hora for p in lista):
                # Duplicado exacto (misma hora, misma prioridad): solo actualizar si prioridad mayor
                for p in lista:
                    if p['inicio'] == hora and p['prioridad'] < prioridad:
                        p['desc'] = texto
                        p['prioridad'] = prioridad
                        p['tipo'] = tipo
                        p['causa'] = causa

            else:
                # Nueva apertura a hora diferente sin entradas de menor prioridad → nuevo ciclo
                _df_eq_apertura = df_eq[df_eq[col_id].astype(str).str.strip().str.upper() == equipo]
                if _df_eq_apertura.empty:
                    continue
                eq_row = _df_eq_apertura.iloc[0]
                lista.append({
                    'inicio': hora,
                    'fecha': fecha,
                    'desc': texto,
                    'tipo': tipo,
                    'causa': causa,
                    'row_info': eq_row,
                    'prioridad': prioridad,
                    'fin_libranza': None,
                    'num_libranza': ev['num_libranza'],
                    'es_portico': ev['es_portico'],
                })

    # CERRAR EVENTOS PENDIENTES AL FINAL (sin cierre registrado en bitácora)
    for equipo, lista in pendientes.items():
        for info in lista:
            hora_fin = info.get('fin_libranza') if info.get('fin_libranza') else "23:59"
            eventos_finales.append(crear_fila_rtr(info, hora_fin, col_ni, col_nj, col_ckt))
 
    df_result = pd.DataFrame(eventos_finales)
    
    for c in COLUMNAS_RTR:
        if c not in df_result.columns:
            df_result[c] = ''
            
    return df_result[COLUMNAS_RTR]

# --- VISTA PRINCIPAL RTR (ACTUALIZADA) ---

def vista_rtr(archivo_bitacora):
    st.markdown("<style>#MainMenu, footer, header {visibility: hidden;} .stDeployButton {display:none;}</style>", unsafe_allow_html=True)
    st.title("⚡ INDISPONIBILIDADES RTR")
    
    ahora = datetime.now() - timedelta(days=1)
    fecha_reporte_str = ahora.strftime('%d/%m/%Y')

    # --- SECCIÓN: GESTIÓN DE BASE DE DATOS ---
    with st.expander("➕ Configuración de Seguimiento (Base de Datos)", expanded=True):
        c1, c2 = st.columns([1, 2])
        n_linea = c1.text_input("Nombre de la Línea (ej. 230-54B)", key="rtr_n_linea")
        n_desc = c2.text_area("Descripción", key="rtr_n_desc")
        
        if st.button("💾 Registrar / Actualizar Base de Datos", help="Agrega de forma manual una nueva línea"):
            if n_linea and n_desc:
                conn = sqlite3.connect('rtr_seguimiento.db')
                conn.execute("CREATE TABLE IF NOT EXISTS lineas_abiertas (linea TEXT PRIMARY KEY, descripcion_manual TEXT)")
                conn.execute("INSERT OR REPLACE INTO lineas_abiertas (linea, descripcion_manual) VALUES (?,?)", 
                             (n_linea.strip().upper(), n_desc.strip()))
                conn.commit()
                conn.close()
                st.session_state.pop("rtr_file_id", None)  # fuerza reprocesamiento con la nueva línea
                st.rerun()

        conn = sqlite3.connect('rtr_seguimiento.db')
        try:
            df_actual = pd.read_sql_query("SELECT * FROM lineas_abiertas", conn)
            if not df_actual.empty:
                st.write("### Líneas en Seguimiento:")
                for _, row in df_actual.iterrows():
                    col_info, col_borrar = st.columns([0.8, 0.2])
                    col_info.info(f"📍 **{row['linea']}** | {row['descripcion_manual']}")
                    if col_borrar.button("🗑️ Quitar", key=f"del_{row['linea']}"):
                        conn.execute("DELETE FROM lineas_abiertas WHERE linea = ?", (row['linea'],))
                        conn.commit()
                        conn.close()
                        st.session_state.pop("rtr_file_id", None)  # fuerza reprocesamiento sin la línea eliminada
                        st.rerun()
        except: pass
        finally: conn.close()

    # ── Cargar reporte RTR ya generado ──────────────────────────────────────────
    with st.expander("📂 Cargar indisponibilidades RTR ya generado", expanded=False):
        st.caption("Sube indisponibilidades RTR generado anteriormente.")
        uploaded_rtr_prev = st.file_uploader(
            "Reporte RTR (.xlsx)", type=["xlsx", "xls"], key="rtr_upload_previo"
        )
        if uploaded_rtr_prev is not None:
            try:
                _raw = uploaded_rtr_prev.read()
                _df_prev = pd.read_excel(
                    BytesIO(_raw),
                    sheet_name='Indisponibilidad Líneas y Trans',
                    skiprows=5, header=0,
                    dtype=str,          # leer TODO como texto para evitar conflictos de tipo
                )
                # La primera columna ('L') se escribe vacía en el Excel generado;
                # pandas la leerá con nombre vacío o 'Unnamed: 0'. Normalizarla.
                _first = str(_df_prev.columns[0])
                if _first == '' or _first.startswith('Unnamed'):
                    _df_prev.rename(columns={_df_prev.columns[0]: 'L'}, inplace=True)
                elif 'L' not in _df_prev.columns:
                    _df_prev.insert(0, 'L', '')

                # Normalizar todas las celdas: reemplazar 'nan'/'NaT'/'None' por ''
                _df_prev = _df_prev.fillna('').replace({'nan': '', 'NaT': '', 'None': ''})

                # Eliminar filas donde todas las columnas relevantes estén vacías
                _cols_datos = [c for c in _df_prev.columns if c != 'L']
                _df_prev = _df_prev[
                    _df_prev[_cols_datos].apply(lambda r: r.str.strip().ne('').any(), axis=1)
                ].reset_index(drop=True)

                # Asegurar que las columnas SelectboxColumn tengan valores válidos
                _sel_cols = {
                    'Tipo de indisponibilidad': ('PROGRAMADA', 'NO PROGRAMADA'),
                    'Tipo de Causa':            ('MANTENIMIENTO', 'DISPARO'),
                }
                for _scol, _opts in _sel_cols.items():
                    if _scol in _df_prev.columns:
                        _df_prev[_scol] = _df_prev[_scol].apply(
                            lambda v: v.strip().upper() if v.strip().upper() in _opts else _opts[0]
                        )

                # Guardar y activar el flag para evitar que la bitácora sobreescriba
                st.session_state["rtr_resultado"]         = _df_prev
                st.session_state["rtr_resultado_editado"] = _df_prev
                st.session_state["rtr_from_upload"]       = True
                st.session_state.pop("rtr_editor", None)
                st.success(f"✅ Reporte RTR cargado correctamente — {len(_df_prev)} indisponibilidad(es).")
            except Exception as _e:
                st.error(f"No se pudo leer el reporte RTR: {_e}")

        if st.session_state.get("rtr_from_upload"):
            
            if st.button("🔄 Volver a procesar desde bitácora", key="rtr_clear_upload"):
                st.session_state.pop("rtr_from_upload", None)
                st.session_state.pop("rtr_resultado", None)
                st.session_state.pop("rtr_resultado_editado", None)
                st.session_state.pop("rtr_file_id", None)
                st.rerun()

    if archivo_bitacora:
        # ── Botón para forzar reprocesamiento aunque sea el mismo archivo ──────
        # Útil cuando se actualiza el código pero el archivo no cambia.
        col_proc, _ = st.columns([1, 3])
        if col_proc.button("🔄 Forzar reprocesamiento", key="rtr_forzar"):
            st.session_state.pop("rtr_file_id", None)
            st.session_state.pop("rtr_from_upload", None)

        # Solo reprocesar si la bitácora es diferente a la última procesada
        # Esto evita que un rerun (ej: al eliminar filas) sobreescriba el session_state.
        # También se omite si el usuario cargó un reporte manualmente.
        file_id = f"{archivo_bitacora.name}_{archivo_bitacora.size}"
        if not st.session_state.get("rtr_from_upload") and st.session_state.get("rtr_file_id") != file_id:
            try:
                db_equipos = cargar_equipos()

                # Leer la bitácora de forma robusta: .xls necesita engine explícito
                # en versiones recientes de xlrd/openpyxl. Intentamos varias opciones.
                archivo_bitacora.seek(0)
                _bytes_bit = archivo_bitacora.read()
                archivo_bitacora.seek(0)
                df_bit = None
                for _engine in (None, "xlrd", "openpyxl"):
                    try:
                        _kw = {"engine": _engine} if _engine else {}
                        df_bit = pd.read_excel(BytesIO(_bytes_bit), header=None, **_kw)
                        break
                    except Exception:
                        continue
                if df_bit is None:
                    st.error("No se pudo leer la bitácora (.xls/.xlsx). Verifica que el archivo sea válido.")
                    st.stop()
                
                # Usamos la función de lógica mejorada
                resultado = procesar_logica_rtr(df_bit, db_equipos)

                # Cargar líneas de la DB para inyectarlas
                conn = sqlite3.connect('rtr_seguimiento.db')
                try: df_db = pd.read_sql_query("SELECT * FROM lineas_abiertas", conn)
                except: df_db = pd.DataFrame()
                conn.close()
                
                if not df_db.empty:
                    db_equipos.columns = [str(c).upper().strip() for c in db_equipos.columns]
                    col_id = next((c for c in db_equipos.columns if "IDENTIFICACION" in c), db_equipos.columns[0])
                    col_ni = next((c for c in db_equipos.columns if "FROM" in c or "NODO I" in c), None)
                    col_nj = next((c for c in db_equipos.columns if "TO" in c or "NODO J" in c), None)
                    col_ckt = next((c for c in db_equipos.columns if any(x in c for x in ["CKT", "CIRCUITO", "EOR"])), None)
                    
                    filas_db = []
                    for _, r_db in df_db.iterrows():
                        buscar = str(r_db['linea']).strip().upper()
                        eq = db_equipos[db_equipos[col_id].astype(str).str.strip().str.upper() == buscar]
                        
                        ni, nj, ckt_val = "", "", ""
                        if not eq.empty:
                            ni = str(eq.iloc[0][col_ni]).split('.')[0] if col_ni else ""
                            nj = str(eq.iloc[0][col_nj]).split('.')[0] if col_nj else ""
                            ckt_val = str(eq.iloc[0][col_ckt]).split('.')[0] if col_ckt else ""

                        desc = r_db['descripcion_manual']
                        desc = str(desc) if pd.notnull(desc) else ''
                        filas_db.append(dict(zip(COLUMNAS_RTR, [
                            ' ', 'LINEA', 'PROGRAMADA', 'MANTENIMIENTO',
                            ni, nj, '', ckt_val,
                            fecha_reporte_str,
                            '00:00', '23:59', '23:59', '',
                            desc
                        ])))
                    
                    if filas_db:
                        resultado = pd.concat([pd.DataFrame(filas_db), resultado], ignore_index=True)

                # Limpieza y ORDENAMIENTO POR HORA
                resultado = resultado.reindex(columns=COLUMNAS_RTR).fillna("")
                resultado = resultado.sort_values(by='Hora Inicio CA+1', ascending=True).reset_index(drop=True)
                # Guardar resultado y marcar el archivo como procesado
                st.session_state["rtr_resultado"] = resultado
                st.session_state["rtr_file_id"] = file_id

            except Exception as e:
                st.error(f"Error crítico: {e}")

    # ── Tabla editable — fuera del bloque de archivo para sobrevivir reruns ──
    if "rtr_resultado" in st.session_state:

        st.subheader("📝 Revisión de indisponibilidades RTR")

        # Construir DataFrame base para el editor.
        # NO se modifica rtr_resultado durante el render para evitar el conflicto
        # interno de Streamlit que hace que los cambios reboten.
        # Las ediciones viven en st.session_state["rtr_editor"] (manejado por Streamlit)
        # y se aplican explícitamente al guardar o al generar archivos.
        resultado = st.session_state["rtr_resultado"]
        df_con_sel = resultado.copy()
        df_con_sel.insert(0, "🗑️", False)

        column_config = {
            "🗑️": st.column_config.CheckboxColumn("Eliminar", default=False, width="small"),
            "L": None,
            "Tipo de Elemento": st.column_config.TextColumn("Elemento", width="small"),
            "Tipo de indisponibilidad": st.column_config.SelectboxColumn(
                "Indisponibilidad", options=["PROGRAMADA", "NO PROGRAMADA"], required=True, width="small",
            ),
            "Tipo de Causa": st.column_config.SelectboxColumn(
                "Causa", options=["MANTENIMIENTO", "DISPARO"], required=True, width="small",
            ),
            "Nodo K": st.column_config.TextColumn("Nodo K", width="small"),
            "F. Indisponibilidad (dd-mmm-aaaa)": st.column_config.TextColumn("Fecha", width="small"),
            "Hora Inicio CA+1": st.column_config.TextColumn("Inicio", width="small"),
            "Hora Fin CA+1": st.column_config.TextColumn("Fin", width="small"),
            "Hora de duración": st.column_config.TextColumn("Duración", width="small"),
            "Energía no servida (MWh)": st.column_config.TextColumn("ENS (MWh)", width="small"),
            "Descripción": st.column_config.TextColumn("Descripción", width="large"),
        }

        df_editado = st.data_editor(
            df_con_sel,
            hide_index=True,
            use_container_width=True,
            row_height=65,
            column_config=column_config,
            key="rtr_editor",
        )

        # df_editado ya contiene el estado completo actual (base + todos los cambios
        # del usuario en esta sesión). Es la fuente de verdad para exportar.
        # NO actualizamos rtr_resultado con df_editado para evitar el ciclo de rerender
        # de Streamlit, pero sí lo usamos directamente para Excel y XML.

        # Filas marcadas para eliminar (checkbox "🗑️" = True)
        _marcadas_mask = df_editado["🗑️"] == True
        _marcadas = set(df_editado.index[_marcadas_mask].tolist())
        n_sel = len(_marcadas)

        if st.button(f"🗑️ Eliminar seleccionadas ({n_sel})", disabled=(n_sel == 0), type="secondary", key="rtr_btn_eliminar"):
            resultado_limpio = (
                df_editado[~_marcadas_mask]
                .drop(columns=["🗑️"])
                .reset_index(drop=True)
            )
            st.session_state["rtr_resultado"] = resultado_limpio
            st.session_state["rtr_resultado_editado"] = resultado_limpio
            st.session_state.pop("rtr_editor", None)
            st.rerun()

        # resultado_final: df_editado sin la columna 🗑️ y sin las filas marcadas
        resultado_final = (
            df_editado[~_marcadas_mask]
            .drop(columns=["🗑️"])
            .reset_index(drop=True)
        )
        # Sincronizar para el XML del EOR (siempre refleja los cambios actuales)
        st.session_state["rtr_resultado_editado"] = resultado_final
        
        # --- FORMATO EXCEL ORIGINAL ---
        output = BytesIO()
        fecha_ayer = fecha_reporte_str
        
        with pd.ExcelWriter(output, engine='xlsxwriter', engine_kwargs={'options': {'nan_inf_to_errors': True}}) as writer:
            res_sheet = 'Indisponibilidad Líneas y Trans'
            resultado_final.to_excel(writer, index=False, sheet_name=res_sheet, startrow=5, header=False)
            wb, ws = writer.book, writer.sheets[res_sheet]
            
            fmt_titulo = wb.add_format({'bold': True, 'font_name': 'Arial', 'font_size': 11})
            fmt_amarillo = wb.add_format({'bold': True, 'bg_color': '#FFFF00', 'border': 1, 'align': 'center', 'font_name': 'Arial', 'font_size': 10, 'text_wrap': True})
            fmt_celda = wb.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter', 'font_name': 'Arial', 'font_size': 9, 'text_wrap': True})
            
            ws.hide_gridlines(2)
            
            ws.write(0, 1, "REGISTRO DE INDISPONIBILIDADES DE LOS ELEMENTOS DE TRANSMISION _PREDESPACHO REGIONAL", fmt_titulo)
            ws.write(1, 1, f"OS/OM : PANAMA {fecha_ayer}", fmt_titulo)
            ws.write(2, 1, "HORA   CA+1", fmt_titulo)
            
            for i, col in enumerate(resultado_final.columns):
                if i == 0: ws.write(5, i, "", None)
                else: ws.write(5, i, col, fmt_amarillo)
            
            anchos = [12, 15, 20, 18, 8, 8, 8, 7, 17, 10, 10, 10, 12, 80]
            for i, w in enumerate(anchos): ws.set_column(i, i, w)
            
            for r in range(6, 6 + len(resultado_final)):
                ws.set_row(r, 35)
                for c in range(len(resultado_final.columns)):
                    if c == 0: ws.write(r, c, "", None)
                    else: ws.write(r, c, resultado_final.iloc[r-6, c], fmt_celda)

        st.download_button("⏬ Descargar Reporte RTR", output.getvalue(), f"Indisponibilidades en tiempo real _PANAMA_{fecha_ayer.replace('/','_')}.xlsx")

def vista_unidades(archivo_bitacora_usuario):
    st.title("⚙️ INDISPONIBILIDAD DE UNIDADES")
    
    MESES_ES = {
        1: 'ene', 2: 'feb', 3: 'mar', 4: 'abr', 5: 'may', 6: 'jun',
        7: 'jul', 8: 'ago', 9: 'sep', 10: 'oct', 11: 'nov', 12: 'dic'
    }

    def formatear_fecha_excel(fecha_str):
        fecha_str = str(fecha_str).strip()

        palabras_especiales = ['SIN FECHA', 'PENDIENTE', 'VER BIT', 'NAN', 'NONE']
        if not fecha_str or any(p in fecha_str.upper() for p in palabras_especiales):
            return fecha_str

        # Quitar microsegundos
        fecha_str = re.sub(r'(\d{2}:\d{2}:\d{2})\.\d+', r'\1', fecha_str)
        # Fecha y hora pegadas sin espacio
        fecha_str = re.sub(r'(\d{2}/\d{2}/\d{4})(\d{2}:\d{2})', r'\1 \2', fecha_str)

        # ── Formato largo español almacenado: D-mmm-YYYY[ HH:MM] ─────────────
        # Ej: "30-abr-2026 22:59"  "8-may-2026 19:00"  "5-sep-2026"
        # Estos llegan de registros importados antes de la corrección del parser.
        _MESES_ES_INV_L = {
            'ene': 1, 'feb': 2, 'mar': 3, 'abr': 4, 'may': 5, 'jun': 6,
            'jul': 7, 'ago': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dic': 12
        }
        _ml = re.match(
            r'^(\d{1,2})-([a-z]{3})-(\d{4})(?:\s+(\d{1,2}):(\d{2}))?$',
            fecha_str, re.IGNORECASE
        )
        if _ml:
            mes = _MESES_ES_INV_L.get(_ml.group(2).lower())
            if mes:
                try:
                    dt = datetime(
                        int(_ml.group(3)), mes, int(_ml.group(1)),
                        int(_ml.group(4) or 0), int(_ml.group(5) or 0)
                    )
                    if _ml.group(4):
                        return f"{dt.day}-{MESES_ES[dt.month]}-{dt.year} {dt.hour:02d}:{dt.minute:02d}"
                    return f"{dt.day}-{MESES_ES[dt.month]}-{dt.year}"
                except ValueError:
                    pass

        formatos_entrada = [
            ('%d/%m/%Y %H:%M',    True),
            ('%d/%m/%Y %H:%M:%S', True),
            ('%d-%m-%Y %H:%M',    True),
            ('%d-%m-%Y %H:%M:%S', True),
            ('%Y-%m-%d %H:%M:%S', True),
            ('%Y-%m-%d %H:%M',    True),
            ('%d/%m/%Y',          False),
            ('%d-%m-%Y',          False),
            ('%Y-%m-%d',          False),
        ]

        for fmt, tiene_hora in formatos_entrada:
            try:
                dt = datetime.strptime(fecha_str, fmt)
                if tiene_hora:
                    return f"{dt.day}-{MESES_ES[dt.month]}-{dt.year} {dt.hour:02d}:{dt.minute:02d}"
                else:
                    return f"{dt.day}-{MESES_ES[dt.month]}-{dt.year}"
            except ValueError:
                continue

        # Último recurso: pandas
        try:
            dt = pd.to_datetime(fecha_str, dayfirst=True)
            if dt.hour == 0 and dt.minute == 0:
                return f"{dt.day}-{MESES_ES[dt.month]}-{dt.year}"
            return f"{dt.day}-{MESES_ES[dt.month]}-{dt.year} {dt.hour:02d}:{dt.minute:02d}"
        except Exception:
            return fecha_str

    hoy_str = datetime.now().strftime("%d/%m/%Y")
    hoy_dt  = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

    _MESES_ES_INV = {
        'ene': 1, 'feb': 2, 'mar': 3, 'abr': 4, 'may': 5, 'jun': 6,
        'jul': 7, 'ago': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dic': 12
    }
    def _parsear_entrada(fecha_str):
        fecha_str = str(fecha_str).strip()
        m = re.match(r'(\d{1,2})-([a-z]{3})-(\d{4})(?:\s+(\d{2}):(\d{2}))?', fecha_str, re.IGNORECASE)
        if m:
            mes = _MESES_ES_INV.get(m.group(2).lower())
            if mes:
                try:
                    return datetime(int(m.group(3)), mes, int(m.group(1)),
                                    int(m.group(4) or 0), int(m.group(5) or 0))
                except ValueError:
                    pass
        for fmt in ['%d/%m/%Y %H:%M', '%d/%m/%Y %H:%M:%S', '%d-%m-%Y %H:%M',
                    '%d-%m-%Y %H:%M:%S', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M',
                    '%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d']:
            try:
                return datetime.strptime(fecha_str, fmt)
            except ValueError:
                continue
        try:
            return pd.to_datetime(fecha_str, dayfirst=True).to_pydatetime()
        except Exception:
            return None

    with sqlite3.connect('unidades_reporte.db') as conn:
        df_alertas = pd.read_sql_query("SELECT rowid, * FROM reporte_unidades ORDER BY rowid DESC", conn)

    if not df_alertas.empty:
        hay_alertas = False
        for _, row in df_alertas.iterrows():
            fecha_entrada = str(row['entrada']).upper()
            unidad = row['unidad']
            if hoy_str in fecha_entrada:
                st.error(f"👀 **{unidad} FINALIZA HOY ({hoy_str})**.")
                hay_alertas = True
            elif "PENDIENTE" in fecha_entrada or "SIN FECHA" in fecha_entrada:
                st.warning(f"**{unidad}: {row['entrada']}**")
                hay_alertas = True
            else:
                dt_entrada = _parsear_entrada(str(row['entrada']))
                if dt_entrada is not None and dt_entrada < hoy_dt:
                    st.error(f"⚠️ **{unidad}: fecha de entrada {row['entrada']} ya venció — verificar estado.**")
                    hay_alertas = True
        if hay_alertas: st.divider()

    if 'df_candidatos' not in st.session_state:
        st.session_state['df_candidatos'] = pd.DataFrame(columns=["SELECCIONAR", "UNIDAD", "FECHA DE SALIDA", "CAUSA", "LIBRANZA", "FECHA DE ENTRADA", "POTENCIA INDISPONIBLE", "OBSERVACIONES"])

    if archivo_bitacora_usuario:
        if st.button("🔍 ANALIZAR BITÁCORA", type="primary"):
            df_bit = pd.read_excel(archivo_bitacora_usuario, header=None)
            cantidad, df_nuevos = procesar_bitacora_unidades(df_bit)
            
            if cantidad > 0:
                with sqlite3.connect('unidades_reporte.db') as conn:
                    df_existente = pd.read_sql_query("SELECT unidad FROM reporte_unidades", conn)
                    unidades_en_db = set(df_existente['unidad'].str.upper().tolist())

                for i, row in df_nuevos.iterrows():
                    if row['UNIDAD'] in unidades_en_db:
                        if row['OBSERVACIONES']:
                            df_nuevos.at[i, 'OBSERVACIONES'] += " | ⚠️ Ya existe en BD"
                        else:
                            df_nuevos.at[i, 'OBSERVACIONES'] = "⚠️ Ya existe en BD"

                st.success(f"✅ Se detectaron {cantidad} posibles indisponibilidades")
                st.session_state['df_candidatos'] = df_nuevos
            else:
                st.info("ℹ️ No se detectaron nuevas indisponibilidades")
                st.session_state['df_candidatos'] = pd.DataFrame() 

    if not st.session_state['df_candidatos'].empty:
        st.subheader("📝 Detecciones Recientes (Selecciona para guardar)")
        
        df_editor = st.data_editor(
            st.session_state['df_candidatos'],
            column_config={
                "SELECCIONAR": st.column_config.CheckboxColumn("Guardar", default=False),
                "CAUSA": st.column_config.TextColumn(width="large"),
                "OBSERVACIONES": st.column_config.TextColumn(width="medium", disabled=True)
            },
            hide_index=True,
            use_container_width=True,
            key="editor_candidatos"
        )
        
        if st.button("💾 Agregar a Base de Datos", type="secondary"):
            df_a_guardar = df_editor[df_editor['SELECCIONAR'] == True]
            
            if not df_a_guardar.empty:
                try:
                    with sqlite3.connect('unidades_reporte.db') as conn:
                        df_insert = df_a_guardar[["UNIDAD", "FECHA DE SALIDA", "CAUSA", "LIBRANZA", "FECHA DE ENTRADA", "POTENCIA INDISPONIBLE"]]
                        df_insert.columns = ["unidad", "salida", "causa", "libranza", "entrada", "potencia"]
                        
                        for _, row in df_insert.iterrows():
                            conn.execute("""
                                INSERT INTO reporte_unidades (unidad, salida, causa, libranza, entrada, potencia)
                                VALUES (?, ?, ?, ?, ?, ?)
                                ON CONFLICT(unidad) DO UPDATE SET 
                                salida = excluded.salida, causa = excluded.causa, 
                                libranza = excluded.libranza, entrada = 'PENDIENTE', potencia = excluded.potencia
                            """, (row['unidad'], row['salida'], row['causa'], row['libranza'], row['entrada'], row['potencia']))
                        conn.commit()
                    
                    st.success(f"✅ Se guardaron {len(df_a_guardar)} unidades.")
                    st.session_state['df_candidatos'] = pd.DataFrame() 
                    st.rerun()
                except Exception as e:
                    st.error(f"Error al guardar: {e}")
            else:
                st.warning("⚠️ No has seleccionado ninguna fila.")
        
        st.divider()

    st.subheader("💾 Base de Datos de Indisponibilidades")

    with st.expander("📥 Importar desde archivo Excel de Indisponibilidades"):
        archivo_import = st.file_uploader(
            "Sube el archivo Excel de indisponibilidades (ej. INDISPONIBILIDAD_DE_UNIDADES_2026.xlsx)",
            type=["xlsx", "xls"],
            key="import_excel_unidades"
        )
        modo_import = st.radio(
            "Modo de importación:",
            ["Agregar (no sobreescribir existentes)", "Reemplazar todo"],
            key="modo_import_unidades"
        )
        if archivo_import and st.button("📤 Importar a Base de Datos", type="primary", key="btn_import_excel"):
            try:
                wb_imp = openpyxl.load_workbook(archivo_import, read_only=True, data_only=True)
                ws_imp = wb_imp.active

                # Meses en español para parsear fechas largas (ej. "30-abr-2026 22:59")
                _MESES_IMP = {
                    'ene': 1, 'feb': 2, 'mar': 3, 'abr': 4, 'may': 5, 'jun': 6,
                    'jul': 7, 'ago': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dic': 12
                }

                def _fmt_fecha_import(val):
                    """Normaliza cualquier valor de fecha al formato interno dd/mm/YYYY HH:MM.

                    Soporta:
                    - datetime de Python/Excel  → siempre normaliza
                    - 'dd/mm/YYYY HH:MM'        → pasa sin cambio
                    - 'D-mmm-YYYY HH:MM'        → formato largo español (ej. 8-may-2026 19:00)
                    - ISO YYYY-MM-DD HH:MM[:SS] → normaliza
                    - Texto especial (SIN FECHA, PENDIENTE…) → devuelve tal cual
                    """
                    ESPECIALES = ('NONE', 'NAT', 'NAN', '')
                    SIN_FECHA  = "SIN FECHA DE SINCRONIZACION"

                    if val is None:
                        return SIN_FECHA
                    if isinstance(val, datetime):
                        return val.strftime("%d/%m/%Y %H:%M")

                    s = str(val).strip()
                    if not s or s.upper() in ESPECIALES:
                        return SIN_FECHA
                    # Texto especial que debe devolverse tal cual
                    if any(p in s.upper() for p in ('SIN FECHA', 'PENDIENTE', 'VER BIT')):
                        return s

                    # Formato largo español: D-mmm-YYYY[ H:MM] o D-mmm-YYYY[ HH:MM]
                    _m = re.match(
                        r'^(\d{1,2})-([a-záéíóú]{3})-(\d{4})(?:\s+(\d{1,2}):(\d{2}))?$',
                        s, re.IGNORECASE
                    )
                    if _m:
                        mes = _MESES_IMP.get(_m.group(2)[:3].lower())
                        if mes:
                            try:
                                dt = datetime(
                                    int(_m.group(3)), mes, int(_m.group(1)),
                                    int(_m.group(4) or 0), int(_m.group(5) or 0)
                                )
                                return dt.strftime("%d/%m/%Y %H:%M")
                            except ValueError:
                                pass

                    # Formatos numéricos estándar
                    for fmt in ["%d/%m/%Y %H:%M", "%d/%m/%Y %H:%M:%S",
                                "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M",
                                "%d-%m-%Y %H:%M", "%d-%m-%Y %H:%M:%S",
                                "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"]:
                        try:
                            return datetime.strptime(s, fmt).strftime("%d/%m/%Y %H:%M")
                        except ValueError:
                            continue

                    return s  # último recurso: devolver tal cual

                # ── Detección automática de la fila de cabecera ──────────────────
                # Busca la fila donde la columna B contiene 'UNIDAD' (cabecera de datos).
                # Esto evita el problema del min_row hardcodeado que saltaba la 1ª fila.
                primera_fila_datos = None
                for _r in ws_imp.iter_rows(max_row=30, values_only=True):
                    b_val = str(_r[1] if len(_r) > 1 else '').strip().upper()
                    if b_val == 'UNIDAD':
                        # La fila de datos empieza en la siguiente
                        primera_fila_datos = _r  # marcador; iter_rows seguirá desde aquí
                        break

                # Calcular el número de fila real de la cabecera para usar min_row
                fila_cabecera = 12  # valor por defecto si la hoja tiene la estructura estándar
                for idx, _r in enumerate(ws_imp.iter_rows(max_row=30, values_only=True), start=1):
                    b_val = str(_r[1] if len(_r) > 1 else '').strip().upper()
                    if b_val == 'UNIDAD':
                        fila_cabecera = idx
                        break

                filas_import = []
                for row in ws_imp.iter_rows(min_row=fila_cabecera + 1, values_only=True):
                    unidad = row[1] if len(row) > 1 else None
                    if not unidad or str(unidad).strip() in ('', 'UNIDAD', 'None'):
                        continue
                    unidad_str = str(unidad).strip().upper()
                    # Ignorar filas de pie de tabla (leyenda de colores, totales, etc.)
                    if unidad_str in ('▼', '▲', 'INDISPONIBILIDAD TOTAL:', ''):
                        continue
                    fecha_salida  = _fmt_fecha_import(row[2] if len(row) > 2 else None)
                    causa         = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                    libranza      = str(row[4]).strip() if len(row) > 4 and row[4] else "S/L"
                    fecha_entrada = _fmt_fecha_import(row[5] if len(row) > 5 else None)
                    try:
                        potencia = float(row[6]) if len(row) > 6 and row[6] is not None else 0.0
                    except (ValueError, TypeError):
                        potencia = 0.0
                    filas_import.append((unidad_str, fecha_salida, causa, libranza, fecha_entrada, potencia))

                wb_imp.close()

                if not filas_import:
                    st.warning("No se encontraron filas con datos en el archivo.")
                else:
                    with sqlite3.connect('unidades_reporte.db') as conn:
                        if modo_import == "Reemplazar todo":
                            conn.execute("DELETE FROM reporte_unidades")
                            for f in filas_import:
                                conn.execute(
                                    "INSERT INTO reporte_unidades (unidad,salida,causa,libranza,entrada,potencia) VALUES (?,?,?,?,?,?)", f
                                )
                        else:  # Agregar sin sobreescribir
                            for f in filas_import:
                                conn.execute("""
                                    INSERT OR IGNORE INTO reporte_unidades
                                    (unidad,salida,causa,libranza,entrada,potencia)
                                    VALUES (?,?,?,?,?,?)
                                """, f)
                        conn.commit()
                    st.success(f"✅ {len(filas_import)} unidades importadas correctamente.")
                    st.rerun()
            except Exception as e:
                st.error(f"Error al importar: {e}")

    with st.expander("➕ Agregar Unidad Manualmente"):
        c1, c2, c3 = st.columns([1, 1, 1])
        with c1:
            new_u = st.text_input("Unidad")
            new_salida = st.text_input("Fecha Salida")
        with c2:
            new_causa = st.text_area("Causa")
            new_lib = st.text_input("Libranza")
        with c3:
            new_pot = st.number_input("Potencia (MW)", value=0.0)
            new_ent = st.text_input("Fecha Entrada", value="SIN FECHA DE SINCRONIZACION")
        
        if st.button("💾 Guardar Manual"):
            if new_u:
                with sqlite3.connect('unidades_reporte.db') as conn:
                    conn.execute("INSERT OR REPLACE INTO reporte_unidades VALUES (?, ?, ?, ?, ?, ?)", 
                                 (new_u.upper(), new_salida, new_causa, new_lib, new_ent, new_pot))
                st.success("Guardado"); st.rerun()

    try:
        with sqlite3.connect('unidades_reporte.db') as conn:
            df_db = pd.read_sql_query("SELECT rowid, * FROM reporte_unidades", conn)
        
        if not df_db.empty:
            df_db.columns = ["ID", "UNIDAD", "FECHA DE SALIDA", "CAUSA", "LIBRANZA", "FECHA DE ENTRADA", "POTENCIA INDISPONIBLE"]
            
            # --- ORDENAMIENTO POR FECHA (MÁS RECIENTE PRIMERO) ---
            # pd.to_datetime falla con meses en español (ej. "30-abr-2026"), así que
            # primero normalizamos al formato numérico corto antes de parsear.
            _MESES_SORT = {
                'ene':'01','feb':'02','mar':'03','abr':'04','may':'05','jun':'06',
                'jul':'07','ago':'08','sep':'09','oct':'10','nov':'11','dic':'12'
            }
            def _normalizar_para_sort(s):
                """Convierte 'D-mmm-YYYY HH:MM' → 'DD/MM/YYYY HH:MM' para pd.to_datetime."""
                s = str(s).strip()
                m = re.match(r'^(\d{1,2})-([a-z]{3})-(\d{4})(.*)$', s, re.IGNORECASE)
                if m:
                    mes = _MESES_SORT.get(m.group(2).lower())
                    if mes:
                        resto = m.group(4).strip()
                        if resto:
                            return f"{int(m.group(1)):02d}/{mes}/{m.group(3)} {resto}"
                        return f"{int(m.group(1)):02d}/{mes}/{m.group(3)}"
                return s
            df_db['sort_key'] = pd.to_datetime(
                df_db['FECHA DE SALIDA'].map(_normalizar_para_sort),
                errors='coerce', dayfirst=True
            )
            df_db = df_db.sort_values(by='sort_key', ascending=False, na_position='last').drop(columns='sort_key').reset_index(drop=True)
            
            df_editado = st.data_editor(df_db, hide_index=True, use_container_width=True, num_rows="dynamic", column_config={"ID": None})
            
            col1, col2 = st.columns([1, 4])
            with col1:
                if st.button("💾 Guardar Cambios en Base de Datos"):
                    with sqlite3.connect('unidades_reporte.db') as conn:
                        conn.execute("DELETE FROM reporte_unidades")
                        df_save = df_editado.copy()
                        # Ajuste: Eliminamos la columna ID para el guardado real
                        df_save_to_db = df_save[["UNIDAD", "FECHA DE SALIDA", "CAUSA", "LIBRANZA", "FECHA DE ENTRADA", "POTENCIA INDISPONIBLE"]]
                        df_save_to_db.columns = ["unidad", "salida", "causa", "libranza", "entrada", "potencia"]
                        df_save_to_db.to_sql('reporte_unidades', conn, if_exists='append', index=False)
                    st.toast("¡Base de datos actualizada!"); st.rerun()
            
            st.markdown("---")
            c_ant, c_btn = st.columns([1, 2])
            with c_ant:
                ind_anterior = st.number_input("Indisponibilidad Anterior (MW)", value=0.0, format="%.2f")
            with c_btn:
                st.write("")
                generar_btn = st.button("📥 GENERAR ARCHIVO EXCEL", type="primary")

            if generar_btn:
                try:
                    from openpyxl.styles import PatternFill
                    from copy import copy

                    # ── Constantes de la plantilla ──────────────────────────────
                    FIRST_DATA_ROW  = 13   # primera fila de datos en la plantilla
                    ORIG_DATA_ROWS  = 4    # filas de datos que trae la plantilla (13-16)
                    ORIG_TOTAL_ROW  = 17   # fila con INDISPONIBILIDAD TOTAL y SUM
                    DATA_COLS       = list(range(2, 8))  # columnas B(2)..G(7)

                    def _safe_float(val):
                        try:
                            f = float(val)
                            return 0.0 if f != f else f
                        except (ValueError, TypeError):
                            return 0.0

                    # ── Cargar plantilla ────────────────────────────────────────
                    PLANTILLA_UNIDADES = os.path.join(os.path.dirname(os.path.abspath(__file__)), "plantilla_unidades.xlsx")
                    if not os.path.exists(PLANTILLA_UNIDADES):
                        st.error("⚠️ No se encontró 'plantilla_unidades.xlsx'. Colócala en la misma carpeta que app.py.")
                        st.stop()

                    wb = openpyxl.load_workbook(PLANTILLA_UNIDADES)
                    ws = wb.active

                    # ── Capturar estilos y alto de la fila de datos de referencia ──
                    ref_styles = {}
                    for col in DATA_COLS:
                        src = ws.cell(row=FIRST_DATA_ROW, column=col)
                        ref_styles[col] = {
                            'font':          copy(src.font),
                            'fill':          copy(src.fill),
                            'border':        copy(src.border),
                            'alignment':     copy(src.alignment),
                            'number_format': src.number_format,
                        }
                    ref_row_height = ws.row_dimensions[FIRST_DATA_ROW].height or 42.75

                    # ── Datos a escribir ─────────────────────────────────────────
                    df_rows    = df_editado.reset_index(drop=True)
                    n          = len(df_rows)
                    total_actual = sum(_safe_float(v) for v in df_rows['POTENCIA INDISPONIBLE'])

                    # ── Ajustar número de filas de datos ─────────────────────────
                    # Desmerger rangos en la zona de totales antes de mover filas
                    # (openpyxl no reubica merges al insertar/eliminar filas)
                    for m in list(ws.merged_cells.ranges):
                        if m.min_row >= ORIG_TOTAL_ROW:
                            ws.unmerge_cells(str(m))

                    # La plantilla trae ORIG_DATA_ROWS filas; necesitamos n filas.
                    if n < ORIG_DATA_ROWS:
                        for _ in range(ORIG_DATA_ROWS - n):
                            ws.delete_rows(FIRST_DATA_ROW + n)
                        total_row = ORIG_TOTAL_ROW - (ORIG_DATA_ROWS - n)
                    elif n > ORIG_DATA_ROWS:
                        extra = n - ORIG_DATA_ROWS
                        ws.insert_rows(ORIG_TOTAL_ROW, extra)
                        total_row = ORIG_TOTAL_ROW + extra
                        for i in range(ORIG_DATA_ROWS, n):
                            r = FIRST_DATA_ROW + i
                            ws.row_dimensions[r].height = ref_row_height
                            for col in DATA_COLS:
                                dst = ws.cell(row=r, column=col)
                                sty = ref_styles[col]
                                dst.font          = copy(sty['font'])
                                dst.fill          = copy(sty['fill'])
                                dst.border        = copy(sty['border'])
                                dst.alignment     = copy(sty['alignment'])
                                dst.number_format = sty['number_format']
                            ws.cell(row=r, column=8).value = (
                                f'=+IF(B{r}="","",IF(F{r}<NOW(),"REVISAR FECHA DE ENTRADA",""))'
                            )
                    else:
                        total_row = ORIG_TOTAL_ROW

                    # Re-mergear F y G de la fila de totales (span 2 filas como en plantilla)
                    ws.merge_cells(f'F{total_row}:F{total_row + 1}')
                    ws.merge_cells(f'G{total_row}:G{total_row + 1}')

                    # ── Actualizar fórmula SUM y referencia de C9 ────────────────
                    last_data_row = FIRST_DATA_ROW + n - 1
                    ws.cell(row=total_row, column=7).value = f'=SUM(G{FIRST_DATA_ROW}:G{last_data_row})'
                    ws['C9'].value = f'=G{total_row}'

                    # ── Escribir solo C8 (ind. anterior); C9 y C10 son fórmulas ──
                    ws['C8'].value = ind_anterior

                    # Color condicional C10
                    cambio = total_actual - ind_anterior
                    if ind_anterior > total_actual:
                        ws['C10'].fill = PatternFill("solid", fgColor="14A35C")   # verde
                    elif total_actual > ind_anterior:
                        ws['C10'].fill = PatternFill("solid", fgColor="B92738")   # rojo

                    # ── Escribir filas de datos ───────────────────────────────────
                    for i, row in df_rows.iterrows():
                        r = FIRST_DATA_ROW + i
                        ws.cell(row=r, column=2).value = row['UNIDAD']
                        ws.cell(row=r, column=3).value = formatear_fecha_excel(row['FECHA DE SALIDA'])
                        causa_cell = ws.cell(row=r, column=4)
                        causa_cell.value = row['CAUSA']
                        from openpyxl.styles import Alignment as _Align
                        causa_cell.alignment = _Align(horizontal='left', vertical='center', wrap_text=True)
                        ws.cell(row=r, column=5).value = row['LIBRANZA']
                        ws.cell(row=r, column=6).value = formatear_fecha_excel(row['FECHA DE ENTRADA'])
                        ws.cell(row=r, column=7).value = _safe_float(row['POTENCIA INDISPONIBLE'])

                    # ── Guardar en BytesIO ────────────────────────────────────────
                    output = BytesIO()
                    wb.save(output)
                    output.seek(0)

                    st.success("✅ Archivo listo para descargar.")
                    st.download_button(
                        "⏬ Descargar Archivo de indisponibilidades",
                        data=output.getvalue(),
                        file_name=f"INDISPONIBILIDAD DE UNIDADES {datetime.now().year}.xlsx"
                    )
                except Exception as e:
                    st.error(f"Error generando Excel: {e}")
        else:
            st.info("La base de datos está vacía.")
            
    except Exception as e:
        st.error(f"Error general: {e}")

def _natural_case(text):
    """Escritura natural: todo-mayúsculas → oración capitalizada conservando siglas y códigos."""
    if not text:
        return text
    letras = [c for c in text if c.isalpha()]
    if letras and not all(c.isupper() for c in letras):
        return text[0].upper() + text[1:]
    result = text.lower()
    result = re.sub(r'(^|(?<=[.!?])\s+)([a-z])',
                    lambda m: m.group(1) + m.group(2).upper(), result)
    for s in ['SIN','AGC','ETESA','SPEAR','CND','MW','MWH','GWH','KW','KWH',
              'KV','HZ','CPS','ASEP','EDEMET','ENSA','EDESUR','NATURGY','AES',
              'GENA','BLM','IDEAL','EOR','RTR','CNI','TLB','CF','CPP','CPPFB',
              'CFT','TLTB','CEPSCO','INDE','ING','SR','SRA','RHH']:
        result = re.sub(r'(?<![A-Za-z])' + s.lower() + r'(?![A-Za-z])', s, result)
    result = re.sub(r'(\d)([a-z])', lambda m: m.group(1) + m.group(2).upper(), result)
    return result


def generar_xml_oficial_eor(df_pan, df_rtr, datos_op, datos_df=None):
    """
    Genera el XML oficial del reporte EOR con el formato requerido por el CND.
    datos_df: dict con claves dem_max_h, dem_max_v, dem_min_h, dem_min_v,
                               frec_max_h, frec_max_v, frec_min_h, frec_min_v
    """
    fecha_dt  = datetime.now() - timedelta(days=1)
    fecha_xml = fecha_dt.strftime("%Y-%m-%d")   # formato requerido: YYYY-MM-DD
    fecha_indis = fecha_dt.strftime("%d/%m/%Y") # formato interno para <FECHA> de indisponibilidad

    def _asegurar_hhmmss(hora: str) -> str:
        """Convierte HH:MM → HH:MM:SS. Si ya tiene segundos, lo deja igual."""
        hora = str(hora).strip()
        if re.match(r'^\d{2}:\d{2}$', hora):
            return hora + ":00"
        return hora if hora else "00:00:00"

    def _calcular_duracion(ini: str, fin: str) -> str:
        """Calcula HH:MM:SS de diferencia entre dos tiempos HH:MM o HH:MM:SS."""
        try:
            fmt = '%H:%M:%S'
            t_ini = datetime.strptime(ini, fmt)
            t_fin = datetime.strptime(fin, fmt)
            if t_fin < t_ini:
                t_fin += timedelta(days=1)
            dur = int((t_fin - t_ini).total_seconds())
            return f"{dur // 3600:02d}:{(dur % 3600) // 60:02d}:{dur % 60:02d}"
        except Exception:
            return "00:00:00"

    # Mapeo de valores internos a texto natural requerido por el formato
    _ELEMENTO = {
        "LINEA":         "Línea de Transmisión",
        "TRANSFORMADOR": "Transformador",
        "INTERRUPTOR":   "Interruptor",
        "BARRA":         "Barra",
    }
    _TIPO_INDIS = {
        "PROGRAMADA":    "Indisponibilidad programada",
        "NO PROGRAMADA": "Indisponibilidad no programada",
    }
    _CAUSA = {
        "MANTENIMIENTO": "Desconexión",
        "DISPARO":       "Disparo",
    }

    root = ET.Element("INFORME")
    root.set("fecha", fecha_xml)
    ET.SubElement(root, "Titulo").text = "Reporte Diario EOR"
    ET.SubElement(root, "Area").text   = "Panamá"

    # ── AGC ──────────────────────────────────────────────────────────────────
    agc = ET.SubElement(root, "AGC")
    ET.SubElement(agc, "MODO").text          = str(datos_op.get('modo', ''))
    ET.SubElement(agc, "BIAS").text          = str(datos_op.get('bias', ''))
    ET.SubElement(agc, "TICONTROLADOR").text = "150"

    # ── Demanda y Frecuencia ─────────────────────────────────────────────────
    df_vals = datos_df or {}
    for tag, hmax_k, vmax_k, hmin_k, vmin_k in [
        ("Demanda",    "dem_max_h",  "dem_max_v",  "dem_min_h",  "dem_min_v"),
        ("Frecuencia", "frec_max_h", "frec_max_v", "frec_min_h", "frec_min_v"),
    ]:
        bloque = ET.SubElement(root, tag)
        for tipo, hk, vk in [("Maxima", hmax_k, vmax_k), ("Minima", hmin_k, vmin_k)]:
            sub = ET.SubElement(bloque, tipo)
            ET.SubElement(sub, "Hora").text  = _asegurar_hhmmss(df_vals.get(hk, ""))
            ET.SubElement(sub, "Valor").text = str(df_vals.get(vk, ""))

    # ── Eventos ──────────────────────────────────────────────────────────────
    eventos_tag = ET.SubElement(root, "EVENTOS")
    if not df_pan.empty:
        for _, row in df_pan.iterrows():
            desc = str(row.get("Descripción", ""))
            ev = ET.SubElement(eventos_tag, "EVENTO")
            if "NO SE REGISTR" in desc.upper():
                ET.SubElement(ev, "INICIO").text      = "00:00:00"
                ET.SubElement(ev, "FIN").text         = "23:59:00"
                ET.SubElement(ev, "DESCRIPCION").text = _natural_case(desc.strip())
            else:
                hora_ev = _asegurar_hhmmss(str(row.get("Hora", "00:00")))
                ET.SubElement(ev, "INICIO").text      = hora_ev
                ET.SubElement(ev, "FIN").text         = " "
                ET.SubElement(ev, "DESCRIPCION").text = _natural_case(desc.strip())
    else:
        ev = ET.SubElement(eventos_tag, "EVENTO")
        ET.SubElement(ev, "INICIO").text      = "00:00:00"
        ET.SubElement(ev, "FIN").text         = "23:59:00"
        ET.SubElement(ev, "DESCRIPCION").text = "No se registraron eventos relevantes"

    # ── Indisponibilidades ───────────────────────────────────────────────────
    indis_tag = ET.SubElement(root, "INDISPONIBILIDADES")
    indis_tag.text = indis_tag.text or ""
    if not df_rtr.empty:
        for _, row in df_rtr.iterrows():
            h_ini = _asegurar_hhmmss(str(row.get('Hora Inicio CA+1', '')))
            h_fin = _asegurar_hhmmss(str(row.get('Hora Fin CA+1',   '')))
            dur   = _calcular_duracion(h_ini, h_fin)

            tipo_raw  = str(row.get('Tipo de Elemento',       '')).strip().upper()
            indis_raw = str(row.get('Tipo de indisponibilidad','')). strip().upper()
            causa_raw = str(row.get('Tipo de Causa',          '')).strip().upper()

            elemento = _ELEMENTO.get(tipo_raw,  str(row.get('Tipo de Elemento', '')).strip())
            tipo_ind = _TIPO_INDIS.get(indis_raw, str(row.get('Tipo de indisponibilidad', '')).strip())
            causa    = _CAUSA.get(causa_raw,     str(row.get('Tipo de Causa', '')).strip())
            desc_raw = str(row.get('Descripción', '')).strip()
            descripcion = _natural_case(desc_raw) if desc_raw else ""

            ind = ET.SubElement(indis_tag, "INDISPONIBILIDAD")
            ET.SubElement(ind, "FECHA").text       = fecha_indis
            ET.SubElement(ind, "INICIO").text      = h_ini
            ET.SubElement(ind, "FIN").text         = h_fin
            ET.SubElement(ind, "DURACION").text    = dur
            ET.SubElement(ind, "ELEMENTO").text    = elemento
            ET.SubElement(ind, "TIPO").text        = tipo_ind
            ET.SubElement(ind, "CAUSA").text       = causa
            ET.SubElement(ind, "NODOI").text       = str(row.get('Nodo I',  '') or '')
            ET.SubElement(ind, "NODOJ").text       = str(row.get('Nodo J',  '') or '')
            ET.SubElement(ind, "NODOK").text       = str(row.get('Nodo K',  '') or '')
            ET.SubElement(ind, "CKT").text         = str(row.get('CKT',     '') or '')
            ET.SubElement(ind, "ENERGIA").text     = str(row.get('Energía no servida (MWh)', '0') or '0')
            ET.SubElement(ind, "DESCRIPCION").text = descripcion

    xml_str = ET.tostring(root, encoding='utf-8')
    pretty  = minidom.parseString(xml_str).toprettyxml(indent="  ")
    # Expandir tags vacíos auto-cerrados (<TAG/> → <TAG></TAG>) según el formato requerido
    pretty = re.sub(r'<(\w+)/>', r'<\1></\1>', pretty)
    return pretty

def obtener_mes_nombre(fecha_dt):
    meses = {1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio", 
             7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"}
    return meses.get(fecha_dt.month, "")

def _extraer_eventos_pan(df_bit):
    """Extrae eventos relevantes de la bitácora para el reporte EOR. Retorna lista de dicts."""
    keywords = [
        "FRECUENCIA BAJO", "ABIERTA INTERCONEXION",
        "NORMALIZADO(A) SIN", "REANUDA", "SUSPENDE", "CERRADA INTERCONEXION",
        "CAMBIAR EL BIAS", "DESVIACION", "AJUSTE DE BIAS", "NORMALIZACION DE BIAS", "DISPARO INTERCONEXION",
        "EL INTERCAMBIO OSCILO", "FRECUENCIA SUBIO", "FRECUENCIA OSCILO", "EL INTERCAMBIO BAJO","EL INTERCAMBIO SUBIO",
        "ESQUEMABAJA FRECUENCIA", "ESQUEMA BAJA FRECUENCIA", "BAJA FRECUENCIA",
        "ESQUEMABAJO VOLTAJE", "ESQUEMA BAJO VOLTAJE", "BAJO VOLTAJE",
        "ACTIVADO(A)", "DESACTIVADO(A)", "PRIMER ESCALON", "SEGUNDO ESCALON", "TERCER ESCALON","CODIGO AMARILLO","CODIGO BLANCO",
    ]
    excluir = ["LIMITADO", "LIMITADA", "RESTRICCION", "LIMITA", "EQUIPOS ESPECIALES", "LIBRANZA","CPS1"]
    keywords_frec_intercambio = [
        "FRECUENCIA BAJO", "FRECUENCIA SUBIO", "FRECUENCIA OSCILO",
        "DISPARO INTERCONEXION", "EL INTERCAMBIO OSCILO", "EL INTERCAMBIO BAJO",
        "ABIERTA INTERCONEXION", "CERRADA INTERCONEXION","EL INTERCAMBIO SUBIO",
    ]

    import unicodedata as _ud
    def _norm(t): return _ud.normalize('NFKD', str(t)).encode('ascii', 'ignore').decode('ascii').upper()

    eventos = []
    for _, row in df_bit.iterrows():
        texto = " ".join(str(x) for x in row.values if pd.notnull(x)).upper()
        texto_norm = _norm(texto)

        if any(_norm(e) in texto_norm for e in excluir):
            continue

        es_esquema = any(_norm(s) in texto_norm for s in ["ESQUEMA", "FRECUENCIA", "VOLTAJE", "ESCALON"])
        es_disparo = "DISPARO" in texto_norm
        es_frec_intercambio = any(_norm(k) in texto_norm for k in keywords_frec_intercambio)
        es_bloque_carga = "BLOQUE DE CARGA" in texto_norm  # SPEAR activó descarga de carga

        def _valor_hz(t):
            m = re.search(r'(\d{2}[.,]\d+)\s*(?:HZ|CPS)?', _norm(t))
            if m:
                return float(m.group(1).replace(',', '.'))
            return None

        pasa_filtro = False
        if es_frec_intercambio:
            es_evento_frec = any(_norm(k) in texto_norm for k in [
                "FRECUENCIA BAJO", "FRECUENCIA SUBIO", "FRECUENCIA OSCILO"
            ])
            if es_evento_frec:
                hz = _valor_hz(texto)
                pasa_filtro = hz is None or hz <= 59.8 or hz >= 60.2
            else:
                pasa_filtro = True
        elif es_disparo:
            match_mw = re.search(r'(\d+(?:\.\d+)?)\s*MW', texto_norm)
            pasa_filtro = float(match_mw.group(1)) >= 60 if match_mw else False
        else:
            if "ACTIVADO(A)" in texto_norm or "DESACTIVADO(A)" in texto_norm:
                # Incluir esquemas (BF, BV, etc.) Y bloques de carga por SPEAR
                pasa_filtro = es_esquema or es_bloque_carga
            elif "NORMALIZADO" in texto_norm:
                # Incluir NORMALIZADO de esquemas Y de SPEAR/BLOQUE DE CARGA
                pasa_filtro = es_esquema or es_bloque_carga or "SPEAR" in texto_norm
            else:
                pasa_filtro = any(_norm(k) in texto_norm for k in keywords)

        if pasa_filtro:
            h_m = re.search(r'(\d{2}:\d{2})', texto)
            hora = h_m.group(1) if h_m else "00:00"
            desc = re.sub(r'\d{2}/\d{2}/\d{4}', '', texto).replace(hora, "").strip(" :,-")
            eventos.append({"Hora": hora, "Descripción": desc})

    return eventos


def _mejorar_descripciones_eor(eventos: list) -> list:
    """Envía los eventos del EOR a Claude para mejorar la redacción de las descripciones.

    Recibe una lista de dicts {'Hora': str, 'Descripción': str} y devuelve la misma
    lista con las descripciones parafraseadas en lenguaje técnico formal.
    Si la API falla devuelve la lista original intacta.
    """
    if not eventos or not _api_key_libranzas:
        return eventos

    # Construir el bloque de texto con los eventos numerados
    bloques = "\n".join(
        f"{i+1}. [{ev['Hora']}] {ev['Descripción']}"
        for i, ev in enumerate(eventos)
    )

    prompt = (
        "Eres un redactor técnico del Centro Nacional de Despacho de Panamá (CND/ETESA). "
        "Tu tarea es parafrasear las siguientes descripciones de eventos del Sistema Eléctrico Interconectado Nacional (SEIN) "
        "para el Reporte Diario del EOR.\n\n"
        "REGLAS ESTRICTAS:\n"
        "- Devuelve EXACTAMENTE el mismo número de descripciones, en el mismo orden.\n"
        "- Conserva los datos técnicos: valores numéricos, códigos de equipo, frecuencias, tensiones, nombres propios.\n"
        "- Redacta en tercera persona, tiempo pasado, lenguaje técnico formal pero claro.\n"
        "- Elimina ruido de bitácora: fechas, horas, nombres de operadores, frases redundantes como 'SE INFORMA QUE', 'SE COMUNICA'.\n"
        "- Máximo 2 oraciones por evento.\n"
        "- Usa mayúsculas solo para siglas y nombres propios de equipos (ej. AGC, SPEAR, ETESA). El resto en minúsculas.\n"
        "- Responde ÚNICAMENTE con un JSON array de strings con las descripciones mejoradas, sin explicación, sin markdown.\n\n"
        f"EVENTOS ({len(eventos)}):\n{bloques}"
    )

    try:
        client = anthropic.Anthropic(api_key=_api_key_libranzas)
        resp = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=2000,
            messages=[{"role": "user", "content": prompt}]
        )
        text = resp.content[0].text.strip()
        # Limpiar bloques markdown si la IA los añade por error
        text = re.sub(r'^```[a-z]*\n?', '', text)
        text = re.sub(r'\n?```$', '', text)
        nuevas = json.loads(text)
        if isinstance(nuevas, list) and len(nuevas) == len(eventos):
            return [
                {"Hora": ev["Hora"], "Descripción": str(desc).strip()}
                for ev, desc in zip(eventos, nuevas)
            ]
    except Exception:
        pass

    return eventos  # fallback: devuelve los originales


def vista_pan(archivo_bitacora):
    st.title("📋 REPORTE DIARIO DEL EOR")

    # ── Cargar reporte EOR ya generado ──────────────────────────────────────────
    with st.expander("📂 Cargar Reporte EOR ya generado", expanded=False):
        st.caption("Sube el reporte del EOR generado anteriormente Excel/XLSM. ")
        uploaded_pan_prev = st.file_uploader(
            "Reporte EOR (.xlsx / .xlsm)", type=["xlsx", "xlsm", "xls"], key="pan_upload_previo"
        )
        if uploaded_pan_prev is not None:
            try:
                _raw_pan = uploaded_pan_prev.read()
                # Intentar leer como Excel con columnas Hora / Descripción (formato simple)
                _eventos_cargados = None
                for _engine in (None, "openpyxl", "xlrd"):
                    try:
                        _kw = {"engine": _engine} if _engine else {}
                        _df_pan_prev = pd.read_excel(BytesIO(_raw_pan), header=0, **_kw)
                        # Buscar columnas Hora y Descripción (insensible a mayúsculas)
                        _col_map = {c.strip().lower(): c for c in _df_pan_prev.columns}
                        _h_col = _col_map.get("hora")
                        _d_col = _col_map.get("descripción") or _col_map.get("descripcion")
                        if _h_col and _d_col:
                            _eventos_cargados = [
                                {"Hora": str(r[_h_col]).strip(), "Descripción": str(r[_d_col]).strip()}
                                for _, r in _df_pan_prev.iterrows()
                                if pd.notnull(r[_d_col]) and str(r[_d_col]).strip()
                            ]
                            break
                    except Exception:
                        continue

                # Si no encontró columnas con encabezado, intentar leer la plantilla xlsm
                # buscando filas donde columna A tenga formato HH:MM y col B tenga texto
                if not _eventos_cargados:
                    for _engine in (None, "openpyxl", "xlrd"):
                        try:
                            _kw2 = {"engine": _engine} if _engine else {}
                            _df_raw = pd.read_excel(BytesIO(_raw_pan), header=None, **_kw2)
                            _ev_tmp = []
                            for _, _row in _df_raw.iterrows():
                                _hora_v = str(_row.iloc[0]).strip() if pd.notnull(_row.iloc[0]) else ""
                                _desc_v = str(_row.iloc[1]).strip() if len(_row) > 1 and pd.notnull(_row.iloc[1]) else ""
                                if re.match(r'^\d{1,2}:\d{2}$', _hora_v) and _desc_v:
                                    _ev_tmp.append({"Hora": _hora_v, "Descripción": _desc_v})
                            if _ev_tmp:
                                _eventos_cargados = _ev_tmp
                                break
                        except Exception:
                            continue

                if _eventos_cargados:
                    st.session_state["pan_eventos"]      = _eventos_cargados
                    st.session_state["pan_from_upload"]  = True
                    st.session_state.pop("pan_archivo_id", None)
                    st.success(f"✅ Reporte EOR cargado correctamente — {len(_eventos_cargados)} evento(s).")
                else:
                    st.error(
                        "No se encontraron eventos en el archivo. "
                        "El Excel debe tener columnas **Hora** y **Descripción**, "
                        "o bien ser el reporte EOR generado por la app."
                    )
            except Exception as _e:
                st.error(f"No se pudo leer el reporte EOR: {_e}")

        if st.session_state.get("pan_from_upload"):
            
            if st.button("🔄 Volver a procesar desde bitácora", key="pan_clear_upload"):
                st.session_state.pop("pan_from_upload", None)
                st.session_state.pop("pan_eventos", None)
                st.session_state.pop("pan_archivo_id", None)
                st.rerun()

    if not archivo_bitacora and not st.session_state.get("pan_from_upload"):
        st.info("👈 Sube la bitácora en el menú lateral para habilitar los reportes.")
        return

    try:
        # ── Botón para forzar reprocesamiento ────────────────────────────────
        col_proc, _ = st.columns([1, 3])
        if col_proc.button("🔄 Forzar reprocesamiento", key="pan_forzar"):
            st.session_state.pop("pan_archivo_id", None)
            st.session_state.pop("pan_from_upload", None)
            st.session_state.pop("pan_ia_aplicada", None)
            st.session_state.pop("pan_eventos_original", None)
            # Limpiar archivos generados anteriormente
            st.session_state.pop("eor_excel_bytes", None)
            st.session_state.pop("eor_excel_nombre", None)
            st.session_state.pop("eor_xml_bytes", None)
            st.session_state.pop("eor_xml_nombre", None)

        # ── Carga inicial: solo cuando cambia el archivo y no hay upload manual ──
        if not st.session_state.get("pan_from_upload") and archivo_bitacora:
            archivo_id = f"{archivo_bitacora.name}_{archivo_bitacora.size}"
            if st.session_state.get("pan_archivo_id") != archivo_id:
                df_bit = pd.read_excel(archivo_bitacora, header=None)
                eventos = _extraer_eventos_pan(df_bit)
                if not eventos:
                    eventos = [{"Hora": "23:59", "Descripción": "NO SE REGISTRARON EVENTOS."}]
                st.session_state["pan_eventos"] = eventos
                st.session_state["pan_archivo_id"] = archivo_id
                st.session_state.pop("pan_ia_aplicada", None)
                st.session_state.pop("pan_eventos_original", None)

        # ── Estado en sesión ──────────────────────────────────────────────────
        if "pan_eventos" not in st.session_state:
            st.session_state["pan_eventos"] = []

        # ── Tabla principal con columna Eliminar ─────────────────────────────
        st.subheader("📋 Eventos del Reporte")

        df_tabla = pd.DataFrame(st.session_state["pan_eventos"])
        df_tabla.insert(0, "🗑️", False)   # Checkbox de selección para borrar

        df_editado = st.data_editor(
            df_tabla,
            hide_index=True,
            use_container_width=True,
            column_config={
                "🗑️": st.column_config.CheckboxColumn("🗑️", help="Marcar para eliminar", width="small"),
                "Hora": st.column_config.TextColumn("Hora", width="small"),
                "Descripción": st.column_config.TextColumn("Descripción"),
            },
            key="editor_pan_v5",
        )

        # Sincronizar ediciones del data_editor de vuelta a session_state inmediatamente
        if "Hora" in df_editado.columns and "Descripción" in df_editado.columns:
            filas_sync = df_editado[df_editado["🗑️"] == False][["Hora", "Descripción"]].to_dict("records")
            st.session_state["pan_eventos"] = filas_sync

        # Botón eliminar seleccionados
        col_del, col_ia, col_rev, col_space = st.columns([1, 1.3, 1, 2])
        with col_del:
            if st.button("🗑️ Eliminar seleccionados", type="secondary", help= "Elimina los registros no deseados del reporte"):
                filas_mantener = df_editado[df_editado["🗑️"] == True]
                if not filas_mantener.empty:
                    st.session_state["pan_eventos"] = df_editado[df_editado["🗑️"] == False][["Hora", "Descripción"]].to_dict("records")
                    st.rerun()

        with col_ia:
            if st.button("✨ Mejorar redacción con IA", type="primary", help="Parafrasea las anotaciones de bitácora en lenguaje técnico formal"):
                eventos_actuales = df_editado[df_editado["🗑️"] == False][["Hora", "Descripción"]].to_dict("records")
                if not eventos_actuales:
                    st.warning("No hay registros para mejorar.")
                elif not _api_key_libranzas:
                    st.error("No hay API key configurada en config.txt.")
                else:
                    # Guardar original para poder revertir
                    st.session_state["pan_eventos_original"] = eventos_actuales.copy()
                    with st.spinner("Mejorando descripciones con IA…"):
                        mejorados = _mejorar_descripciones_eor(eventos_actuales)
                    if mejorados is eventos_actuales:  # fallback → no cambió nada
                        st.error("No se pudo conectar con la IA. Verifica la API key.")
                    else:
                        st.session_state["pan_eventos"] = mejorados
                        st.session_state["pan_ia_aplicada"] = True
                        st.rerun()

        with col_rev:
            if st.session_state.get("pan_ia_aplicada") and st.session_state.get("pan_eventos_original"):
                if st.button("↩️ Revertir", help="Vuelve a las descripciones originales de la bitácora"):
                    st.session_state["pan_eventos"] = st.session_state.pop("pan_eventos_original")
                    st.session_state.pop("pan_ia_aplicada", None)
                    st.rerun()

        if st.session_state.get("pan_ia_aplicada"):
            st.caption("✅ Descripciones mejoradas por IA — puedes editarlas libremente en la tabla.")

        st.markdown("---")

        # ── Agregar registro manual ───────────────────────────────────────────
        with st.expander("➕ Agregar registro manualmente"):
            c1, c2, c3 = st.columns([1, 4, 1])
            hora_nueva = c1.text_input("Hora (HH:MM)", placeholder="07:45", key="pan_hora_nueva")
            desc_nueva = c2.text_input("Descripción", placeholder="Descripción del evento", key="pan_desc_nueva")
            agregar = c3.button("Agregar", type="primary", use_container_width=True)

            if agregar:
                hora_nueva = hora_nueva.strip()
                desc_nueva = desc_nueva.strip()
                if not hora_nueva or not desc_nueva:
                    st.warning("Completa la hora y la descripción antes de agregar.")
                elif not re.match(r'^\d{2}:\d{2}$', hora_nueva):
                    st.warning("El formato de hora debe ser HH:MM (ej. 07:45).")
                else:
                    # Guardar cambios manuales del editor antes de agregar
                    filas_actuales = df_editado[df_editado["🗑️"] == False][["Hora", "Descripción"]].to_dict("records")
                    filas_actuales.append({"Hora": hora_nueva, "Descripción": desc_nueva.upper()})
                    # Ordenar por hora
                    filas_actuales.sort(key=lambda x: x["Hora"])
                    st.session_state["pan_eventos"] = filas_actuales
                    st.rerun()

        st.markdown("---")

        # ── CSS: fuentes más grandes para la sección AGC / Demanda ──────────
        st.markdown("""
        <style>
        /* Labels de los inputs de AGC */
        div[data-testid="stSelectbox"] label,
        div[data-testid="stTextInput"] label {
            font-size: 15px !important;
        }
        /* Texto dentro de los inputs de AGC */
        div[data-testid="stSelectbox"] div[data-baseweb="select"] span,
        div[data-testid="stTextInput"] input {
            font-size: 15px !important;
        }
        /* Cabecera del data_editor (nombres de columnas) */
        div[data-testid="stDataFrame"] th,
        div[data-testid="stDataFrameResizable"] th {
            font-size: 15px !important;
        }
        /* Celdas del data_editor */
        div[data-testid="stDataFrame"] td,
        div[data-testid="stDataFrameResizable"] td,
        div[data-testid="stDataFrame"] div[class*="cell"],
        .dvn-scroller div[role="gridcell"],
        .dvn-scroller div[role="columnheader"] {
            font-size: 20px !important;
        }
        </style>
        """, unsafe_allow_html=True)

        # ── AGC + Demanda + Frecuencia (layout tipo tabla Excel) ─────────────
        st.subheader("⚙️ AGC / 📊 Demanda y Frecuencia")
        st.caption("Completa todos los valores antes de exportar el XML. Puedes pegar directamente desde el Excel.")

        _col_agc, _spacer, _col_df = st.columns([1.0, 0.08, 1.6])

        # ── Bloque AGC ────────────────────────────────────────────────────────
        with _col_agc:
            modo_agc = st.selectbox("Modo de control de AGC", ["TLB", "CPPFB", "CNI", "CF", "TLTB", "CFT", "CPP"])
            bias_val = st.text_input("Bias (MW/0.1Hz)", key="pan_bias_val", value="19.15")

        # ── Bloque Demanda / Frecuencia como tabla editable ───────────────────
        with _col_df:
            st.markdown("<p style='font-size:15px;font-weight:600;margin-bottom:4px;'>Demanda y Frecuencia</p>", unsafe_allow_html=True)

            # ── Importar desde reporte EOR previo ─────────────────────────────
            with st.expander("📂 Importar desde reporte EOR (.xlsm / .xlsx)"):
                archivo_eor_imp = st.file_uploader(
                    "Sube el reporte EOR del día anterior",
                    type=["xlsm", "xlsx"],
                    key="eor_import_file"
                )
                if archivo_eor_imp and st.button("⬆️ Extraer valores", key="eor_import_btn"):
                    try:
                        wb_imp = openpyxl.load_workbook(archivo_eor_imp, data_only=True)
                        ws_imp = wb_imp.active

                        def _leer_celda(ws, row, col):
                            v = ws.cell(row=row, column=col).value
                            return v

                        def _fmt_hora(v):
                            if v is None:
                                return ""
                            if hasattr(v, 'strftime'):
                                return v.strftime("%H:%M")
                            s = str(v).strip()
                            # Si viene como "HH:MM:SS" truncar a HH:MM
                            if re.match(r'^\d{2}:\d{2}', s):
                                return s[:5]
                            return s

                        def _fmt_valor(v, decimales=2):
                            if v is None:
                                return ""
                            try:
                                return str(round(float(v), decimales))
                            except (ValueError, TypeError):
                                return str(v)

                        dem_max_v_imp  = _fmt_valor(_leer_celda(ws_imp, 23, 6), 0)
                        dem_max_h_imp  = _fmt_hora(_leer_celda(ws_imp, 23, 7))
                        dem_min_v_imp  = _fmt_valor(_leer_celda(ws_imp, 24, 6), 0)
                        dem_min_h_imp  = _fmt_hora(_leer_celda(ws_imp, 24, 7))
                        frec_max_v_imp = _fmt_valor(_leer_celda(ws_imp, 25, 6), 4)
                        frec_max_h_imp = _fmt_hora(_leer_celda(ws_imp, 25, 7))
                        frec_min_v_imp = _fmt_valor(_leer_celda(ws_imp, 26, 6), 4)
                        frec_min_h_imp = _fmt_hora(_leer_celda(ws_imp, 26, 7))
                        wb_imp.close()

                        st.session_state["pan_df_tabla"] = pd.DataFrame({
                            "Parámetro": ["Demanda Máxima, MW", "Demanda Mínima, MW",
                                          "Frecuencia Máxima, Hz", "Frecuencia Mínima, Hz"],
                            "Valor":     [dem_max_v_imp, dem_min_v_imp,
                                          frec_max_v_imp, frec_min_v_imp],
                            "HORA CA+1": [dem_max_h_imp, dem_min_h_imp,
                                          frec_max_h_imp, frec_min_h_imp],
                        })
                        st.success("✅ Valores extraídos correctamente.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Error al leer el archivo: {e}")

            _df_init = pd.DataFrame({
                "Parámetro":  ["Demanda Máxima, MW", "Demanda Mínima, MW",
                               "Frecuencia Máxima, Hz", "Frecuencia Mínima, Hz"],
                "Valor":      ["", "", "", ""],
                "HORA CA+1":  ["", "", "", ""],
            })

            # Pre-poblar con lo que ya hay en session_state (para no perder datos en reruns)
            if "pan_df_tabla" not in st.session_state:
                st.session_state["pan_df_tabla"] = _df_init.copy()

            _df_edit = st.data_editor(
                st.session_state["pan_df_tabla"],
                hide_index=True,
                use_container_width=True,
                column_config={
                    "Parámetro": st.column_config.TextColumn("Parámetro", disabled=True, width="medium"),
                    "Valor":     st.column_config.TextColumn("Valor",     width="small"),
                    "HORA CA+1": st.column_config.TextColumn("HORA CA+1", width="small"),
                },
                row_height=42,
                key="editor_df_tabla",
            )
            # Persistir ediciones
            st.session_state["pan_df_tabla"] = _df_edit

        # Extraer valores de la tabla para usarlos en el XML
        def _tv(row_idx, col):
            try:    return str(_df_edit.iloc[row_idx][col]).strip()
            except: return ""

        dem_max_v  = _tv(0, "Valor");     dem_max_h  = _tv(0, "HORA CA+1")
        dem_min_v  = _tv(1, "Valor");     dem_min_h  = _tv(1, "HORA CA+1")
        frec_max_v = _tv(2, "Valor");     frec_max_h = _tv(2, "HORA CA+1")
        frec_min_v = _tv(3, "Valor");     frec_min_h = _tv(3, "HORA CA+1")

        try:    bias_num = float(bias_val.replace(",", "."))
        except: bias_num = 19.15

        datos_op = {'modo': modo_agc, 'bias': bias_num}
        datos_df = {
            "dem_max_h":  dem_max_h,  "dem_max_v":  dem_max_v,
            "dem_min_h":  dem_min_h,  "dem_min_v":  dem_min_v,
            "frec_max_h": frec_max_h, "frec_max_v": frec_max_v,
            "frec_min_h": frec_min_h, "frec_min_v": frec_min_v,
        }

        st.markdown("---")

        # ── Generación de archivos ────────────────────────────────────────────
        # El df limpio (sin col checkbox) que se usa para generar Excel/XML
        df_final_clean = df_editado[df_editado["🗑️"] == False][["Hora", "Descripción"]].reset_index(drop=True)

        col_excel, col_xml = st.columns(2)

        with col_excel:
            if st.button("✏️ Generar Excel EOR", type="primary"):
                if os.path.exists("plantilla_pan.xlsm"):
                    try:
                        wb = openpyxl.load_workbook("plantilla_pan.xlsm", keep_vba=True)
                        ws = wb.active
                        border_style = Border(
                            left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin')
                        )
                        font_eor = Font(name='Arial', size=10)
                        fila_inicio   = 30
                        CHARS_POR_LINEA = 86
                        ALTO_POR_LINEA  = 13.5
                        ALTO_MINIMO     = 15.0

                        # Calcular cuántas filas necesitamos (evento + separador por cada uno)
                        n_eventos = len(df_final_clean)
                        filas_necesarias = max(1, n_eventos * 2 - 1)  # evento + separador, sin separador final
                        fila_fin_area = fila_inicio + filas_necesarias - 1

                        # Si necesitamos más filas que las que hay en la plantilla,
                        # insertar las filas adicionales ANTES de escribir
                        FILAS_PLANTILLA = 21  # originalmente filas 30-50
                        if filas_necesarias > FILAS_PLANTILLA:
                            filas_extra = filas_necesarias - FILAS_PLANTILLA
                            ws.insert_rows(fila_inicio + FILAS_PLANTILLA, filas_extra)

                        # ── Paso 1: desmerge y limpiar TODO el área ──
                        for m in list(ws.merged_cells.ranges):
                            if fila_inicio <= m.min_row <= fila_fin_area:
                                ws.unmerge_cells(str(m))
                        for fila in range(fila_inicio, fila_fin_area + 1):
                            for col in range(1, 8):
                                c = ws.cell(row=fila, column=col)
                                c.value = None
                                c.border = border_style
                                c.alignment = Alignment()
                                c.font = font_eor
                            ws.row_dimensions[fila].height = 15
                            ws.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=7)
                            ws.cell(row=fila, column=2).alignment = Alignment(horizontal='center', vertical='center')

                        # ── Paso 2: escribir eventos, uno por fila (sin límite) ──
                        f = fila_inicio
                        for i, row in df_final_clean.iterrows():
                            texto = str(row["Descripción"])

                            # Col A: hora
                            ca = ws.cell(row=f, column=1)
                            ca.value = str(row["Hora"])
                            ca.alignment = Alignment(horizontal='center', vertical='top')
                            ca.border = border_style
                            ca.font = font_eor

                            # Col B:G mergeadas: descripcion
                            ws.merge_cells(start_row=f, start_column=2, end_row=f, end_column=7)
                            cb = ws.cell(row=f, column=2)
                            cb.value = texto
                            cb.alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')
                            cb.border = border_style
                            cb.font = font_eor

                            for col in range(3, 8):
                                c = ws.cell(row=f, column=col)
                                c.border = border_style
                                c.font = font_eor

                            lineas = sum(
                                max(1, -(-len(p) // CHARS_POR_LINEA))
                                for p in texto.split('\n')
                            )
                            ws.row_dimensions[f].height = max(ALTO_MINIMO, lineas * ALTO_POR_LINEA)
                            f += 1

                            # Fila separadora (excepto después del último evento)
                            if i < len(df_final_clean) - 1:
                                ws.row_dimensions[f].height = 20
                                for col in range(1, 8):
                                    c = ws.cell(row=f, column=col)
                                    c.border = border_style
                                    c.font = font_eor
                                ws.merge_cells(start_row=f, start_column=2, end_row=f, end_column=7)
                                f += 1
                        ahora = datetime.now() - timedelta(days=1)
                        nombre_archivo = f"PAN-Diario_{ahora.strftime('%d')}_{obtener_mes_nombre(ahora)}.xlsm"
                        output = BytesIO()
                        wb.save(output)
                        # ── Guardar en session_state para descarga persistente ──
                        st.session_state["eor_excel_bytes"]  = output.getvalue()
                        st.session_state["eor_excel_nombre"] = nombre_archivo
                        st.success("✅ Excel generado. Haz clic en el botón de descarga.")
                    except Exception as e:
                        st.error(f"🩻 Error al escribir Excel: {e}")
                else:
                    st.error("No se encontró 'plantilla_pan.xlsm'.")

            # ── Botón de descarga Excel (persistente, fuera del callback) ──
            if st.session_state.get("eor_excel_bytes"):
                st.download_button(
                    f"⏬ Descargar {st.session_state['eor_excel_nombre']}",
                    st.session_state["eor_excel_bytes"],
                    st.session_state["eor_excel_nombre"],
                    mime="application/vnd.ms-excel.sheet.macroEnabled.12",
                    key="eor_excel_dl",
                )

        with col_xml:
            # Validar campos MIN/MAX para habilitar generación
            _campos_vacios = [
                nombre for nombre, val in [
                    ("Demanda Máxima — Valor",         dem_max_v),
                    ("Demanda Máxima — Hora",          dem_max_h),
                    ("Demanda Mínima — Valor",         dem_min_v),
                    ("Demanda Mínima — Hora",          dem_min_h),
                    ("Frecuencia Máxima — Valor",      frec_max_v),
                    ("Frecuencia Máxima — Hora",       frec_max_h),
                    ("Frecuencia Mínima — Valor",      frec_min_v),
                    ("Frecuencia Mínima — Hora",       frec_min_h),
                ] if not str(val).strip()
            ]
            if st.button("✏️ Generar XML", disabled=bool(_campos_vacios)):
                if _campos_vacios:
                    st.warning(
                        "⚠️ Debe llenar los valores MIN/MAX antes de exportar el XML.\n\n"
                        "**Campos pendientes:** " + " · ".join(_campos_vacios)
                    )
                else:
                    try:
                        # Usar los datos RTR con las ediciones del usuario aplicadas.
                        # rtr_resultado_editado se actualiza en cada render del data_editor RTR.
                        # Si el modulo RTR no se abrio en esta sesion, caer al original.
                        df_rtr_final = st.session_state.get(
                            "rtr_resultado_editado",
                            st.session_state.get("rtr_resultado", pd.DataFrame())
                        )
                        xml_data = generar_xml_oficial_eor(
                            df_final_clean, df_rtr_final, datos_op, datos_df
                        )
                        fecha_xml = (datetime.now() - timedelta(days=1)).strftime("%Y%m%d")
                        # ── Guardar en session_state para descarga persistente ──
                        st.session_state["eor_xml_bytes"]  = xml_data
                        st.session_state["eor_xml_nombre"] = f"PAN-INF{fecha_xml}.xml"
                        st.success("✅ XML generado. Haz clic en el botón de descarga.")
                    except Exception as e:
                        st.error(f"Error XML: {e}")
            elif _campos_vacios:
                st.caption(f"⚠️ Faltan: {' · '.join(_campos_vacios)}")

            # ── Botón de descarga XML (persistente, fuera del callback) ──
            if st.session_state.get("eor_xml_bytes"):
                st.download_button(
                    "📥 Descargar XML",
                    st.session_state["eor_xml_bytes"],
                    st.session_state["eor_xml_nombre"],
                    mime="application/xml",
                    key="eor_xml_dl",
                )

    except Exception as e:
        st.error(f"Error procesando EOR: {e}")

# --- 7. PLANILLA DEL COLABORADOR ---

_PLANILLA_NUM_EMPLEADOS = {
    'A': 27034, 'B': 27203, 'C': 26888, 'E': 27576, 'F': 26954,
    'G': 27712, 'H': 26510, 'J': 5318,  'L': 27962, 'M': 26842,
    'N': 27262, 'O': 26951, 'Q': 28045, 'S': 27033, 'T': 27705,
    'W': 27863, 'Y': 27836, 'Z': 27714,
}
_PLANILLA_CEDULAS = {
    'A': '8-791-1066', 'B': '9-708-1261', 'C': '2-706-1039',
    'E': '6-715-261',  'F': '8-270-462',  'G': '4-772-2491',
    'H': '8-369-717',  'J': '8-226-2448', 
    'L': '8-927-2400', 'M': '8-342-668',  'N': '8-813-181',
    'O': '6-702-948',  'Q': '4-756-1872', 
    'S': '9-713-81',   'T': '8-890-999',  'W': '8-832-1925',
    'X': '8-980-2399', 'Y': '4-770-190',  'Z': '4-768-675',
}
_DIAS_ES_PL   = ['lunes','martes','miércoles','jueves','viernes','sábado','domingo']
_MESES_ES_PL  = {1:'Enero',2:'Febrero',3:'Marzo',4:'Abril',5:'Mayo',6:'Junio',
                 7:'Julio',8:'Agosto',9:'Septiembre',10:'Octubre',11:'Noviembre',12:'Diciembre'}
_MESES_MIN_PL = {1:'enero',2:'febrero',3:'marzo',4:'abril',5:'mayo',6:'junio',
                 7:'julio',8:'agosto',9:'septiembre',10:'octubre',11:'noviembre',12:'diciembre'}
_TURNO_TIEMPOS = {'23-07':('23:00','07:00'),'07-15':('07:00','15:00'),'15-23':('15:00','23:00')}
_NEXT_TURNO_DOBLE = {
    '23-07': ('07-15', 8),
    '07-15': ('15-23', 9),
    '15-23': ('23-07', 9),
}

_FERIADOS_FIJOS_PL = {
    (1,  1): "Año Nuevo",         (1,  9): "Día de los Mártires",
    (5,  1): "Día del Trabajador",
    (11, 3): "Fiestas Patrias",   (11, 5):  "Fiestas Patrias",
    (11,10): "Fiestas Patrias",   (11,28):  "Fiestas Patrias",
    (12, 8): "Día de las Madres", (12,20): "Día de Duelo Nacional",
    (12,25): "Navidad",
}
_FERIADOS_MOVIBLES_PL = [
    (2025, 3,  4, "Martes de Carnaval"), (2025, 4, 18, "Viernes Santo"),
    (2026, 2, 17, "Martes de Carnaval"), (2026, 4,  3, "Viernes Santo"),
]

def _nombre_feriado(fecha):
    fijo = _FERIADOS_FIJOS_PL.get((fecha.month, fecha.day))
    if fijo:
        return fijo
    for anio, mes, dia, nombre in _FERIADOS_MOVIBLES_PL:
        if fecha.year == anio and fecha.month == mes and fecha.day == dia:
            return nombre
    return None

def _es_feriado_fecha(fecha):
    return _nombre_feriado(fecha) is not None


def _fetch_horario_planilla(nombre_archivo):
    try:
        import base64
        token = __import__("os").environ.get("GITHUB_TOKEN", "")
        headers = {"Authorization": f"token {token}", "Accept": "application/vnd.github+json"}
        api_url = f"https://api.github.com/repos/LilisophiG24/Reportes-del-EOR/contents/{nombre_archivo}"
        r = requests.get(api_url, headers=headers, timeout=15)
        if r.status_code == 200:
            data = r.json()
            if isinstance(data, dict) and data.get("encoding") == "base64":
                return base64.b64decode(data["content"]), None
            return r.content, None
        return None, f"HTTP {r.status_code} al descargar {nombre_archivo}"
    except Exception as e:
        return None, str(e)


def _leer_mes_horario(horario_bytes):
    wb = openpyxl.load_workbook(BytesIO(horario_bytes), read_only=True)
    ws = wb['LISTA']
    for c in range(1, 50):
        v = ws.cell(5, c).value
        if hasattr(v, 'date'):
            wb.close()
            return v.month, v.year
    wb.close()
    return None, None


def _leer_empleados_horario(horario_bytes):
    VALIDOS = {'A','B','C','E','F','G','H','J','L','M','N','O','Q','S','T','W','X','Y','Z'}
    wb = openpyxl.load_workbook(BytesIO(horario_bytes))
    ws = wb['LISTA']
    empleados, vistos = [], set()
    for r in range(8, 36):
        cod = ws.cell(r, 2).value
        nom = ws.cell(r, 3).value
        if cod and str(cod).strip() in VALIDOS and nom and str(nom).strip():
            cod = str(cod).strip()
            if cod in vistos:
                continue
            vistos.add(cod)
            empleados.append({
                'cod':          cod,
                'nombre':       str(nom).strip().title(),
                'num_empleado': _PLANILLA_NUM_EMPLEADOS.get(cod, ''),
                'cedula':       _PLANILLA_CEDULAS.get(cod, ''),
            })
    wb.close()
    return sorted(empleados, key=lambda x: x['cod'])


def _leer_rotacion_empleado(horario_bytes, cod, fecha_inicio, fecha_fin):
    """Lee rotación. Extiende 1 día para capturar turno 23-07 del día 16/1."""
    VALIDOS = {'A','B','C','E','F','G','H','J','L','M','N','O','Q','S','T','W','X','Y','Z'}
    wb = openpyxl.load_workbook(BytesIO(horario_bytes))
    ws = wb['LISTA']
    date_to_col = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(5, c).value
        if hasattr(v, 'date'):
            date_to_col[v.date()] = c

    fecha_fin_ext = fecha_fin + timedelta(days=1)
    dias = []
    fecha = fecha_inicio
    while fecha <= fecha_fin_ext:
        col = date_to_col.get(fecha.date())
        if col is None:
            fecha += timedelta(days=1)
            continue
        turno, es_primer_libre = None, False
        for t_nombre, t_range in [('23-07', range(8,15)),
                                   ('07-15', range(15,28)),
                                   ('15-23', range(28,35))]:
            for ri in t_range:
                v = ws.cell(ri, col).value
                if v:
                    partes = [p.strip() for p in str(v).strip().split('/')]
                    if cod in partes and all(p in VALIDOS for p in partes):
                        turno = t_nombre
                        try:
                            fg = ws.cell(ri, col).fill.fgColor
                            if fg and fg.type == 'rgb' and fg.rgb == 'FFFF0000':
                                es_primer_libre = True
                        except Exception:
                            pass
                        break
            if turno:
                break
        if fecha > fecha_fin and turno != '23-07':
            fecha += timedelta(days=1)
            continue
        es_libre = (turno is None) and any(
            ws.cell(ri, col).value and str(ws.cell(ri, col).value).strip() == cod
            for ri in range(45, 56)
        )
        nombre_fer = _nombre_feriado(fecha)
        dias.append({
            'fecha': fecha, 'turno': turno,
            'es_libre': es_libre, 'es_primer_libre': es_primer_libre,
            'es_domingo': fecha.weekday() == 6,
            'es_feriado': nombre_fer is not None,
            'nombre_feriado': nombre_fer or '',
            'libre_orden': 0, 'credito_festivo': None,
            'credito_festivo_trabajado': False,
        })
        fecha += timedelta(days=1)
    wb.close()

    for i, d in enumerate(dias):
        if d['es_libre']:
            d['libre_orden'] = 2 if (i > 0 and dias[i-1]['es_libre']) else 1

    i = 0
    while i < len(dias):
        if dias[i]['es_libre']:
            festivo_orden = None
            while i < len(dias) and dias[i]['es_libre']:
                if dias[i]['es_feriado']:
                    festivo_orden = dias[i]['libre_orden']
                i += 1
            if festivo_orden is not None:
                j = i
                while j < len(dias) and dias[j]['turno'] is None:
                    j += 1
                if j < len(dias) and dias[j]['turno'] is not None:
                    dias[j]['credito_festivo'] = festivo_orden
        else:
            i += 1

    for i, d in enumerate(dias):
        if not d['es_libre'] and d['turno'] and d['es_feriado']:
            if i + 1 < len(dias):
                nd = dias[i + 1]
                if not nd['es_libre'] and nd['turno']:
                    nd['credito_festivo_trabajado'] = True
    return dias


def _add_minutes_to_time_str(time_str, extra_min):
    """Suma minutos a un string 'HH:MM' y retorna nuevo string 'HH:MM'."""
    h, m = map(int, time_str.split(':'))
    total = h * 60 + m + extra_min
    return f"{total // 60:02d}:{total % 60:02d}"


def _llenar_gasto(ws_gasto, nombre, num_empleado, cedula,
                  fecha, monto_fila18, monto_fila23,
                  descripcion, ord_ini, ord_fin, ext_ini, ext_fin):
    """Escribe los campos del formulario Gasto de Alimentación.
    Los montos van en col 17 (Q). El total de la fila 28 lo calcula Excel.
    """
    _MESES_G = {1:'Enero',2:'Febrero',3:'Marzo',4:'Abril',5:'Mayo',6:'Junio',
                7:'Julio',8:'Agosto',9:'Septiembre',10:'Octubre',11:'Noviembre',12:'Diciembre'}
    ws_gasto.cell(11, 5).value  = nombre
    ws_gasto.cell(11, 14).value = cedula or ""
    ws_gasto.cell(11, 26).value = str(num_empleado)
    if monto_fila18:
        ws_gasto.cell(18, 17).value = monto_fila18
    if monto_fila23:
        ws_gasto.cell(23, 17).value = monto_fila23
    ws_gasto.cell(30, 2).value  = descripcion

    # ── Bordes en el cuadrante de Descripcion del Trabajo (filas 30-32, cols B-Z) ──
    _thin_side = Side(style='thin')
    _borde_all = Border(left=_thin_side, right=_thin_side, top=_thin_side, bottom=_thin_side)
    for _r in range(30, 33):
        for _c in range(2, 27):
            _cel = ws_gasto.cell(row=_r, column=_c)
            if not isinstance(_cel, MergedCell):
                _cel.border = _borde_all

    ws_gasto.cell(35, 8).value  = ord_ini
    ws_gasto.cell(35, 11).value = ord_fin
    if ext_ini:
        ws_gasto.cell(35, 25).value = ext_ini
    if ext_fin:
        ws_gasto.cell(35, 30).value = ext_fin
    ws_gasto.cell(36, 9).value  = fecha.day
    ws_gasto.cell(36, 11).value = _MESES_G[fecha.month]
    ws_gasto.cell(36, 14).value = fecha.year


def generar_planilla_colaborador(horario_bytes, cod, nombre, num_empleado,
                                  periodo_inicio, periodo_fin,
                                  eventos_extra=None,
                                  cedula=""):
    """
    eventos_extra: lista de dicts con:
      tipo='doble':      {tipo, fecha, justificacion}
      tipo='extendida':  {tipo, fecha, horas, minutos, justificacion}
      tipo='anticipada': {tipo, fecha, hora_anticipada (HH:MM), justificacion}
    """
    from datetime import time as dtime

    PLANTILLA = "planilla.xlsx"
    if not os.path.exists(PLANTILLA):
        return None, f"No se encontró '{PLANTILLA}' en la carpeta de la app."

    dias = _leer_rotacion_empleado(horario_bytes, cod, periodo_inicio, periodo_fin)
    dias_map = {d['fecha'].date(): d for d in dias}

    wb = openpyxl.load_workbook(PLANTILLA)
    ws = wb['Horas Extras']

    ws.cell(4, 7).value  = datetime.now()
    ws.cell(4, 15).value = periodo_inicio
    ws.cell(4, 19).value = periodo_fin
    ws.cell(6, 7).value  = nombre
    ws.cell(6, 16).value = f"04-50-01-01-{num_empleado}" if num_empleado else nombre

    # Separar eventos por tipo
    evs_doble     = [ev for ev in (eventos_extra or []) if ev['tipo'] == 'doble']
    evs_cambio    = [ev for ev in (eventos_extra or []) if ev['tipo'] == 'cambio_turno']
    evs_extra_map = {ev['fecha']: ev for ev in (eventos_extra or [])
                     if ev['tipo'] in ('extendida', 'anticipada')}

    # Fechas manejadas por bloques doble (excluir de rotación normal)
    fechas_doble = set()
    for ev_d in evs_doble:
        fd = ev_d['fecha']
        fechas_doble.add(fd)
        sig = (datetime.combine(fd, datetime.min.time()) + timedelta(days=1)).date()
        d_sig = dias_map.get(sig)
        if d_sig and d_sig['turno']:
            fechas_doble.add(sig)

    entries = []

    def _t(total_min):
        return dtime(total_min // 60, total_min % 60)

    def _min_from_str(t_str):
        h, m = map(int, t_str.split(':'))
        return h * 60 + m

    def _agregar(fecha_dt, p_in, p_fin, r_in, r_fin, h12_min, h16_min, h18_min, h20_min, just):
        total = h12_min + h16_min + h18_min + h20_min
        if total == 0:
            return
        entries.append({
            'fecha':   fecha_dt, 'dia': _DIAS_ES_PL[fecha_dt.weekday()],
            'p_in':    p_in, 'p_fin': p_fin, 'r_in': r_in, 'r_fin': r_fin,
            'h12': h12_min, 'h16': h16_min, 'h18': h18_min, 'h20': h20_min,
            'total': total, 'just': just,
        })

    for d in dias:
        if d['turno'] is None:
            continue
        if d['fecha'].date() in fechas_doble:
            continue

        fecha   = d['fecha']
        turno   = d['turno']
        noc     = turno in ('23-07', '15-23')
        credito = d['credito_festivo']
        cft     = d['credito_festivo_trabajado']
        es_pl   = d.get('es_primer_libre', False)
        ini, fin = _TURNO_TIEMPOS[turno]
        ev = evs_extra_map.get(fecha.date())

        # Caso 1: Sexto día (celda roja) — no suma recargo domingo
        if es_pl:
            h = (9 if noc else 8) * 60
            just_parts = ["Laborando en primer dia libre"]
            if ev and ev['tipo'] == 'extendida':
                ext_min = ev['horas'] * 60 + ev['minutos']
                h += ext_min
                just_parts.append(ev['justificacion'])
                real_fin = _add_minutes_to_time_str(fin, ext_min)
                _agregar(fecha, ini, fin, fin, real_fin, h, 0, 0, 0, " + ".join(just_parts))
            else:
                _agregar(fecha, ini, fin, ini, fin, h, 0, 0, 0, " + ".join(just_parts))
            continue

        # Caso 2: Festivo trabajado
        if d['es_feriado']:
            h = (9 if noc else 8) * 60
            just_parts = [f"Día Nacional: {fecha.day} de {_MESES_MIN_PL[fecha.month]}"]
            if noc:
                just_parts.append("Jornada nocturna")
            if ev and ev['tipo'] == 'extendida':
                ext_min = ev['horas'] * 60 + ev['minutos']
                h += ext_min
                just_parts.append(ev['justificacion'])
                real_fin = _add_minutes_to_time_str(fin, ext_min)
                _agregar(fecha, ini, fin, fin, real_fin, h, 0, 0, 0, " + ".join(just_parts))
            else:
                _agregar(fecha, ini, fin, ini, fin, h, 0, 0, 0, " + ".join(just_parts))
            continue

        # Caso 3/4: crédito, nocturna, festivo trabajado, domingo
        h12, h20, just_parts = 0, 0, []
        if credito:
            h12 += 4 * 60
            # credito indica en qué libre cayó el festivo.
            # El trabajador entra en el libre CONTRARIO:
            #   festivo en 1er libre → entra en 2do libre
            #   festivo en 2do libre → entra en 1er libre
            orden = "2do" if credito == 1 else "1er"
            just_parts.append(f"4 Hrs por coincidir con su {orden} dia libre")
        if noc:
            h12 += 60
            just_parts.append("Jornada nocturna")
        if cft:
            h20 = 4 * 60
            just_parts.insert(0, "4 horas compensatorias por haber laborado dia libre nacional")
        if d['es_domingo'] and not es_pl and not credito and not cft:
            h20 += 4 * 60
            just_parts.insert(0, "Domingo")

        # Evento extra en este día
        if ev:
            if ev['tipo'] == 'extendida':
                ext_min = ev['horas'] * 60 + ev['minutos']
                h12 += ext_min
                just_parts.append(ev['justificacion'])
                real_fin = _add_minutes_to_time_str(fin, ext_min)
                _agregar(fecha, ini, fin, fin, real_fin, h12, 0, 0, h20, " + ".join(just_parts))
            elif ev['tipo'] == 'anticipada':
                hora_ant = ev['hora_anticipada']   # "HH:MM"
                ant_min  = _min_from_str(ini) - _min_from_str(hora_ant)
                h12 += max(ant_min, 0)
                just_parts.append(ev['justificacion'])
                # programado = anticipada→turno normal; realizado = turno normal completo
                _agregar(fecha, hora_ant, ini, ini, fin, h12, 0, 0, h20, " + ".join(just_parts))
            continue

        _agregar(fecha, ini, fin, ini, fin, h12, 0, 0, h20, " + ".join(just_parts))

    # ── Doble jornada ─────────────────────────────────────────────────────────
    # ── Dobles jornadas ──────────────────────────────────────────────────────
    for ev_d in evs_doble:
        fd        = ev_d['fecha']
        d_doble   = dias_map.get(fd)
        turno_reg = d_doble['turno'] if d_doble and d_doble['turno'] else '07-15'
        noc_reg   = turno_reg in ('23-07', '15-23')
        ini_reg, fin_reg = _TURNO_TIEMPOS[turno_reg]
        turno_extra, h_extra = _NEXT_TURNO_DOBLE[turno_reg]
        ini_ext, fin_ext = _TURNO_TIEMPOS[turno_extra]
        h_total = ((1 if noc_reg else 0) + h_extra) * 60
        just_d  = ev_d.get('justificacion') or "Doble jornada"
        if noc_reg:
            just_d += " + Jornada nocturna"
        fecha_doble_dt = datetime.combine(fd, datetime.min.time())
        _agregar(fecha_doble_dt, ini_reg, fin_reg, ini_ext, fin_ext, h_total, 0, 0, 0, just_d)
        sig = (fecha_doble_dt + timedelta(days=1)).date()
        d_sig = dias_map.get(sig)
        if d_sig and d_sig['turno']:
            noc_sig  = d_sig['turno'] in ('23-07', '15-23')
            ini_s, fin_s = _TURNO_TIEMPOS[d_sig['turno']]
            just_sig = "4 horas compensatorias por doble jornada" + (" + Jornada nocturna" if noc_sig else "")
            _agregar(d_sig['fecha'], ini_s, fin_s, ini_s, fin_s,
                     (60 if noc_sig else 0), 0, 0, 4*60, just_sig)

    # ── Cambio de turno ──────────────────────────────────────────────────────────
    # Si el turno al que se cambió es nocturno (15-23 o 23-07): genera 1 hr recargo.
    # Si es 07-15: no genera horas extra.
    for ev_c in evs_cambio:
        fc         = ev_c['fecha']
        turno_nuevo = ev_c['turno_nuevo']
        noc_nuevo  = turno_nuevo in ('23-07', '15-23')
        if not noc_nuevo:
            continue   # 07-15 → sin extra
        ini_c, fin_c = _TURNO_TIEMPOS[turno_nuevo]
        just_c = ev_c.get('justificacion') or f"Cambio de turno al {turno_nuevo}"
        just_c += " + Recargo nocturno"
        _agregar(datetime.combine(fc, datetime.min.time()),
                 ini_c, fin_c, ini_c, fin_c,
                 60, 0, 0, 0, just_c)   # 1 hr = 60 min recargo nocturno

    # Eventos extra en días sin rotación registrada
    fechas_ya = {e['fecha'].date() for e in entries}
    for ev in (eventos_extra or []):
        if ev['tipo'] not in ('extendida', 'anticipada'):
            continue
        fd = ev['fecha']
        if fd not in fechas_ya and fd in dias_map:
            d = dias_map[fd]
            if not d['turno']:
                continue
            ini, fin = _TURNO_TIEMPOS[d['turno']]
            noc = d['turno'] in ('23-07', '15-23')
            if ev['tipo'] == 'extendida':
                ext_min = ev['horas'] * 60 + ev['minutos']
                real_fin = _add_minutes_to_time_str(fin, ext_min)
                _agregar(datetime.combine(fd, datetime.min.time()),
                         ini, fin, fin, real_fin, ext_min, 0, 0, 0, ev['justificacion'])
            elif ev['tipo'] == 'anticipada':
                hora_ant = ev['hora_anticipada']
                ant_min  = _min_from_str(ini) - _min_from_str(hora_ant)
                noc_extra = 60 if noc else 0
                _agregar(datetime.combine(fd, datetime.min.time()),
                         hora_ant, ini, ini, fin,
                         max(ant_min, 0) + noc_extra, 0, 0, 0, ev['justificacion'])

    entries.sort(key=lambda x: x['fecha'])

    for idx, e in enumerate(entries[:13]):
        r_hdr = 13 + idx * 2
        r_tot = r_hdr + 1
        ws.cell(r_hdr, 1).value  = 0
        ws.cell(r_hdr, 2).value  = e['dia']
        ws.cell(r_hdr, 3).value  = e['fecha']
        ws.cell(r_hdr, 12).value = e['p_in']
        ws.cell(r_hdr, 13).value = e['p_fin']
        ws.cell(r_hdr, 14).value = e['r_in']
        ws.cell(r_hdr, 15).value = e['r_fin']
        ws.cell(r_hdr, 27).value = e['just']
        ws.cell(r_tot, 3).value  = 'TOTAL'
        ws.cell(r_tot, 20).value = None
        ws.cell(r_tot, 20).number_format = 'h:mm'
        total_min = 0
        for col, key in [(12,'h12'),(16,'h16'),(18,'h18'),(20,'h20')]:
            m = e[key]
            if m > 0:
                cell = ws.cell(r_tot, col)
                cell.value         = _t(m)
                cell.number_format = 'h:mm'
                total_min += m
        t_cell = ws.cell(r_tot, 24)
        t_cell.value         = _t(total_min)
        t_cell.number_format = 'h:mm'

    # ── Gasto de Alimentación: un sheet por evento cualificante ───────────────
    _tini = {t: v[0] for t, v in _TURNO_TIEMPOS.items()}
    _tfin = {t: v[1] for t, v in _TURNO_TIEMPOS.items()}

    # Recopilar todos los eventos que generan Gasto
    gasto_events = []

    # Dobles
    for ev_d in evs_doble:
        fd  = ev_d['fecha']
        d_d = dias_map.get(fd)
        tr  = d_d['turno'] if d_d and d_d['turno'] else '07-15'
        te, _ = _NEXT_TURNO_DOBLE[tr]
        e_ent = next((e for e in entries if e['fecha'].date() == fd), None)
        desc  = e_ent['just'] if e_ent else (ev_d.get('justificacion') or "Doble jornada")
        gasto_events.append({
            'fecha': datetime.combine(fd, datetime.min.time()),
            'f18': None, 'f23': 10, 'desc': desc,
            'ord_ini': _tini[tr], 'ord_fin': _tfin[tr],
            'ext_ini': _tini[te], 'ext_fin': _tfin[te],
        })

    # Entrada anticipada (3+ horas)
    for ev in (eventos_extra or []):
        if ev['tipo'] == 'anticipada':
            fd   = ev['fecha']
            d_ev = dias_map.get(fd)
            if not d_ev or not d_ev['turno']:
                continue
            tr       = d_ev['turno']
            hora_ant = ev['hora_anticipada']
            mins_ant = _min_from_str(_tini[tr]) - _min_from_str(hora_ant)
            if mins_ant >= 180:
                gasto_events.append({
                    'fecha': datetime.combine(fd, datetime.min.time()),
                    'f18': 6, 'f23': None, 'desc': ev['justificacion'],
                    'ord_ini': hora_ant, 'ord_fin': _tini[tr],
                    'ext_ini': _tini[tr], 'ext_fin': _tfin[tr],
                })

    # Primer día libre
    for e in entries:
        fd  = e['fecha'].date()
        d_e = dias_map.get(fd)
        if d_e and d_e.get('es_primer_libre'):
            tr = d_e['turno']
            gasto_events.append({
                'fecha': e['fecha'], 'f18': 6, 'f23': None, 'desc': e['just'],
                'ord_ini': "Libre", 'ord_fin': "Libre",
                'ext_ini': _tini[tr], 'ext_fin': _tfin[tr],
            })

    # Día nacional/festivo trabajado
    for e in entries:
        fd  = e['fecha'].date()
        d_e = dias_map.get(fd)
        if d_e and d_e.get('es_feriado') and not d_e.get('es_primer_libre'):
            tr = d_e['turno']
            gasto_events.append({
                'fecha': e['fecha'], 'f18': 6, 'f23': None, 'desc': e['just'],
                'ord_ini': _tini[tr], 'ord_fin': _tfin[tr],
                'ext_ini': None, 'ext_fin': None,
            })

    if gasto_events:
        ws_gasto_orig = wb['Gasto']
        # Copiar sheet en blanco para cada evento adicional ANTES de llenar ninguno
        gasto_sheets = [ws_gasto_orig]
        for i in range(1, len(gasto_events)):
            new_ws = wb.copy_worksheet(ws_gasto_orig)
            new_ws.title = f'Gasto {i+1}'
            gasto_sheets.append(new_ws)
        ws_gasto_orig.title = 'Gasto 1'
        # Llenar cada sheet
        for ws_g, ge in zip(gasto_sheets, gasto_events):
            ws_g.sheet_view.showGridLines = False
            _llenar_gasto(ws_g, nombre, num_empleado, cedula,
                          ge['fecha'], ge['f18'], ge['f23'], ge['desc'],
                          ge['ord_ini'], ge['ord_fin'], ge['ext_ini'], ge['ext_fin'])

    output = BytesIO()
    wb.save(output)
    return output.getvalue(), None


# --- 8. INFORME RHH ---

_COLOR_DESC = {
    'FFFFFF00': 'Media Hora',
    'FFFF0000': 'Primer día libre',
    'FF30803D': 'Compensatorio',
    'FF074F69': 'Compensatorio',
    'FF0066FF': 'Libre (Disponible)',
    'FF8497B0': 'Libre (Feriado)',
}
# Filas Excel 36-38 → Vacaciones (chocolate, formato condicional)
_FILAS_VACACIONES = {36, 37, 38}

DESCRIPCIONES_AUSENCIA = [
    "Calamidad doméstica", "Compensatorio", "Licencia Sindical", "Duelo",
    "Fiesta", "Gira o Seminario", "Incapacidad médica", "Libre",
    "Libre (Disponible)", "Libre (Feriado)", "Media Hora", "Permiso", "Vacaciones",
    "Primer día libre", "Asignación especial",
]


def _leer_turnos_horario(horario_bytes, fecha):
    """
    Lee el horario (bytes) y extrae códigos + descripción por turno para la fecha.
    Returns dict con t1, t2, t3 (listas de códigos) y
    ausentes (lista de dicts {Código, Descripción}).
    """
    VALIDOS = {'A','B','C','E','F','G','H','J','L','M','N','O','Q','S','T','W','X','Y','Z'}

    wb = openpyxl.load_workbook(BytesIO(horario_bytes))   # con colores → NO data_only
    ws = wb['LISTA']

    # Fila 5: buscar columna del día
    fila_fechas = [ws.cell(5, c+1).value for c in range(ws.max_column)]
    col_dia = None
    for i, v in enumerate(fila_fechas):
        if hasattr(v, 'date') and v.date() == fecha.date():
            col_dia = i + 1   # 1-indexed para openpyxl
            break

    if col_dia is None:
        wb.close()
        return None

    def _codigo(ri):
        v = ws.cell(row=ri, column=col_dia).value
        return str(v).strip() if v and str(v).strip() in VALIDOS else None

    def _desc_ausencia(ri):
        """Detecta la descripción de ausencia por fila (vacaciones) y luego por color de celda."""
        if ri in _FILAS_VACACIONES:
            return 'Vacaciones'
        cell = ws.cell(row=ri, column=col_dia)
        fg = cell.fill.fgColor if cell.fill else None
        if fg and fg.type == 'rgb' and fg.rgb in _COLOR_DESC:
            return _COLOR_DESC[fg.rgb]
        return None   # sin color reconocido → no incluir en ausentes

    # VARIOS (rows 36-43): extraer PRIMERO para detectar Media Hora
    varios_raw = []
    for ri in range(36, 44):
        cod = _codigo(ri)
        if cod:
            desc = _desc_ausencia(ri)
            if desc:   # solo incluir si tiene color reconocido (no None)
                varios_raw.append({'Código': cod, 'Descripción': desc})

    # Códigos con Media Hora en VARIOS → no deben aparecer en los turnos
    media_hora_codigos = {item['Código'] for item in varios_raw if item['Descripción'] == 'Media Hora'}

    # Helper robusto: detecta celda roja (FFFF0000)
    def _es_rojo(ri):
        try:
            fg = ws.cell(row=ri, column=col_dia).fill.fgColor
            return fg is not None and fg.type == 'rgb' and fg.rgb == 'FFFF0000'
        except Exception:
            return False

    # Helper: valor raw de celda (puede incluir splits como J/G)
    def _raw(ri):
        v = ws.cell(row=ri, column=col_dia).value
        return str(v).strip() if v else None

    # Primer día libre: letra sola en rojo
    primer_dia_libre = {}
    for ri in range(8, 34):
        cod = _codigo(ri)
        if cod and _es_rojo(ri) and cod not in media_hora_codigos:
            primer_dia_libre[cod] = ri

    # Turno dividido: celda con X/Y (puede estar en rojo)
    turno_dividido = []
    for ri in range(8, 34):
        raw = _raw(ri)
        if raw and '/' in raw:
            partes = [p.strip() for p in raw.split('/')]
            if len(partes) == 2 and all(p in VALIDOS for p in partes):
                turno_dividido.append({
                    'primero': partes[0],
                    'segundo': partes[1],
                    'es_rojo': _es_rojo(ri),
                })

    # Turnos principales: excluir quien está en Media Hora
    t1 = [c for c in [_codigo(i) for i in range(8, 15)]  if c and c not in media_hora_codigos]
    t2 = [c for c in [_codigo(i) for i in range(15, 28)] if c and c not in media_hora_codigos]
    t3 = [c for c in [_codigo(i) for i in range(28, 34)] if c and c not in media_hora_codigos]

    # Y en turno → Jose Samudio asistencia DEN
    tiene_y = 'Y' in t1 or 'Y' in t2 or 'Y' in t3

    # LIBRES (rows 45-55): Libre por defecto; Compensatorio si tiene verde oscuro
    libres_raw = []
    for ri in range(45, 56):
        cod = _codigo(ri)
        if cod:
            desc = _desc_ausencia(ri)
            if desc not in ('Compensatorio', 'Asignación especial', 'Media Hora'):
                desc = 'Libre'
            libres_raw.append({'Código': cod, 'Descripción': desc})

    wb.close()

    # Merge ausentes sin duplicados
    seen = set()
    ausentes = []
    for item in varios_raw + libres_raw:
        if item['Código'] not in seen:
            seen.add(item['Código'])
            ausentes.append(item)

    # ── Sábado y domingo: empleados sin turno asignado → Libre automático ──────
    # _codigo(ri) lee col_dia (puede estar vacía el fin de semana).
    # Se lee la col B del horario (código permanente, siempre relleno).
    if fecha.weekday() in (5, 6):   # 5=sábado, 6=domingo
        en_turno    = set(t1 + t2 + t3)
        en_ausentes = seen  # ya construido arriba
        todos_en_horario = set()
        for ri in range(8, 35):
            cod_perm = ws.cell(row=ri, column=2).value   # col B = código permanente
            if cod_perm and str(cod_perm).strip() in VALIDOS:
                todos_en_horario.add(str(cod_perm).strip())
        for cod in sorted(todos_en_horario - en_turno - en_ausentes - media_hora_codigos):
            ausentes.append({'Código': cod, 'Descripción': 'Libre'})

    return {
        't1': t1, 't2': t2, 't3': t3,
        'ausentes':         ausentes,
        'primer_dia_libre': list(primer_dia_libre.keys()),
        'turno_dividido':   turno_dividido,
        'tiene_y':          tiene_y,
    }


def _fetch_horario_github():
    """Descarga el horario desde GitHub. Soporta repos privados via token en st.secrets."""
    url = GITHUB_RAW + "/Horario.xlsx"
    try:
        headers = {}
        token = st.secrets.get("GITHUB_TOKEN", "")
        if token:
            headers["Authorization"] = f"token {token}"
        r = requests.get(url, headers=headers, timeout=15)
        if r.status_code == 200:
            return r.content, url.split("/")[-1]
        return None, f"HTTP {r.status_code} al descargar el horario desde GitHub."
    except Exception as e:
        return None, str(e)


def generar_informe_tiempo(horario_bytes, fecha, ausentes_editados, observaciones="", turnos_override=None, codigos_4hrs=None):
    """Llena la plantilla local con los datos del horario para la fecha."""
    PLANTILLA_IT = "plantilla_Informe_de_tiempo.xlsx"
    if not os.path.exists(PLANTILLA_IT):
        return None, f"No se encontró '{PLANTILLA_IT}' en la carpeta de la app."

    if turnos_override is not None:
        turnos = turnos_override
    else:
        turnos = _leer_turnos_horario(horario_bytes, fecha)
    if turnos is None:
        return None, f"No se encontró el día {fecha.strftime('%d/%m/%Y')} en el horario."

    wb = openpyxl.load_workbook(PLANTILLA_IT)
    ws = wb['Normal']

    ws['F6'] = fecha.day
    ws['G6'] = fecha.month
    ws['H6'] = fecha.year

    # Corrección Y: helper para parchear col C
    def _fix_y(fila):
        f = ws.cell(row=fila, column=3).value
        if f and isinstance(f, str) and f'A{fila}="K"' in f and f'A{fila}="Y"' not in f:
            ws.cell(row=fila, column=3).value = f.replace(f'A{fila}="K"', f'A{fila}="K",A{fila}="Y"')

    for i, fila in enumerate([11, 13, 15, 17, 19, 21]):
        ws.cell(row=fila, column=1).value = turnos['t1'][i] if i < len(turnos['t1']) else None
        _fix_y(fila)

    for i, fila in enumerate([23, 25, 27, 29, 31, 33, 35, 37, 39, 41]):
        ws.cell(row=fila, column=1).value = turnos['t2'][i] if i < len(turnos['t2']) else None
        _fix_y(fila)

    for i, fila in enumerate([43, 45, 47, 49, 51]):
        ws.cell(row=fila, column=1).value = turnos['t3'][i] if i < len(turnos['t3']) else None
        _fix_y(fila)

    _es_dia_especial = fecha.weekday() == 6 or _es_feriado_fecha(fecha)
    if _es_dia_especial:
        for i, fila in enumerate([11, 13, 15, 17, 19, 21]):
            if i < len(turnos['t1']) and turnos['t1'][i]:
                ws.cell(row=fila, column=6).value = None
                ws.cell(row=fila, column=7).value = 9
        for i, fila in enumerate([23, 25, 27, 29, 31, 33, 35, 37, 39, 41]):
            if i < len(turnos['t2']) and turnos['t2'][i]:
                ws.cell(row=fila, column=6).value = None
                ws.cell(row=fila, column=7).value = 8
        for i, fila in enumerate([43, 45, 47, 49, 51]):
            if i < len(turnos['t3']) and turnos['t3'][i]:
                ws.cell(row=fila, column=6).value = None
                ws.cell(row=fila, column=7).value = 9
    if not _es_dia_especial:
        primer_libre_set = set(turnos.get('primer_dia_libre', []))
        if primer_libre_set:
            for i, fila in enumerate([11, 13, 15, 17, 19, 21]):
                cod = turnos['t1'][i] if i < len(turnos['t1']) else None
                if cod and cod in primer_libre_set:
                    ws.cell(row=fila, column=7).value = 9
            for i, fila in enumerate([23, 25, 27, 29, 31, 33, 35, 37, 39, 41]):
                cod = turnos['t2'][i] if i < len(turnos['t2']) else None
                if cod and cod in primer_libre_set:
                    ws.cell(row=fila, column=7).value = 8
            for i, fila in enumerate([43, 45, 47, 49, 51]):
                cod = turnos['t3'][i] if i < len(turnos['t3']) else None
                if cod and cod in primer_libre_set:
                    ws.cell(row=fila, column=7).value = 9

    # ── 4 hrs (+ 1 nocturna) por haber laborado en día nacional ─────────────
    # Col G (columna 7): t1 y t3 → 5 hrs (4+1 nocturna), t2 → 4 hrs
    if codigos_4hrs:
        _hrs_por_turno = {'t1': 5, 't2': 4, 't3': 5}
        for i, fila in enumerate([11, 13, 15, 17, 19, 21]):
            cod = turnos['t1'][i] if i < len(turnos['t1']) else None
            if cod and cod in codigos_4hrs:
                ws.cell(row=fila, column=7).value = _hrs_por_turno['t1']
        for i, fila in enumerate([23, 25, 27, 29, 31, 33, 35, 37, 39, 41]):
            cod = turnos['t2'][i] if i < len(turnos['t2']) else None
            if cod and cod in codigos_4hrs:
                ws.cell(row=fila, column=7).value = _hrs_por_turno['t2']
        for i, fila in enumerate([43, 45, 47, 49, 51]):
            cod = turnos['t3'][i] if i < len(turnos['t3']) else None
            if cod and cod in codigos_4hrs:
                ws.cell(row=fila, column=7).value = _hrs_por_turno['t3']

    # ── Empleados ausentes → filas 55-62 ─────────────────────────────────────
    # ── Empleados ausentes → filas 55-66 (hasta 12 registros) ────────────────
    for i, fila in enumerate([55, 56, 57, 58, 59, 60, 61, 62, 63, 64, 65, 66]):
        if i < len(ausentes_editados):
            ws.cell(row=fila, column=1).value = ausentes_editados[i]['Código']
            ws.cell(row=fila, column=4).value = ausentes_editados[i]['Descripción']
        else:
            ws.cell(row=fila, column=1).value = None
            ws.cell(row=fila, column=4).value = None

    # ── Observaciones → filas 68-72, col B (mergeadas B:H) ───────────────────
    if observaciones.strip():
        lineas = observaciones.strip().splitlines()
        for i, fila_obs in enumerate([68, 69, 70, 71, 72]):
            ws.cell(row=fila_obs, column=2).value = lineas[i] if i < len(lineas) else None

    output = BytesIO()
    wb.save(output)
    return output.getvalue(), None


def _analizar_festivo_anterior(horario_bytes, fecha_dt, feriados_check):
    """Busca el festivo más reciente (hasta 14 días). Acepta callable o set de (mes,dia)."""
    def _es_fest(d):
        return feriados_check(d) if callable(feriados_check) else (d.month, d.day) in feriados_check
    festivo = None
    for d in range(1, 15):
        cand = fecha_dt - timedelta(days=d)
        if _es_fest(cand):
            festivo = cand
            break
    if festivo is None:
        return {}

    wb = openpyxl.load_workbook(BytesIO(horario_bytes))
    ws = wb['LISTA']

    VALIDOS     = {'A','B','C','E','F','G','H','J','K','L','M','N','O','Q','R','S','T','W','X','Y','Z'}
    COMP_COLORS = {'FF30803D', 'FF074F69'}

    # Mapear fecha → columna una sola vez
    fila_fechas = [ws.cell(5, c + 1).value for c in range(ws.max_column)]
    date_to_col = {}
    for i, v in enumerate(fila_fechas):
        if hasattr(v, 'date'):
            date_to_col[v.date()] = i + 1

    col_festivo = date_to_col.get(festivo.date())
    if col_festivo is None:
        wb.close()
        return {}

    TURNO_ROWS = list(range(8, 34))
    COMP_ROWS  = list(range(36, 44)) + list(range(45, 56))

    def _cods_en_col(col, filas):
        result = set()
        for ri in filas:
            v = ws.cell(row=ri, column=col).value
            if v and str(v).strip() in VALIDOS:
                result.add(str(v).strip())
        return result

    def _es_comp(ri, col):
        try:
            fg = ws.cell(row=ri, column=col).fill.fgColor
            return fg is not None and fg.type == 'rgb' and fg.rgb in COMP_COLORS
        except Exception:
            return False

    workers_festivo = _cods_en_col(col_festivo, TURNO_ROWS)
    if not workers_festivo:
        wb.close()
        return {}

    # Días intermedios entre festivo y hoy
    dias_intermedios_cols = []
    d = festivo + timedelta(days=1)
    while d < fecha_dt:
        col_d = date_to_col.get(d.date())
        if col_d:
            dias_intermedios_cols.append(col_d)
        d += timedelta(days=1)

    # Primer día de regreso: no trabajó ningún día intermedio
    primer_regreso = set()
    for cod in workers_festivo:
        if not any(cod in _cods_en_col(c, TURNO_ROWS) for c in dias_intermedios_cols):
            primer_regreso.add(cod)

    if not primer_regreso:
        wb.close()
        return {}

    # ¿Tiene compensatorio después del festivo?
    cols_post = [col for date, col in date_to_col.items() if date > festivo.date()]
    resultado = {}
    for cod in primer_regreso:
        tiene_comp = False
        for col in cols_post:
            for ri in COMP_ROWS:
                v = ws.cell(row=ri, column=col).value
                if v and str(v).strip() == cod and _es_comp(ri, col):
                    tiene_comp = True
                    break
            if tiene_comp:
                break
        resultado[cod] = tiene_comp

    wb.close()
    return resultado


def vista_informe_tiempo():
    st.title("🗓️ INFORME DE TIEMPO")

    PLANTILLA_IT = "plantilla_Informe_de_tiempo.xlsx"
    if not os.path.exists(PLANTILLA_IT):
        st.error(f"⚠️ No se encontró `{PLANTILLA_IT}` en la carpeta de la app.")
        return

    # ── Cargar horario desde GitHub (se cachea en session_state por mes) ─────
    if "it_horario_bytes" not in st.session_state or not st.session_state["it_horario_bytes"]:
        with st.spinner("📥 Descargando horario desde GitHub..."):
            horario_bytes, msg = _fetch_horario_github()
        if not horario_bytes:
            st.error(f"❌ No se pudo descargar el horario: {msg}")
            st.info("Verifica que `GITHUB_RAW` apunte al repositorio correcto y el archivo esté subido.")
            return
        st.session_state["it_horario_bytes"]      = horario_bytes
        st.session_state["it_horario_nombre"]     = msg
        st.session_state["it_turnos_cache"]       = None
        st.session_state["it_turnos_cache_fecha"] = None

    # Botón para forzar re-descarga (ej: cuando se sube un horario actualizado)
    col_info, col_refetch = st.columns([4, 1])
    if col_refetch.button("🔄 Actualizar horario", help="En caso de actualización de la lista de turno"):
        st.session_state["it_horario_bytes"]      = None
        st.session_state["it_turnos_cache"]       = None
        st.session_state["it_turnos_cache_fecha"] = None
        st.rerun()


    # ── Selector de fecha (fuera de los tabs para que los return no bloqueen tab_pl) ──
    st.markdown("<p style='font-size:1.4rem; font-weight:600; margin-bottom:0.3rem;'>📅 Fecha</p>", unsafe_allow_html=True)
    fecha = st.date_input("fecha", value=datetime.now().date(), label_visibility="collapsed")
    fecha_dt = datetime.combine(fecha, datetime.min.time())

    # Cache de turnos por fecha
    cache_key = f"it_turnos_{fecha}"
    if st.session_state.get("it_turnos_cache_fecha") != cache_key:
        try:
            turnos = _leer_turnos_horario(st.session_state["it_horario_bytes"], fecha_dt)
            st.session_state["it_turnos_cache"]       = turnos
            st.session_state["it_turnos_cache_fecha"] = cache_key
            if turnos:
                st.session_state["it_ausentes_base"] = list(turnos['ausentes'])
        except Exception as e:
            st.error(f"Error leyendo horario: {e}")
            return

    turnos = st.session_state.get("it_turnos_cache")
    if not turnos:
        st.error(f"No se encontró {fecha.strftime('%d/%m/%Y')} en el horario.")
        return

    tab_it, tab_pl = st.tabs(["📄 Informe de Tiempo", "📋 Generar Planilla"])

    with tab_it:
        st.markdown("---")

        # ── Tabla de nombres desde plantilla ────────────────────────────────────
        PLANTILLA_IT = "plantilla_Informe_de_tiempo.xlsx"
        nombres_map = {}
        if os.path.exists(PLANTILLA_IT):
            try:
                wb_n = openpyxl.load_workbook(PLANTILLA_IT, data_only=True)
                ws_n = wb_n['Normal']
                for r in range(11, 32):
                    cod = ws_n.cell(r, 12).value
                    nom = ws_n.cell(r, 13).value
                    if cod and nom:
                        nombres_map[str(cod).strip()] = str(nom).strip()
                wb_n.close()
            except:
                pass

        obs_auto = []
        _nf = _nombre_feriado(fecha_dt)
        if _nf:
            obs_auto.append(f"DIA NACIONAL — {_nf}")
        for cod in turnos.get('primer_dia_libre', []):
            nombre = nombres_map.get(cod, cod)
            obs_auto.append(f"{nombre} laborando en primer día libre")
        for split in turnos.get('turno_dividido', []):
            if split.get('es_rojo'):
                n1 = nombres_map.get(split['primero'], split['primero'])
                n2 = nombres_map.get(split['segundo'], split['segundo'])
                obs_auto.append(f"{n1} laborando en jornada extendida")
                obs_auto.append(f"{n2} entrando anticipadamente")
        if turnos.get('tiene_y'):
            obs_auto.append("Jose Samudio en asistencia a la DEN")
        codigos_4hrs = {}
        horario_bytes_4h = st.session_state.get("it_horario_bytes")
        if horario_bytes_4h:
            try:
                analisis = _analizar_festivo_anterior(horario_bytes_4h, fecha_dt, _es_feriado_fecha)
                cod_a_turno = (
                    {c: 't1' for c in turnos['t1']} |
                    {c: 't2' for c in turnos['t2']} |
                    {c: 't3' for c in turnos['t3']}
                )
                for cod, tiene_comp in analisis.items():
                    if cod in cod_a_turno and not tiene_comp:
                        codigos_4hrs[cod] = cod_a_turno[cod]
                if codigos_4hrs:
                    nombres_acum = ", ".join(nombres_map.get(c, c) for c in codigos_4hrs)
                    obs_auto.append(f"{nombres_acum} + 4 hrs por haber laborado en día nacional")
            except Exception:
                pass

        # Cuando cambia la fecha: actualizar state y forzar rerun para que
        # el text_area reciba el nuevo valor en el siguiente render
        _obs_texto = "\n".join(obs_auto)
        if st.session_state.get("it_obs_fecha") != str(fecha):
            st.session_state["it_obs_fecha"]    = str(fecha)
            st.session_state["it_obs_ta"]       = _obs_texto
            st.session_state["it_codigos_4hrs"] = codigos_4hrs
            st.rerun()

        # ── Distribución de turnos (editable) ────────────────────────────────────
        st.subheader("👷 Distribución de turnos")
        st.caption("✏️ Puedes editar, agregar o eliminar códigos antes de generar el informe.")

        # Inicializar en session_state solo cuando cambia la fecha o faltan las claves base
        if st.session_state.get("it_turnos_fecha") != str(fecha) or "it_t1_base" not in st.session_state:
            st.session_state["it_t1_base"] = [{"Código": c} for c in turnos['t1']]
            st.session_state["it_t2_base"] = [{"Código": c} for c in turnos['t2']]
            st.session_state["it_t3_base"] = [{"Código": c} for c in turnos['t3']]
            st.session_state["it_ausentes_base"] = list(turnos['ausentes'])
            st.session_state["it_turnos_fecha"] = str(fecha)
            # Limpiar estado interno de los editores para que se reinicialicen con la nueva base
            for _k in ["editor_t1", "editor_t2", "editor_t3", "editor_ausentes_it"]:
                st.session_state.pop(_k, None)

        col_cfg = st.column_config.TextColumn("Código", width="small")
        c1, c2, c3 = st.columns(3)
        with c1:
            st.markdown("**🌙 23:00 – 07:00**")
            df_t1_ed = st.data_editor(
                pd.DataFrame(st.session_state["it_t1_base"]) if st.session_state["it_t1_base"]
                else pd.DataFrame(columns=["Código"]),
                hide_index=True, use_container_width=True, num_rows="dynamic",
                column_config={"Código": col_cfg}, key="editor_t1",
            )
        with c2:
            st.markdown("**☀️ 07:00 – 15:00**")
            df_t2_ed = st.data_editor(
                pd.DataFrame(st.session_state["it_t2_base"]) if st.session_state["it_t2_base"]
                else pd.DataFrame(columns=["Código"]),
                hide_index=True, use_container_width=True, num_rows="dynamic",
                column_config={"Código": col_cfg}, key="editor_t2",
            )
        with c3:
            st.markdown("**🌆 15:00 – 23:00**")
            df_t3_ed = st.data_editor(
                pd.DataFrame(st.session_state["it_t3_base"]) if st.session_state["it_t3_base"]
                else pd.DataFrame(columns=["Código"]),
                hide_index=True, use_container_width=True, num_rows="dynamic",
                column_config={"Código": col_cfg}, key="editor_t3",
            )

        st.markdown("---")

        # ── Empleados ausentes editable ──────────────────────────────────────────
        st.subheader("📋 Libres")
        df_aus = pd.DataFrame(st.session_state["it_ausentes_base"]) \
                 if st.session_state.get("it_ausentes_base") \
                 else pd.DataFrame(columns=["Código", "Descripción"])

        st.caption("💡 Opciones comunes: " + " · ".join(DESCRIPCIONES_AUSENCIA))
        df_editado = st.data_editor(
            df_aus, hide_index=True, use_container_width=True, num_rows="dynamic",
            column_config={
                "Código": st.column_config.TextColumn("Código", width="small"),
                "Descripción": st.column_config.SelectboxColumn(
                    "Descripción",
                    width="large",
                    options=DESCRIPCIONES_AUSENCIA,
                    required=False,
                ),
            },
            key="editor_ausentes_it",
        )

        st.markdown("---")

        # ── Observaciones ────────────────────────────────────────────────────────
        st.subheader("📝 Observaciones")
        st.caption("sección Observaciones(máx. 5 líneas).")

        # it_obs_ta fue poblado por el bloque anterior vía st.rerun()
        observaciones = st.text_area(
            "Observaciones",
            placeholder="Escribe aquí las observaciones...\nCada línea es un renglón de la sección Observaciones.",
            height=130,
            max_chars=1000,
            label_visibility="collapsed",
            key="it_obs_ta",
        )

        lineas_obs = [l for l in observaciones.splitlines() if l.strip()]
        if len(lineas_obs) > 5:
            st.warning("⚠️ Solo se escribirán las primeras 5 líneas en la plantilla.")

        st.markdown("---")

        # ── Generar ──────────────────────────────────────────────────────────────
        if st.button("📄 Generar Informe de Tiempo", type="primary"):
            ausentes_final = [r for r in df_editado.to_dict("records")
                              if r.get("Código") and str(r["Código"]).strip()]
            t1_final = [r["Código"] for r in df_t1_ed.to_dict("records") if r.get("Código") and str(r["Código"]).strip()]
            t2_final = [r["Código"] for r in df_t2_ed.to_dict("records") if r.get("Código") and str(r["Código"]).strip()]
            t3_final = [r["Código"] for r in df_t3_ed.to_dict("records") if r.get("Código") and str(r["Código"]).strip()]
            turnos_editados = {**turnos, 't1': t1_final, 't2': t2_final, 't3': t3_final}
            datos, error = generar_informe_tiempo(
                st.session_state["it_horario_bytes"], fecha_dt, ausentes_final, observaciones,
                turnos_override=turnos_editados,
                codigos_4hrs=st.session_state.get("it_codigos_4hrs", {})
            )
            if error:
                st.error(error)
            else:
                dias_semana = ["Lunes","Martes","Miércoles","Jueves","Viernes","Sábado","Domingo"]
                dia_nombre = dias_semana[fecha.weekday()]
                nombre = f"{fecha.day:02d}-{fecha.month:02d}-{fecha.year} {dia_nombre}.xlsx"
                st.download_button(
                    f"⏬ Descargar {nombre}", datos, nombre,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                st.success("✅ Informe generado correctamente.")



    with tab_pl:
        import calendar as _cal_pl

        hor_actual   = st.session_state.get("it_horario_bytes")
        hor_anterior = st.session_state.get("pl_horario_ant")

        if not hor_anterior:
            with st.spinner("Descargando horario anterior..."):
                _b, _e = _fetch_horario_planilla("Horario_Anterior.xlsx")
                st.session_state["pl_horario_ant"] = _b
                hor_anterior = _b

        def _mes_lbl(hb):
            if not hb: return None, None
            m, a = _leer_mes_horario(hb)
            return (m, a) if m else (None, None)

        mes_act, anio_act = _mes_lbl(hor_actual)
        mes_ant, anio_ant = _mes_lbl(hor_anterior)
        today_pl = datetime.now()
        periodos_pl = []
        _en_1ra_act = today_pl.day <= 15

        if mes_act and anio_act:
            dias_act_pl = _cal_pl.monthrange(anio_act, mes_act)[1]
            periodos_pl.append({
                'label':  f"1ra quincena — {_MESES_ES_PL[mes_act]} {anio_act}" + ("  ✦" if _en_1ra_act else ""),
                'inicio': datetime(anio_act, mes_act, 1),
                'fin':    datetime(anio_act, mes_act, 15),
                'horario': hor_actual, 'mes': mes_act, 'anio': anio_act,
            })
            periodos_pl.append({
                'label':  f"2da quincena — {_MESES_ES_PL[mes_act]} {anio_act}" + ("" if _en_1ra_act else "  ✦"),
                'inicio': datetime(anio_act, mes_act, 16),
                'fin':    datetime(anio_act, mes_act, dias_act_pl),
                'horario': hor_actual, 'mes': mes_act, 'anio': anio_act,
            })

        if mes_ant and anio_ant:
            dias_ant_pl = _cal_pl.monthrange(anio_ant, mes_ant)[1]
            periodos_pl.append({
                'label':  f"1ra quincena — {_MESES_ES_PL[mes_ant]} {anio_ant}",
                'inicio': datetime(anio_ant, mes_ant, 1),
                'fin':    datetime(anio_ant, mes_ant, 15),
                'horario': hor_anterior, 'mes': mes_ant, 'anio': anio_ant,
            })
            periodos_pl.append({
                'label':  f"2da quincena — {_MESES_ES_PL[mes_ant]} {anio_ant}",
                'inicio': datetime(anio_ant, mes_ant, 16),
                'fin':    datetime(anio_ant, mes_ant, dias_ant_pl),
                'horario': hor_anterior, 'mes': mes_ant, 'anio': anio_ant,
            })

        if not periodos_pl:
            st.warning("No hay horarios disponibles.")
        else:
            lbl_act_pl = f"✅ {_MESES_ES_PL[mes_act]} {anio_act}" if mes_act else "❌ No disponible"
            lbl_ant_pl = f"✅ {_MESES_ES_PL[mes_ant]} {anio_ant}" if mes_ant else "❌ No disponible"
            st.caption(f"Horario.xlsx → {lbl_act_pl}  ·  Horario_Anterior.xlsx → {lbl_ant_pl}")

            _idx_default = 0 if _en_1ra_act else min(1, len(periodos_pl) - 1)
            sel_per = st.selectbox(
                "📅 Período",
                [p["label"] for p in periodos_pl],
                index=_idx_default,
                key="pl_sel_periodo",
            )
            periodo_pl = next(p for p in periodos_pl if p["label"] == sel_per)

            if not periodo_pl["horario"]:
                st.error("El horario para este período no está disponible.")
            else:
                try:
                    emps_pl = _leer_empleados_horario(periodo_pl["horario"])
                except Exception as _ex:
                    st.error(f"Error leyendo empleados: {_ex}")
                    emps_pl = []

                if emps_pl:
                    _per_key = sel_per[:10].replace(" ", "_")
                    opc_pl = {f"{e['cod']} — {e['nombre']}": e for e in emps_pl}
                    sel_emp_pl = st.selectbox("👤 Empleado", list(opc_pl.keys()),
                                              key=f"pl_emp_{_per_key}")
                    emp_pl = opc_pl[sel_emp_pl]

                    p_ini = periodo_pl["inicio"]
                    p_fin = periodo_pl["fin"]
                    st.caption(f"Período: **{p_ini.strftime('%d/%m/%Y')}** al **{p_fin.strftime('%d/%m/%Y')}**")

                    eventos_extra_list = []

                    # ── Doble jornada ─────────────────────────────────────────
                    st.markdown("**🔁 Doble jornada**")
                    hay_doble = st.checkbox("¿Hubo doble jornada en el período?", key="pl_hay_doble")
                    if hay_doble:
                        n_dobles_pl = st.number_input("Cantidad de dobles en el período",
                                                      min_value=1, max_value=5, value=1, step=1,
                                                      key="pl_n_dobles")
                        for _nd in range(int(n_dobles_pl)):
                            cd1, cd2 = st.columns([1, 2])
                            _fd = cd1.date_input(
                                f"Fecha doble {_nd+1}",
                                value=p_ini.date(), min_value=p_ini.date(), max_value=p_fin.date(),
                                key=f"pl_doble_fecha_{_nd}",
                            )
                            _jd = cd2.text_input(
                                f"Justificación doble {_nd+1}",
                                key=f"pl_doble_just_{_nd}",
                                placeholder="Ej: Doble jornada por falta de relevo"
                            )
                            eventos_extra_list.append({
                                'tipo': 'doble', 'fecha': _fd, 'justificacion': _jd,
                            })
                        st.caption("El turno extra se determina automáticamente según el horario.")

                    # ── Entrada anticipada ────────────────────────────────────
                    st.markdown("**⏪ Entrada anticipada**")
                    hay_ant = st.checkbox("¿Hubo entrada anticipada en el período?", key="pl_hay_ant")
                    if hay_ant:
                        n_ant = st.number_input("Número de días con entrada anticipada",
                                                min_value=1, max_value=5, value=1, step=1, key="pl_n_ant")
                        for _na in range(int(n_ant)):
                            ca1, ca2, ca3 = st.columns([2, 1, 2])
                            _fa = ca1.date_input(
                                f"Fecha anticipada {_na+1}",
                                value=p_ini.date(), min_value=p_ini.date(), max_value=p_fin.date(),
                                key=f"pl_ant_fecha_{_na}",
                            )
                            _ha = ca2.number_input("Hora entrada", min_value=0, max_value=23,
                                                   value=12, step=1, key=f"pl_ant_h_{_na}")
                            _ma = ca2.number_input("Minutos", min_value=0, max_value=59,
                                                   value=0, step=5, key=f"pl_ant_m_{_na}")
                            _ja = ca3.text_input(f"Justificación {_na+1}",
                                                 key=f"pl_ant_just_{_na}",
                                                 placeholder="entrando anticipadamente para cubrir...")
                            eventos_extra_list.append({
                                'tipo':             'anticipada',
                                'fecha':            _fa,
                                'hora_anticipada':  f"{int(_ha):02d}:{int(_ma):02d}",
                                'justificacion':    _ja,
                            })

                    # ── Jornada extendida ─────────────────────────────────────
                    st.markdown("**⏱️ Jornada extendida**")
                    hay_ext = st.checkbox("¿Hubo jornada extendida en el período?", key="pl_hay_ext")
                    if hay_ext:
                        n_ext = st.number_input("Número de días con jornada extendida",
                                                min_value=1, max_value=10, value=1, step=1, key="pl_n_ext")
                        for _ne in range(int(n_ext)):
                            cx1, cx2, cx3, cx4 = st.columns([2, 1, 1, 2])
                            _fext = cx1.date_input(
                                f"Fecha extendida {_ne+1}",
                                value=p_ini.date(), min_value=p_ini.date(), max_value=p_fin.date(),
                                key=f"pl_ext_fecha_{_ne}",
                            )
                            _hext = cx2.number_input("Horas", min_value=0, max_value=12,
                                                     value=0, step=1, key=f"pl_ext_h_{_ne}")
                            _mext = cx3.number_input("Minutos", min_value=0, max_value=59,
                                                     value=0, step=5, key=f"pl_ext_m_{_ne}")
                            _jext = cx4.text_input(f"Justificación {_ne+1}",
                                                   key=f"pl_ext_just_{_ne}",
                                                   placeholder="horas extras en espera del relevo...")
                            if int(_hext) > 0 or int(_mext) > 0:
                                eventos_extra_list.append({
                                    'tipo':          'extendida',
                                    'fecha':         _fext,
                                    'horas':         int(_hext),
                                    'minutos':       int(_mext),
                                    'justificacion': _jext,
                                })

                    # ── Cambio de turno ───────────────────────────────────────
                    st.markdown("**🔄 Cambio de turno**")
                    hay_cambio = st.checkbox("¿Hubo cambio de turno en el período?", key="pl_hay_cambio")
                    if hay_cambio:
                        st.caption("Si el empleado cambió su turno con otro compañero, indica el turno que realmente trabajó. "
                                   "Cambiar a 15-23 o 23-07 acumula recargo nocturno; cambiar a 07-15 no genera extra.")
                        n_cambios = st.number_input("Cantidad de cambios de turno", min_value=1, max_value=5,
                                                    value=1, step=1, key="pl_n_cambio")
                        for _nc in range(int(n_cambios)):
                            cc1, cc2, cc3 = st.columns([2, 1, 2])
                            _fc = cc1.date_input(
                                f"Fecha cambio {_nc+1}",
                                value=p_ini.date(), min_value=p_ini.date(), max_value=p_fin.date(),
                                key=f"pl_cambio_fecha_{_nc}",
                            )
                            _tc = cc2.selectbox(
                                f"Turno trabajado",
                                ["07-15", "15-23", "23-07"],
                                key=f"pl_cambio_turno_{_nc}",
                            )
                            _jc = cc3.text_input(
                                f"Justificación {_nc+1}",
                                key=f"pl_cambio_just_{_nc}",
                                placeholder="Cambio de turno con compañero..."
                            )
                            if _tc in ("15-23", "23-07"):
                                st.caption(f"↑ Turno {_tc}: acumula 1 hr recargo nocturno.")
                            else:
                                st.caption(f"↑ Turno {_tc}: no genera horas extra.")
                            eventos_extra_list.append({
                                'tipo':        'cambio_turno',
                                'fecha':       _fc,
                                'turno_nuevo': _tc,
                                'justificacion': _jc or f"Cambio de turno al {_tc}",
                            })

                    if st.button("📄 Generar Planilla", type="primary", key="btn_gen_planilla"):
                        with st.spinner("Procesando..."):
                            datos_pl, err_pl = generar_planilla_colaborador(
                                periodo_pl["horario"], emp_pl["cod"], emp_pl["nombre"],
                                str(emp_pl["num_empleado"]), p_ini, p_fin,
                                eventos_extra=eventos_extra_list if eventos_extra_list else None,
                                cedula=emp_pl.get("cedula", ""),
                            )
                        if err_pl:
                            st.error(err_pl)
                        else:
                            _partes_nombre = emp_pl['nombre'].strip().split()
                            _inicial  = _partes_nombre[0][0].upper() if _partes_nombre else ""
                            _apellido = _partes_nombre[-1].title() if len(_partes_nombre) > 1 else (_partes_nombre[0].title() if _partes_nombre else "")
                            _mes_nombre = _MESES_ES_PL[periodo_pl['mes']]
                            fname_pl = (
                                f"{_inicial}{_apellido} {p_ini.day} a {p_fin.day} "
                                f"de {_mes_nombre} de {periodo_pl['anio']}.xlsx"
                            )
                            st.download_button(
                                f"⬇️ Descargar {fname_pl}", datos_pl, fname_pl,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            )
                            st.success("✅ Planilla generada.")

# ══════════════════════════════════════════════════════════════════════════════
# MÓDULO: VALIDADOR DE LIBRANZAS
# ══════════════════════════════════════════════════════════════════════════════

# Palabras clave — si el nombre del campo contiene alguna, se oculta
CAMPOS_OCULTOS_KEYWORDS = [
    "numeraci", "id libranza", "id/numeraci", "tipo de libranza", "tipo libranza",
    "solicitante", "duraci", "cargo del responsable", "cargo responsable",
    "ultimo estado", "último estado", "fecha ingreso", "fecha de ingreso",
    "fecha solicitud", "fecha de solicitud",
]

def _campo_oculto(nombre: str) -> bool:
    n = nombre.lower().strip()
    return any(k in n for k in CAMPOS_OCULTOS_KEYWORDS)

SYSTEM_PROMPT_LIB = """Eres un sistema experto en validacion de libranzas electricas del sistema de transmision de Panama (ETESA/CND).

Tu tarea es:
1. Detectar automaticamente si la libranza es de tipo INTERRUPTOR, LINEA o LINEA_DOBLE segun su contenido.
2. Validarla aplicando las reglas correspondientes al tipo detectado.
3. Devolver UNICAMENTE un objeto JSON valido, sin texto adicional, sin backticks, sin explicaciones fuera del JSON.

REGLAS COMUNES A TODOS LOS TIPOS

DATOS GENERALES (verificar todos):
- Numeracion/ID de la libranza (ej: ETESA-200-2026)
- Tipo de libranza: debe ser PROGRAMADA, EMERGENCIA o FORZADA
- Fecha de Ingreso
- Fecha de Solicitud
- Fecha de Inicio
- Fecha Final
- Nombre del Solicitante
- Duracion Programada
- DESCRIPCION: debe mencionar los trabajos que se van a realizar y mencionar explicitamente el equipo principal que se va a desenergizar. Si no lo menciona marcar como OBSERVACION.
- OBSERVACIONES: debe mencionar el equipo principal y las condiciones requeridas. Si no lo menciona marcar como OBSERVACION.
- Indicacion de Solmant (Si o No). Si es RTR y no tiene numero de Solmant marcar como OBSERVACION.
- EQUIPOS: debe incluir el equipo principal con su nomenclatura correcta. Para linea doble deben aparecer AMBAS lineas. Si falta alguno marcar como OBSERVACION.
- Indicacion de Libranzas Vinculadas
- Responsable de Campo (nombre)
- Cargo del Responsable de Campo
- AGENTES INFORMADOS: si hay agentes listados confirmar que esten. Si no hay es OK.
- AGENTES INVOLUCRADOS: si hay agentes involucrados listados, buscar la seccion "Avales" del documento (puede aparecer en cualquier pagina, generalmente despues de los agentes involucrados). Esta seccion tiene columnas: Agente, Usuario, Status. Verificar que cada agente involucrado tenga una fila en esa seccion con Status "Aprobado" y una fecha. Si un agente involucrado tiene su aval registrado en la seccion Avales con status Aprobado, marcar como OK. Si algun agente involucrado NO tiene su fila en la seccion Avales o su status no es Aprobado, marcar como OBSERVACION indicando cual agente falta. Si no hay seccion Avales visible pero hay agentes involucrados, marcar como OBSERVACION. LIBRANZAS FORZADAS: el aval automatico NO esta permitido, si se detecta marcar OBSERVACION.
- Area Afectada, Energia No Servida, Generacion Obligada (si las hay)
- Ultimo Estado: APROBADO, CORREGIR, RECHAZADO, REPROGRAMAR o CANCELADO. Extraer comentarios en comentarios_correccion (si los hay).

NOMENCLATURA: Interruptores 115kV: 11A## o 11B##. Interruptores 230kV: 23A## o 23B##. Cuchillas manuales: N-1 y N+1 del ultimo par.

REGLAS PARA INTERRUPTORES

DISTINCION IMPORTANTE — QUIEN EJECUTA LOS PASOS EN CAMPO:
Antes de validar cualquier paso, debes identificar dos variables criticas:

RESPONSABLE_CAMPO = nombre del Responsable de Campo del encabezado de la libranza (campo "Responsable de Campo").
RESPONSABLE_SE = persona fisica que ejecuta los pasos operativos en la subestacion.

REGLA PARA DETERMINAR RESPONSABLE_SE segun el CARGO del Responsable de Campo:
- Si el cargo es TECNICO, SUPERVISOR o ASISTENTE → RESPONSABLE_SE = RESPONSABLE_CAMPO (la misma persona ejecuta los pasos).
- Si el cargo es ESPECIALISTA o COORDINADOR → RESPONSABLE_SE es una persona diferente. En este caso buscar en los pasos de la libranza: a partir del paso 2, el primer nombre que aparece como ejecutor (que no sea CND ni un agente externo) es RESPONSABLE_SE. Ejemplo: si el paso 2 lo ejecuta "RIVERA", entonces RESPONSABLE_SE = "RIVERA".
- Si no aparece el cargo en la libranza → comparar nombres: si en los pasos el ejecutor coincide con RESPONSABLE_CAMPO (nombre completo o apellido), entonces RESPONSABLE_SE = RESPONSABLE_CAMPO. Si el ejecutor es un nombre diferente a RESPONSABLE_CAMPO, ese nombre es RESPONSABLE_SE.

COMPARACION FLEXIBLE (aplica a todas las variables de nombre):
En los pasos puede aparecer solo el apellido o parte del nombre. Si el apellido del ejecutor coincide con alguna parte del nombre completo de la variable, considerarlo valido.
Ejemplo: RESPONSABLE_CAMPO="JOSUE MARTINEZ", en pasos aparece "MARTINEZ" → valido.
Ejemplo: RESPONSABLE_SE="JOSE RIVERA", en pasos aparece "RIVERA" → valido.

En el JSON usa SIEMPRE el nombre real en "ejecutor_esperado", nunca textos genericos como "RESPONSABLE_CAMPO" o "RESPONSABLE_SE".

VARIABLES PARA INTERRUPTORES: antes de validar extrae del encabezado:
- RESPONSABLE_CAMPO = nombre del Responsable de Campo.
- RESPONSABLE_SE = determinar segun la regla del cargo descrita arriba.
- En interruptores, RESPONSABLE_SE es quien ejecuta todos los pasos operativos (P1, P3, P4, P5, P6 de inicio y P1, P2, P3 de finalizacion).
- En el paso P7 de inicio (entrega al responsable), verificar que el destinatario sea RESPONSABLE_CAMPO (no RESPONSABLE_SE si son diferentes).
- En el JSON usa el NOMBRE REAL en "ejecutor_esperado", no el texto generico.

INICIO INTERRUPTOR (7 pasos en orden):
P1: Confirmar reunion de trabajo y verificar el ATS. Por: RESPONSABLE_CAMPO
P2: Abrir el interruptor [X] de S/E [X]. Por: CND
P3: Verificar posicion de abierto en patio del interruptor [X] de S/E [X]. Por: RESPONSABLE_CAMPO
P4: Colocar en posicion LOCAL el interruptor [X] de S/E [X] + Colocar Roja. Por: RESPONSABLE_CAMPO
P5: Utilizar adecuadamente el permisivo electromecanico o mecanico previo a operar las cuchillas manuales. Por: RESPONSABLE_CAMPO
P6: Abrir las cuchillas manuales [N-1] y [N+1] del interruptor [X] de S/E [X] + Colocar Roja. Por: RESPONSABLE_CAMPO
P7: Entregar el interruptor [X] de S/E [X] a RESPONSABLE_CAMPO. Por: CND. Verificar que el destinatario sea RESPONSABLE_CAMPO.
Paso adicional opcional valido: VERIFICAR QUE ESTE HABILITADO EL RECIERRE DE LA LINEA [X]. Marcarlo OK.

FINALIZACION INTERRUPTOR (4 pasos en orden):
P1: Entregar el interruptor [X] de S/E [X] al CND. Por: RESPONSABLE_CAMPO
P2: Cerrar las cuchillas manuales [N-1] y [N+1] del interruptor [X] de S/E [X] + Retirar Roja. Por: RESPONSABLE_CAMPO
P3: Colocar en posicion REMOTO el interruptor [X] de S/E [X] + Retirar Roja. Por: RESPONSABLE_CAMPO
P4: Cerrar el interruptor [X] de S/E [X]. Por: CND

REGLAS PARA LINEAS

IDENTIFICACION DE VARIABLES PARA LIBRANZAS DE LINEA:
Aplica la misma regla de cargo descrita arriba para determinar RESPONSABLE_SE:
- RESPONSABLE_CAMPO = nombre del encabezado.
- RESPONSABLE_SE = quien confirma ATS (paso 1) y entrega/recibe la linea:
  * Cargo TECNICO, SUPERVISOR o ASISTENTE → RESPONSABLE_SE = RESPONSABLE_CAMPO.
  * Cargo ESPECIALISTA o COORDINADOR → RESPONSABLE_SE es diferente; buscar desde el paso 2 el primer nombre ejecutor que no sea CND ni agente externo.
  * Sin cargo → si el nombre en los pasos coincide con RESPONSABLE_CAMPO es la misma persona, sino el otro nombre es RESPONSABLE_SE.
- OPERADOR_SEX = persona en campo en S/E X: verifica patio, coloca local/remoto, opera cuchillas. Se detecta por las operaciones que ejecuta en S/E X. PUEDE SER DISTINTO A RESPONSABLE_SE.
- OPERADOR_SEY = igual pero en S/E Y.

COMO DETECTAR OPERADOR_SEX y OPERADOR_SEY:
Busca quien ejecuta "VERIFICAR POSICION EN PATIO", "COLOCAR EN POSICION LOCAL", "ABRIR/CERRAR CUCHILLAS MANUALES", "CUCHILLA DE TIERRA" en cada S/E. Esa persona es el OPERADOR de esa S/E.

COMPARACION FLEXIBLE: el apellido del ejecutor puede ser parte del nombre completo. Si coincide → valido. Ej: OPERADOR_SEX="OSCAR CAMAÑO", ejecutor="CAMAÑO" → valido.

En el JSON usa SIEMPRE el nombre real en "ejecutor_esperado".

TABLA DE PERMISOS PARA LIBRANZAS DE LINEA (si alguien opera algo que no le corresponde = ERROR):
- REUNION DE INICIO Y ATS              → RESPONSABLE_CAMPO exclusivamente.
- INTERRUPTORES (abrir o cerrar)       → CND exclusivamente (o agente externo A/B si no es CND). NUNCA OPERADOR_SE.
- CUCHILLA MOTORIZADA (abrir o cerrar) → CND exclusivamente. NUNCA OPERADOR_SE.
- COLOCAR EN POSICION LOCAL/REMOTO     → OPERADOR_SE de esa S/E exclusivamente. NUNCA CND.
- CUCHILLA DE TIERRA (cerrar o abrir)  → OPERADOR_SE de esa S/E exclusivamente. NUNCA CND.
- VERIFICAR POSICION EN PATIO          → OPERADOR_SE de esa S/E exclusivamente. NUNCA CND.
- CUCHILLAS MANUALES (abrir o cerrar)  → OPERADOR_SE de esa S/E exclusivamente. NUNCA CND.
- ENTREGA DE LINEA al responsable      → CND exclusivamente.
- ENTREGA DE LINEA al CND              → RESPONSABLE_CAMPO exclusivamente.

INICIO LINEA (pasos operativos obligatorios — numeracion puede variar por pasos de verificacion intercalados):
P1: Confirmar reunion de trabajo y verificar el ATS. Por: RESPONSABLE_CAMPO (verificar que sea el nombre real del responsable de campo)
P2: Abrir los interruptores de S/E [X]. Por: CND (o agente externo A si no es CND). NUNCA OPERADOR_SEX.
P3: Abrir los interruptores de S/E [Y]. Por: CND (o agente externo B si no es CND). NUNCA OPERADOR_SEY.
P4: Abrir la cuchilla motorizada de S/E [X] + Colocar Roja. Por: CND exclusivamente.
P5: Abrir la cuchilla motorizada de S/E [Y] + Colocar Roja. Por: CND exclusivamente.
P6: Cerrar la cuchilla de tierra de la linea en S/E [X]. Por: OPERADOR_SEX (verificar que sea el nombre real del operador de S/E X).
P7: Cerrar la cuchilla de tierra de la linea en S/E [Y]. Por: OPERADOR_SEY (verificar que sea el nombre real del operador de S/E Y).
P8: Cerrar los interruptores de S/E [X]. Por: CND (o agente externo A). NUNCA OPERADOR_SEX.
P9: Cerrar los interruptores de S/E [Y]. Por: CND (o agente externo B). NUNCA OPERADOR_SEY.
P10: Entregar la linea a RESPONSABLE_CAMPO. Por: CND exclusivamente. Verificar que el nombre del destinatario en este paso coincida con RESPONSABLE_CAMPO.

FINALIZACION LINEA (pasos operativos obligatorios):
P1: Entregar la linea al CND. Por: RESPONSABLE_CAMPO (verificar que sea el nombre real del responsable de campo).
P2: Abrir los interruptores de S/E [X]. Por: CND (o agente externo A). NUNCA OPERADOR_SEX.
P3: Abrir los interruptores de S/E [Y]. Por: CND (o agente externo B). NUNCA OPERADOR_SEY.
P4: Abrir la cuchilla de tierra de la linea en S/E [X] + Retirar Roja. Por: OPERADOR_SEX exclusivamente.
P5: Abrir la cuchilla de tierra de la linea en S/E [Y] + Retirar Roja. Por: OPERADOR_SEY exclusivamente.
P6: Cerrar la cuchilla motorizada de S/E [X] + Retirar Roja. Por: CND exclusivamente.
P7: Cerrar la cuchilla motorizada de S/E [Y] + Retirar Roja. Por: CND exclusivamente.
P8: Cerrar los interruptores de S/E [X]. Por: CND (o agente externo A). NUNCA OPERADOR_SEX.
P9: Cerrar los interruptores de S/E [Y]. Por: CND (o agente externo B). NUNCA OPERADOR_SEY.
P10: Verificar que este habilitado el recierre de la linea. Por: CND exclusivamente.

CONSISTENCIA DE EQUIPOS: las cuchillas motorizadas abiertas en inicio deben cerrarse en fin. Las cuchillas de tierra cerradas en inicio deben abrirse en fin. Si hay inconsistencia de equipos marcar ERROR.

REGLAS LINEA DOBLE (dos lineas al mismo tiempo):
- EQUIPOS: deben aparecer ambas lineas. DESCRIPCION y OBSERVACIONES: deben mencionar ambas.
- INICIO: bloque completo de 11 pasos para Linea 1 + bloque completo de 11 pasos para Linea 2.
- FINALIZACION: bloque completo de 10 pasos para Linea 1 + bloque completo de 10 pasos para Linea 2.
- Pasos SPEAR validos (OK): DESHABILITAR/HABILITAR LA CONTINGENCIA C[X] DEL SPEAR PARA LA LINEA [X]. Por: CND. Asociados a lineas 230-3B y 230-4B, 230-50, 230-49A.

REGLA — LINEA YA DESENERGIZADA:
En algunos casos la libranza indica explicitamente (en la descripcion, observaciones o justificacion de libranza forzada) que la linea YA SE ENCUENTRA DESENERGIZADA, abierta y aterrizada en ambos extremos antes de iniciar los trabajos. En este caso las maniobras son diferentes y mas cortas que el procedimiento estandar. NO marques error por no seguir los 11 pasos normales. Las maniobras validas para este caso son:

INICIO cuando la linea ya esta desenergizada:
- P1: Confirmar que se realizo la reunion de inicio de trabajos y se reviso el ATS. Por: RESPONSABLE DE CAMPO. (obligatorio)
- Pasos de verificacion del CND: el CND verifica que la linea efectivamente este abierta, aterrizada y los interruptores abiertos y seccionados. Estos pasos son validos y correctos. Por: CND.
- Ultimo paso de inicio: el CND indica al responsable de campo que puede proceder con los trabajos (o entrega la linea al responsable). Por: CND.

FINALIZACION cuando la linea ya estaba desenergizada:
- P1: El responsable de campo indica al CND que finalizaron los trabajos. Por: RESPONSABLE DE CAMPO. (obligatorio)
- Pasos de verificacion del CND: el CND verifica que la linea siga en las condiciones requeridas (abierta, aterrizada, interruptores seccionados). Estos pasos son validos. Por: CND.

COMO DETECTAR ESTE CASO: buscar en descripcion, observaciones o justificacion frases como "la linea se encuentra actualmente desenergizada", "linea abierta y aterrizada", "linea se mantiene desenergizada", o similares. Si se detecta este caso, aplicar las reglas de verificacion anteriores en lugar de los 11/10 pasos estandar.

REGLA — PASOS VERIFICAR NO GENERAN RESTRICCIONES (aplica a todos los tipos):
Los pasos del tipo "VERIFICAR QUE EL INTERRUPTOR [X] ESTÉ ABIERTO/CERRADO" o cualquier paso que empiece con VERIFICAR son SOLO CONFIRMACIONES VISUALES. NO son operaciones. NO crean restricciones sobre ningun paso posterior.

PRINCIPIO FUNDAMENTAL: Si los pasos siguientes a un VERIFICAR se están ejecutando, significa que la condicion del VERIFICAR ya fue satisfecha. Por lo tanto:
- NUNCA uses un paso VERIFICAR como razon para marcar ERROR en cualquier otro paso.
- NUNCA generes el mensaje "El interruptor [X] debe permanecer abierto segun observaciones" basandote en un paso VERIFICAR. Esa logica es incorrecta.
- Un paso VERIFICAR solo confirma el estado actual — no implica que el equipo deba mantenerse asi por el resto de la libranza ni que no pueda operarse.

CASOS CONCRETOS QUE DEBEN SER OK:
- "VERIFICAR QUE EL INTERRUPTOR 11A72 ESTE ABIERTO" en paso 5 → NO genera ninguna restriccion. Los pasos 13 (COLOCAR EN POSICION LOCAL EL CONTROL DEL INTERRUPTOR 11A72) y 18 (ABRIR LAS CUCHILLAS MANUALES 11A71 Y 11A73 DEL INTERRUPTOR 11A72) son VALIDOS y deben marcarse OK.
- En general: operar las cuchillas manuales, cuchilla de tierra, o posicion de control (LOCAL/REMOTO) de un interruptor es SIEMPRE valido independientemente de si hay un VERIFICAR previo sobre ese interruptor.

REGLA — EJECUTOR DUAL (CND/NOMBRE o NOMBRE/CND):
Cuando el ejecutor de un paso aparece como "CND/NOMBRE" o "NOMBRE/CND" (ej: "CND/BETHANCOURTH"), significa que el paso requiere coordinacion entre CND y el operador local. Para validar el ejecutor:
- Si el paso es de CUCHILLA MOTORIZADA (abrir o cerrar): el ejecutor esperado es CND. Si aparece "CND/NOMBRE", el paso es VALIDO (OK) porque CND coordina con el operador. NO marcar error solo porque aparezca un segundo nombre junto a CND.
- Si el paso es de INTERRUPTORES: el ejecutor esperado es CND. Si aparece "CND/NOMBRE", tambien es VALIDO.
- Solo marcar ERROR si CND NO aparece en el ejecutor cuando el paso lo requiere exclusivamente (ej: ejecutor = solo "BETHANCOURTH" en un paso de cuchilla motorizada).


Si en las maniobras aparece una linea que sea un encabezado o anotacion informativa como por ejemplo "PASOS PARA DESENERGIZAR LA LINEA X", "MANIOBRAS PARA INTERRUPTOR X", "INICIO DE MANIOBRAS", o cualquier texto que claramente sea un titulo o nota y no una instruccion operativa, NO debe validarse como paso. Debe incluirse en la lista con estado "OK" e indicar en descripcion_encontrada que es un paso informativo, sin verificar contenido ni ejecutor.

REGLA — PASOS DE VERIFICACION EN CAMPO (aplica a lineas):
En las libranzas de lineas pueden aparecer pasos intermedios del tipo "VERIFICAR LA POSICION DE ABIERTO EN PATIO DE..." o similares. Estos pasos son VALIDOS y deben marcarse como OK. Sin embargo, su presencia altera la numeracion de los pasos en el documento. Por lo tanto:
- NO valides los pasos operativos por su numero en el documento, sino por su CONTENIDO.
- Los pasos operativos obligatorios que debes buscar son (independientemente del numero que tengan en el documento):
  INICIO: Confirmar ATS, Abrir interruptores S/E X, Abrir interruptores S/E Y, Abrir cuchilla motorizada S/E X, Abrir cuchilla motorizada S/E Y, Cerrar cuchilla de tierra S/E X, Cerrar cuchilla de tierra S/E Y, Cerrar interruptores S/E X, Cerrar interruptores S/E Y, Entregar linea al responsable.
  FINALIZACION: Entregar linea al CND, Abrir interruptores S/E X, Abrir interruptores S/E Y, Abrir cuchilla de tierra S/E X, Abrir cuchilla de tierra S/E Y, Cerrar cuchilla motorizada S/E X, Cerrar cuchilla motorizada S/E Y, Cerrar interruptores S/E X, Cerrar interruptores S/E Y, Verificar recierre.
- Los pasos de VERIFICACION DE POSICION EN PATIO son opcionales y validos. Si aparecen incluyelos como OK pero NO los cuentes como pasos operativos obligatorios.
- En el JSON, el campo "paso" debe reflejar el numero real del documento, pero la validacion de contenido y ejecutor debe hacerse sobre el paso operativo que representa, no sobre su posicion numerica.
- El ejecutor de los pasos de verificacion en patio suele ser el operador local de la S/E (CARRERA, RUDAS, o quien sea el operador designado), lo cual es correcto.

REGLA — ORDEN DE CUCHILLAS DE TIERRA (aplica a lineas simples y dobles):
En los pasos de cierre de cuchillas de tierra (inicio, pasos 7 y 8) y apertura de cuchillas de tierra (finalizacion, pasos 4 y 5), el ORDEN entre los dos extremos (S/E X y S/E Y) NO es relevante. Lo que importa es que AMBAS cuchillas de tierra esten presentes (una por cada extremo de la linea). Si aparece primero la de S/E Y y luego la de S/E X, eso es CORRECTO y no debe marcarse como error. Solo marcar ERROR si falta una de las dos cuchillas de tierra, no por el orden en que aparecen.

REGLAS PARA TRANSFORMADORES
Cuando el tipo de equipo es TRANSFORMADOR o AUTOTRANSFORMADOR:
- Detectar automaticamente el nombre del transformador (T1, T2, T3, etc.) y la subestacion.
- Si se proporcionan PASOS ESPERADOS en el prompt, validar cada paso de la libranza contra esos pasos esperados.
- Usar RESPONSABLE_CAMPO como variable que almacena el nombre real del responsable segun el encabezado de la libranza.
- Los pasos con ejecutor NOTA son informativos: marcarlos OK sin verificar contenido ni ejecutor.
- Comparacion flexible de nombres: si el apellido del ejecutor coincide con parte del nombre completo, es valido.
- Verificar que los pasos de inicio y finalizacion esten completos y en el orden correcto.
- Si no se proporcionan pasos esperados, validar la estructura general: confirmar ATS al inicio, entrega al responsable al final del inicio, entrega al CND al inicio de finalizacion.
- En el JSON usar tipo_equipo = "TRANSFORMADOR".

REGLA — OPERADOR_SE EN TRANSFORMADORES (dos operadores de campo):
Algunas libranzas de transformador (especialmente pruebas electricas o mantenimientos con personal especializado) tienen DOS personas de campo con roles distintos:
  * RESPONSABLE_CAMPO: el responsable oficial del encabezado (puede ser de protecciones, pruebas, etc.).
  * OPERADOR_SE: un segundo tecnico de subestaciones que ejecuta las maniobras fisicas de switching (verificar posicion en patio, colocar en local/remoto, abrir/cerrar cuchillas manuales).
Para detectar al OPERADOR_SE: busca el primer nombre distinto a RESPONSABLE_CAMPO y a CND que aparezca ejecutando pasos de tipo: VERIFICAR POSICION EN PATIO, COLOCAR EN LOCAL, COLOCAR EN REMOTO, ABRIR/CERRAR CUCHILLAS MANUALES. Ese nombre es el OPERADOR_SE.
Una vez identificado el OPERADOR_SE, todos sus pasos de maniobras fisicas de subestacion son VALIDOS (OK). NO marcar error por el hecho de que no sea RESPONSABLE_CAMPO.
El CND NUNCA ejecuta pasos de COLOCAR EN LOCAL/REMOTO ni de CUCHILLAS MANUALES — esos son siempre de campo (OPERADOR_SE o RESPONSABLE_CAMPO). Si un paso dice "Por: CND" y es de esos tipos, SI es error.

FORMATO JSON:
{"libranza_id":"string","tipo_equipo":"INTERRUPTOR|LINEA|LINEA_DOBLE","equipo_principal":"string","subestacion":"string","tipo_libranza":"string","ultimo_estado":"string","responsable_campo":"string","comentarios_correccion":"string|null","resumen":{"total_verificaciones":0,"total_ok":0,"total_observaciones":0},"datos_generales":[{"campo":"string","estado":"OK|OBSERVACION","detalle":"string"}],"maniobras_inicio":[{"paso":0,"descripcion_esperada":"string","descripcion_encontrada":"string","ejecutor_esperado":"string","ejecutor_encontrado":"string","estado":"OK|ERROR","errores":[],"linea":"LINEA 1|LINEA 2|null"}],"maniobras_finalizacion":[{"paso":0,"descripcion_esperada":"string","descripcion_encontrada":"string","ejecutor_esperado":"string","ejecutor_encontrado":"string","estado":"OK|ERROR","errores":[],"linea":"LINEA 1|LINEA 2|null"}]}"""


def _detectar_tipo_libranza(client, pdf_b64: str) -> dict:
    """Paso 1: detección rápida del tipo, transformador y S/E."""
    msg = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=300,
        system='Eres un experto en libranzas electricas. Analiza el PDF y responde UNICAMENTE con un JSON con estos campos: {tipo_equipo: INTERRUPTOR|LINEA|LINEA_DOBLE|TRANSFORMADOR, equipo_principal: nombre del equipo, subestacion: nombre de la SE}. Sin texto adicional ni backticks.',
        messages=[{"role":"user","content":[
            {"type":"document","source":{"type":"base64","media_type":"application/pdf","data":pdf_b64}},
            {"type":"text","text":"Detecta tipo de equipo, nombre del equipo y subestacion. Solo JSON."},
        ]}],
    )
    raw = msg.content[0].text
    m = re.search(r'\{[\s\S]*?\}', raw)
    clean = m.group(0) if m else raw.replace("```json","").replace("```","").strip()
    return json.loads(clean)


def _validar_libranza_api(pdf_bytes: bytes) -> dict:
    client = anthropic.Anthropic(api_key=_api_key_libranzas)
    pdf_b64 = base64.b64encode(pdf_bytes).decode("utf-8")

    # ── Paso 1: detección rápida ──
    try:
        deteccion = _detectar_tipo_libranza(client, pdf_b64)
        tipo = deteccion.get("tipo_equipo","").upper()
    except Exception:
        tipo = ""

    # ── Paso 2: si es TRANSFORMADOR o LINEA, cargar pasos del Excel ──
    prompt_extra = ""
    if tipo == "TRANSFORMADOR":
        equipo = deteccion.get("equipo_principal","")
        se     = deteccion.get("subestacion","")
        pasos_data = buscar_pasos_transformador(equipo, se)
        if pasos_data:
            prompt_extra = ("\n\n" + "="*50 +
                "\nPASOS ESPERADOS PARA ESTE TRANSFORMADOR (extraidos del archivo pasos_transformadores.xlsx):\n" +
                "="*50 + "\n" + formatear_pasos_para_prompt(pasos_data) +
                "\n\nIMPORTANTE: Valida los pasos de la libranza contra los pasos esperados listados arriba. " +
                "Verifica que cada paso este presente con el ejecutor correcto y en el orden correcto. " +
                "Los pasos con ejecutor RESPONSABLE_CAMPO pueden ser ejecutados por el responsable de campo real " +
                "O por el OPERADOR_SE (segundo tecnico de subestaciones, segun la regla OPERADOR_SE del system prompt). " +
                "Los pasos con ejecutor NOTA son informativos y se validan como OK sin verificar ejecutor. " +
                "Si un paso esperado no aparece en la libranza marcalo como ERROR. " +
                "Si hay pasos adicionales no listados que sean de verificacion en patio o similares, marcarlos como OK.")

    elif tipo in ("LINEA", "LINEA_DOBLE"):
        equipo = deteccion.get("equipo_principal","")
        se     = deteccion.get("subestacion","")
        pasos_data = buscar_pasos_linea(equipo, se)
        if pasos_data:
            prompt_extra = ("\n\n" + "="*50 +
                "\nPASOS ESPERADOS PARA ESTA LINEA (extraidos del archivo pasos_lineas.xlsx):\n" +
                "="*50 + "\n" + formatear_pasos_para_prompt(pasos_data) +
                "\n\nIMPORTANTE: Valida los pasos de la libranza contra los pasos esperados listados arriba. " +
                "Verifica que cada paso este presente con el ejecutor correcto y en el orden correcto. " +
                "Los pasos con ejecutor RESPONSABLE_CAMPO deben ser ejecutados por el responsable de campo real de la libranza. " +
                "Los pasos con ejecutor NOTA son informativos y se validan como OK sin verificar ejecutor. " +
                "Si un paso esperado no aparece en la libranza marcalo como ERROR. " +
                "Si hay pasos adicionales no listados que sean de verificacion en patio o similares, marcarlos como OK.")

    system_final = SYSTEM_PROMPT_LIB + prompt_extra

    msg = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=16000,
        system=system_final,
        messages=[{"role":"user","content":[
            {"type":"document","source":{"type":"base64","media_type":"application/pdf","data":pdf_b64}},
            {"type":"text","text":"Detecta el tipo de libranza y validala. Responde UNICAMENTE con el JSON."},
        ]}],
    )
    raw = msg.content[0].text
    m = re.search(r'\{[\s\S]*\}', raw)
    clean = m.group(0) if m else raw.replace("```json","").replace("```","").strip()
    try:
        return json.loads(clean)
    except json.JSONDecodeError as e:
        raise ValueError(f"No se pudo parsear el JSON.\n\nRespuesta:\n{raw[:600]}") from e


def _badge_lib(valor: str) -> str:
    v = (valor or "").upper()
    mapa = {
        "APROBADO":("pill-lib-ok","✓ APROBADO"),"PROGRAMADA":("pill-lib-ok","PROGRAMADA"),
        "OK":("pill-lib-ok","✓ OK"),"RECHAZADO":("pill-lib-err","✗ RECHAZADO"),
        "CANCELADO":("pill-lib-err","CANCELADO"),"CORREGIR":("pill-lib-amb","⚠ CORREGIR"),
        "REPROGRAMAR":("pill-lib-amb","REPROGRAMAR"),"EMERGENCIA":("pill-lib-amb","EMERGENCIA"),
        "FORZADA":("pill-lib-amb","FORZADA"),"INTERRUPTOR":("pill-lib-blue","⚡ INTERRUPTOR"),
        "LINEA":("pill-lib-blue","〰️ LÍNEA"),"LINEA_DOBLE":("pill-lib-blue","〰️〰️ LÍNEA DOBLE"),"TRANSFORMADOR":("pill-lib-blue","🔄 TRANSFORMADOR"),
    }
    cls, label = mapa.get(v, ("pill-lib-neu", valor or "—"))
    return f'<span class="{cls}">{label}</span>'


def _item_dg(campo, detalle, estado):
    cls   = "lib-item-ok"   if estado=="OK" else "lib-item-warn"
    icono = "✓" if estado=="OK" else "⚠"
    color = "#4cde8f" if estado=="OK" else "#f5a623"
    return f'<div class="{cls}"><span style="font-size:20px;color:#8a9099;font-weight:500;">{campo}</span><br><span style="color:{color};font-size:20px;">{icono} {detalle}</span></div>'


def _item_paso(paso):
    estado  = paso.get("estado","ERROR")
    cls     = "lib-item-ok" if estado=="OK" else "lib-item-err"
    icono   = "✓" if estado=="OK" else "✗"
    color   = "#4cde8f" if estado=="OK" else "#f06060"
    ej_esp  = paso.get("ejecutor_esperado","")
    ej_enc  = paso.get("ejecutor_encontrado","—")
    ej_diff = (f' <span style="color:#f06060;">(esperado:{ej_esp})</span>' if estado == "ERROR" and ej_esp and ej_esp!=ej_enc else "")
    errores = "".join(f'<div style="background:rgba(240,96,96,0.1);border:1px solid rgba(240,96,96,0.25);border-radius:4px;padding:4px 8px;margin-top:4px;font-size:15px;color:#f06060;">✗ {e}</div>' for e in (paso.get("errores") or []))
    linea   = paso.get("linea")
    ltag    = (f'<span style="background:rgba(96,165,250,0.15);border:1px solid rgba(96,165,250,0.3);color:#60a5fa;font-size:10px;padding:2px 7px;border-radius:3px;margin-left:6px;">{linea}</span>' if linea else "")
    desc    = paso.get("descripcion_encontrada") or "NO ENCONTRADO"

    # ── Badges de validación contra DB (uno por equipo encontrado) ─────────
    db_html = ""
    for vdb in paso.get("validaciones_db", []):
        if vdb["estado"] == "OK":
            db_color = "#4cde8f"; db_bg = "rgba(76,222,143,0.08)"; db_brd = "rgba(76,222,143,0.3)"; db_ic = "🗂✓"
        elif vdb["estado"] == "ERROR":
            db_color = "#f06060"; db_bg = "rgba(240,96,96,0.08)";  db_brd = "rgba(240,96,96,0.3)";  db_ic = "🗂✗"
        else:
            db_color = "#f5a623"; db_bg = "rgba(245,166,35,0.08)"; db_brd = "rgba(245,166,35,0.3)"; db_ic = "🗂⚠"

        ejecutor_txt = vdb.get("ejecutor") or ej_enc
        db_header = (
            f'<span style="color:{db_color};font-weight:600;">{db_ic} {vdb["codigo"]}</span>'
            f'<span style="color:#8a9099;font-size:11px;margin-left:6px;">'
            f'{vdb["tipo"]} · {vdb["instalacion_db"]} · Agente: {vdb["agente_db"]}'
            f' · <span style="color:#b0b8c4;">Por: {ejecutor_txt}</span>'
            f'</span>'
        )
        db_errores = "".join(
            f'<div style="color:#f06060;font-size:12px;margin-top:2px;">✗ {e}</div>'
            for e in vdb.get("errores", [])
        )
        db_html += (
            f'<div style="background:{db_bg};border:1px solid {db_brd};border-radius:4px;'
            f'padding:4px 10px;margin-top:5px;font-size:12px;">'
            f'{db_header}{db_errores}</div>'
        )

    return f'<div class="{cls}"><span style="font-size:18px;color:#636b78;">PASO {paso.get("paso")} · Por: {ej_enc}{ej_diff}</span>{ltag}<br><span style="color:{color};font-size:20px;">{icono} {desc}</span>{errores}{db_html}</div>'

def vista_libranzas():
    # Estilos específicos del módulo
    st.markdown("""
    <style>
    .lib-item-ok  {background:rgba(76,222,143,0.06);border-left:3px solid #4cde8f;border-radius:0 6px 6px 0;padding:0.5rem 1rem;margin-bottom:5px;}
    .lib-item-err {background:rgba(240,96,96,0.06); border-left:3px solid #f06060;border-radius:0 6px 6px 0;padding:0.5rem 1rem;margin-bottom:5px;}
    .lib-item-warn{background:rgba(245,166,35,0.06);border-left:3px solid #f5a623;border-radius:0 6px 6px 0;padding:0.5rem 1rem;margin-bottom:5px;}
    .lib-card{background:#111418;border:1px solid rgba(255,255,255,0.08);border-radius:10px;padding:1rem 1.25rem;margin-bottom:0.75rem;}
    .lib-coment{background:rgba(245,166,35,0.08);border:1px solid rgba(245,166,35,0.3);border-radius:6px;padding:8px 12px;margin:6px 0;font-size:13px;color:#f5a623;}
    .pill-lib-ok  {background:#4cde8f22;border:1px solid #4cde8f44;color:#4cde8f;padding:5px 14px;border-radius:5px;font-size:13px;font-weight:600;}
    .pill-lib-err {background:#f0606022;border:1px solid #f0606044;color:#f06060;padding:5px 14px;border-radius:5px;font-size:13px;font-weight:600;}
    .pill-lib-amb {background:#f5a62322;border:1px solid #f5a62344;color:#f5a623;padding:5px 14px;border-radius:5px;font-size:13px;font-weight:600;}
    .pill-lib-blue{background:#3b82f622;border:1px solid #3b82f644;color:#60a5fa;padding:5px 14px;border-radius:5px;font-size:13px;font-weight:600;}
    .pill-lib-neu {background:#1f242c;   border:1px solid #636b7844;color:#e8e4dc;padding:5px 14px;border-radius:5px;font-size:13px;}

    /* Uploader y botón compactos */
    [data-testid="stFileUploader"] {
        max-width: 300px !important;
    }
    /* Botón Validar Libranza */
    div[data-testid="stVerticalBlock"] .stButton > button[kind="primary"] {
        max-width: 300px !important;
        padding: 0.3rem 0.75rem !important;
        font-size: 13px !important;
        height: auto !important;
    }
    /* Mensaje verde de éxito */
    div[data-testid="stVerticalBlock"] div[data-testid="stAlert"] {
        max-width: 300px !important;
        padding: 0.35rem 0.75rem !important;
        font-size: 12px !important;
    }

    </style>""", unsafe_allow_html=True)

    c1, c2 = st.columns([1, 2.4], gap="small")

    with c1:
        st.markdown("#### 📄 Cargar Libranza")
        pdf_file = st.file_uploader("Sube el PDF de la libranza", type=["pdf"],
                                    help="Máximo 10 MB · Interruptores y Líneas",
                                    key="lib_pdf_uploader")
        if pdf_file is not None:
            st.session_state["lib_pdf_bytes"] = pdf_file.read()
            st.session_state["lib_pdf_name"]  = pdf_file.name
            st.session_state["lib_pdf_size"]  = pdf_file.size
        elif "lib_pdf_bytes" not in st.session_state:
            st.session_state["lib_pdf_bytes"] = None

        pdf_bytes = st.session_state.get("lib_pdf_bytes")
        pdf_name  = st.session_state.get("lib_pdf_name", "")
        pdf_size  = st.session_state.get("lib_pdf_size", 0)

        if not _api_key_libranzas:
            st.error("No se encontró config.txt con la API Key.")

        if pdf_bytes:
            st.success(f"✓ {pdf_name} ({pdf_size/1024:.1f} KB)")

        btn = st.button("⚡ Validar Libranza", disabled=not(pdf_bytes and _api_key_libranzas),
                        use_container_width=True, type="primary", key="lib_btn_validar")
        if not pdf_bytes:
            st.info("Sube una libranza para comenzar.")

    with c2:
        st.markdown("#### 📊 Reporte de Validación")

        if btn and pdf_bytes and _api_key_libranzas:
            with st.spinner("Analizando libranza con IA…"):
                try:
                    resultado = _validar_libranza_api(pdf_bytes)
                    st.session_state["lib_resultado"] = resultado
                except (json.JSONDecodeError, ValueError) as e:
                    st.error(f"Error al procesar la respuesta: {e}")
                    st.session_state.pop("lib_resultado", None)
                except anthropic.AuthenticationError:
                    st.error("API Key inválida. Verifica config.txt.")
                    st.session_state.pop("lib_resultado", None)
                except anthropic.RateLimitError:
                    st.error("Límite de uso alcanzado. Espera un momento.")
                    st.session_state.pop("lib_resultado", None)
                except Exception as e:
                    st.error(f"Error inesperado: {e}")
                    st.session_state.pop("lib_resultado", None)

        if "lib_resultado" in st.session_state:
            d   = st.session_state["lib_resultado"]
            res = d.get("resumen", {})
            tipo_eq = d.get("tipo_equipo","")
            if tipo_eq == "INTERRUPTOR":
                elabel = "Interruptor"
            elif tipo_eq == "LINEA_DOBLE":
                elabel = "Líneas"
            elif tipo_eq == "TRANSFORMADOR":
                elabel = "Transformador"
            else:
                elabel = "Línea"
            st.markdown(f"""<div class="lib-card" style="padding:1.25rem 1.5rem;">
              <div style="margin-bottom:0.75rem;">
                <span style="font-size:1.6rem;color:#f5a623;font-weight:700;letter-spacing:-0.02em;">{d.get('libranza_id','—')}</span>
                &nbsp;&nbsp;
                {_badge_lib(tipo_eq)}{_badge_lib(d.get('tipo_libranza',''))}{_badge_lib(d.get('ultimo_estado',''))}
              </div>
              <div style="display:flex;gap:1.5rem;flex-wrap:wrap;align-items:center;">
                <div>
                  <span style="font-size:11px;color:#636b78;text-transform:uppercase;letter-spacing:0.07em;">S/E</span><br>
                  <strong style="font-size:16px;color:#e8e4dc;">{d.get('subestacion','—')}</strong>
                </div>
                <div style="width:1px;height:32px;background:rgba(255,255,255,0.1);"></div>
                <div>
                  <span style="font-size:11px;color:#636b78;text-transform:uppercase;letter-spacing:0.07em;">{elabel}</span><br>
                  <strong style="font-size:16px;color:#e8e4dc;">{d.get('equipo_principal','—')}</strong>
                </div>
                <div style="width:1px;height:32px;background:rgba(255,255,255,0.1);"></div>
                <div>
                  <span style="font-size:11px;color:#636b78;text-transform:uppercase;letter-spacing:0.07em;">Responsable</span><br>
                  <strong style="font-size:16px;color:#e8e4dc;">{d.get('responsable_campo','—')}</strong>
                </div>
              </div>
            </div>""", unsafe_allow_html=True)

            if d.get("ultimo_estado","").upper()=="CORREGIR" and d.get("comentarios_correccion"):
                st.markdown(f'<div class="lib-coment"><strong>📝 Comentarios de corrección:</strong><br>{d["comentarios_correccion"]}</div>', unsafe_allow_html=True)

            m1, m2, m3 = st.columns(3)
            m1.metric("Verificaciones",  res.get("total_verificaciones",0))
            m2.metric("✓ Correctas",     res.get("total_ok",0))
            m3.metric("⚠ Observaciones", res.get("total_observaciones",0))

            # ── Validación contra lista maestra de equipos ──
            eq_result = validar_equipo_en_lista(
                d.get("equipo_principal",""),
                d.get("subestacion",""),
                d.get("tipo_equipo","")
            )
            eq_color = "#4cde8f" if eq_result["estado"]=="OK" else ("#f06060" if eq_result["estado"]=="ERROR" else "#f5a623")
            eq_icono = "✓" if eq_result["estado"]=="OK" else ("✗" if eq_result["estado"]=="ERROR" else "⚠")
            eq_cls   = "lib-item-ok" if eq_result["estado"]=="OK" else ("lib-item-err" if eq_result["estado"]=="ERROR" else "lib-item-warn")
            with st.expander(f"🗂️ Validación contra Lista Maestra — {eq_result['estado']}", expanded=True):
                st.markdown(f'<div class="{eq_cls}"><span style="font-size:12px;color:#8a9099;font-weight:500;">Equipo en lista maestra</span><br><span style="color:{eq_color};font-size:15px;">{eq_icono} {eq_result["detalle"]}</span></div>', unsafe_allow_html=True)

            dg = [x for x in d.get("datos_generales",[]) if not _campo_oculto(x["campo"])]
            obs_dg = sum(1 for x in dg if x["estado"]!="OK")
            with st.expander(f"📋 Datos Generales — {'✓ Sin observaciones' if obs_dg==0 else f'⚠ {obs_dg} observación(es)'}", expanded=True):
                for item in dg:
                    st.markdown(_item_dg(item["campo"],item["detalle"],item["estado"]), unsafe_allow_html=True)

            # ── Validación de agentes vs DB ──────────────────────────────────
            _eq_db = cargar_equipos_libranzas()
            _ag_val = validar_agentes_vs_db(
                d.get("equipo_principal",""),
                d.get("datos_generales", []),
                _eq_db
            )
            if _ag_val["estado"] != "SIN_DATOS":
                _ag_color = "#4cde8f" if _ag_val["estado"]=="OK" else "#f5a623"
                _ag_ic    = "✓" if _ag_val["estado"]=="OK" else "⚠"
                _ag_cls   = "lib-item-ok" if _ag_val["estado"]=="OK" else "lib-item-warn"
                with st.expander(f"🏢 Validación de Agentes (DB) — {_ag_val['estado']}", expanded=(_ag_val["estado"]!="OK")):
                    presentes = _ag_val.get("presentes", [])
                    faltantes = _ag_val.get("faltantes", [])
                    if presentes:
                        st.markdown(
                            "".join(f'<div class="lib-item-ok"><span style="color:#4cde8f;">✓ {ag} — presente en la libranza</span></div>' for ag in presentes),
                            unsafe_allow_html=True
                        )
                    if faltantes:
                        st.markdown(
                            "".join(f'<div class="lib-item-warn"><span style="color:#f5a623;">⚠ {ag} — registrado en DB pero NO aparece en la libranza como agente involucrado/informado</span></div>' for ag in faltantes),
                            unsafe_allow_html=True
                        )
                    if not presentes and not faltantes:
                        st.markdown(f'<div class="lib-item-ok"><span style="color:#4cde8f;">{_ag_val["detalle"]}</span></div>', unsafe_allow_html=True)

            # ── Maniobras con validación DB por paso ─────────────────────────
            _subestacion = d.get("subestacion", "")
            mi = enriquecer_pasos_con_db(d.get("maniobras_inicio",[]), _subestacion, _eq_db)
            mf = enriquecer_pasos_con_db(d.get("maniobras_finalizacion",[]), _subestacion, _eq_db)

            # Validación cruzada: consistencia de responsables entre todos los pasos
            validar_consistencia_responsables(mi + mf)

            def _contar_db_obs(pasos):
                return sum(
                    1 for x in pasos
                    if any(v.get("estado") in ("ERROR","ADVERTENCIA")
                           for v in x.get("validaciones_db", []))
                )

            err_mi    = sum(1 for x in mi if x["estado"]=="ERROR")
            db_obs_mi = _contar_db_obs(mi)
            _label_mi = "✓ Sin errores" if err_mi==0 else f"✗ {err_mi} error(es)"
            if db_obs_mi:
                _label_mi += f" · 🗂 {db_obs_mi} obs. DB"
            with st.expander(f"🔓 Maniobras de Inicio — {_label_mi}", expanded=True):
                for paso in mi:
                    st.markdown(_item_paso(paso), unsafe_allow_html=True)

            err_mf    = sum(1 for x in mf if x["estado"]=="ERROR")
            db_obs_mf = _contar_db_obs(mf)
            _label_mf = "✓ Sin errores" if err_mf==0 else f"✗ {err_mf} error(es)"
            if db_obs_mf:
                _label_mf += f" · 🗂 {db_obs_mf} obs. DB"
            with st.expander(f"🔒 Maniobras de Finalización — {_label_mf}", expanded=True):
                for paso in mf:
                    st.markdown(_item_paso(paso), unsafe_allow_html=True)





def vista_vertimiento(archivo_bitacora):
    st.title("🌊 CENTRALES EN VERTIMIENTO")

    PLANTILLA_VERT = (lambda f: f if os.path.exists(f) else os.path.join(os.getcwd(), "CENTRALES_EN_VERTIMIENTO.xlsm"))(os.path.join(os.path.dirname(os.path.abspath(__file__)), "CENTRALES_EN_VERTIMIENTO.xlsm"))

    def _norm_v(t):
        return unicodedata.normalize('NFKD', str(t)).encode('ascii', 'ignore').decode('ascii').upper()

    def _extraer_vertimientos(df_bit):
        # Seguimiento cronológico: el último evento por central es el que manda.
        # Un FINALIZO anterior no debe cancelar un INICIO posterior.
        estado = {}  # central -> {'activo': bool, 'fecha': str}
        for _, row in df_bit.iterrows():
            texto = " ".join(str(x) for x in row.values if str(x) != 'nan')
            tn = _norm_v(texto)
            if "VERTIMIENTO" not in tn:
                continue
            m_fecha = re.search(r'(\d{2}/\d{2}/\d{4})\s*(\d{2}:\d{2})', texto)
            fecha_str = f"{m_fecha.group(1)} {m_fecha.group(2)}" if m_fecha else ""
            m_central = re.search(r'(?:INICIO\(A\)|FINALIZO\(A\))\s+(.+?)\s+(?:PRESA|CAMARA)', tn)
            if not m_central:
                continue
            central = m_central.group(1).strip()
            if "FINALIZO" in tn:
                estado[central] = {'activo': False, 'fecha': ''}
            else:
                estado[central] = {'activo': True, 'fecha': fecha_str}
        return {c: v['fecha'] for c, v in estado.items() if v['activo']}

    df_bit = pd.read_excel(archivo_bitacora, header=None)
    vertimientos = _extraer_vertimientos(df_bit)

    _cols_esp = {"SELECCIONAR","CENTRAL","FECHA INICIO VERTIMIENTO","CONDICION ACTUAL","OBSERVACIONES"}
    _vc_actual = st.session_state.get('vert_candidatos', pd.DataFrame())
    if _vc_actual.empty or not _cols_esp.issubset(set(_vc_actual.columns)):
        st.session_state['vert_candidatos'] = pd.DataFrame()

    if st.button("🔍 ANALIZAR BITÁCORA", type="primary"):
        if vertimientos:
            with sqlite3.connect('vertimiento.db') as conn:
                existentes = set(pd.read_sql_query(
                    "SELECT central FROM reporte_vertimiento", conn)['central'].str.upper().tolist())
            filas = []
            for c, f in vertimientos.items():
                obs = "⚠️ Ya existe en BD" if c in existentes else ""
                filas.append({"SELECCIONAR": False, "CENTRAL": c,
                               "FECHA INICIO VERTIMIENTO": f,
                               "CONDICION ACTUAL": "", "OBSERVACIONES": obs})
            st.session_state['vert_candidatos'] = pd.DataFrame(filas)
            st.success(f"✅ {len(vertimientos)} central(es) en vertimiento detectada(s).")
        else:
            st.info("ℹ️ No se detectaron centrales en vertimiento activo.")
            st.session_state['vert_candidatos'] = pd.DataFrame()

    if not st.session_state['vert_candidatos'].empty:
        st.subheader("📝 Detecciones Recientes (Selecciona para guardar)")
        df_cand = st.data_editor(
            st.session_state['vert_candidatos'],
            column_config={
                "SELECCIONAR":              st.column_config.CheckboxColumn("Guardar", default=False),
                "CENTRAL":                  st.column_config.TextColumn(disabled=True),
                "FECHA INICIO VERTIMIENTO": st.column_config.TextColumn(disabled=True),
                "CONDICION ACTUAL":         st.column_config.TextColumn(width="large"),
                "OBSERVACIONES":            st.column_config.TextColumn(disabled=True, width="small"),
            },
            hide_index=True, use_container_width=True, key="vert_cand_editor"
        )
        if st.button("💾 Agregar a Base de Datos", type="secondary"):
            sel = df_cand[df_cand['SELECCIONAR'] == True]
            if not sel.empty:
                with sqlite3.connect('vertimiento.db') as conn:
                    for _, r in sel.iterrows():
                        conn.execute("""INSERT INTO reporte_vertimiento (central,fecha_inicio,condicion)
                            VALUES (?,?,?) ON CONFLICT(central) DO UPDATE SET
                            fecha_inicio=excluded.fecha_inicio, condicion=excluded.condicion""",
                            (r['CENTRAL'], r['FECHA INICIO VERTIMIENTO'], r['CONDICION ACTUAL']))
                st.success(f"✅ {len(sel)} central(es) guardada(s)."); st.rerun()
            else:
                st.warning("⚠️ No seleccionaste ninguna fila.")
        st.divider()

    st.subheader("💾 Base de Datos de Centrales en Vertimiento")

    with st.expander("📥 Importar desde archivo de Vertimiento (.xlsm / .xlsx)"):
        st.caption("Sube el archivo CENTRALES_EN_VERTIMIENTO.xlsm. Lee el sheet 'VERTIMIENTO-1' desde fila 12.")
        archivo_import_v = st.file_uploader("Selecciona el archivo",
                                            type=["xlsm","xlsx","xls"], key="import_excel_vert")
        modo_import_v = st.radio("Modo de importación:",
                                 ["Agregar (no sobreescribir existentes)", "Reemplazar todo"],
                                 key="modo_import_vert")
        if archivo_import_v and st.button("📤 Importar", type="primary", key="btn_import_vert"):
            try:
                wb_imp = openpyxl.load_workbook(archivo_import_v, data_only=True, keep_vba=False)
                ws_imp = wb_imp["VERTIMIENTO-1"] if "VERTIMIENTO-1" in wb_imp.sheetnames else wb_imp.active
                filas_imp = []
                for row in ws_imp.iter_rows(min_row=12, values_only=True):
                    central = row[0] if row else None
                    if not central or str(central).strip() in ("", "None", "CENTRAL"):
                        continue
                    central   = str(central).strip().upper()
                    fecha_ini = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                    condicion = str(row[4]).strip() if len(row) > 4 and row[4] else ""
                    filas_imp.append((central, fecha_ini, condicion))
                wb_imp.close()
                if not filas_imp:
                    st.warning("No se encontraron filas con datos.")
                else:
                    with sqlite3.connect('vertimiento.db') as conn:
                        if modo_import_v == "Reemplazar todo":
                            conn.execute("DELETE FROM reporte_vertimiento")
                            for f in filas_imp:
                                conn.execute("INSERT INTO reporte_vertimiento (central,fecha_inicio,condicion) VALUES (?,?,?)", f)
                        else:
                            for f in filas_imp:
                                conn.execute("INSERT OR IGNORE INTO reporte_vertimiento (central,fecha_inicio,condicion) VALUES (?,?,?)", f)
                    st.success(f"✅ {len(filas_imp)} central(es) importada(s)."); st.rerun()
            except Exception as e:
                st.error(f"Error al importar: {e}")

    with st.expander("➕ Agregar Central Manualmente"):
        mc1, mc2 = st.columns([1, 2])
        with mc1:
            m_central = st.text_input("Central", key="vert_m_central")
            m_fecha   = st.text_input("Fecha Inicio Vertimiento", key="vert_m_fecha",
                                      placeholder="DD/MM/YYYY HH:MM")
        with mc2:
            m_cond = st.text_area("Condición Actual", key="vert_m_cond")
        if st.button("➕ Agregar a Candidatos", key="vert_btn_manual"):
            if m_central:
                nueva = {
                    "SELECCIONAR": False,
                    "CENTRAL": m_central.upper(),
                    "FECHA INICIO VERTIMIENTO": m_fecha,
                    "CONDICION ACTUAL": m_cond,
                    "OBSERVACIONES": "✏️ Agregado manualmente",
                }
                df_actual = st.session_state.get('vert_candidatos', pd.DataFrame())
                # Evitar duplicados en la tabla candidatos
                if not df_actual.empty and m_central.upper() in df_actual['CENTRAL'].str.upper().values:
                    df_actual.loc[df_actual['CENTRAL'].str.upper() == m_central.upper()] = list(nueva.values())
                    st.session_state['vert_candidatos'] = df_actual
                    st.info(f"ℹ️ {m_central.upper()} actualizado en la tabla de candidatos.")
                else:
                    st.session_state['vert_candidatos'] = pd.concat(
                        [df_actual, pd.DataFrame([nueva])], ignore_index=True
                    )
                    st.success(f"✅ {m_central.upper()} agregado a la tabla de candidatos. Selecciónalo y pulsa 'Agregar a Base de Datos'.")
                st.rerun()
            else:
                st.warning("⚠️ Escribe el nombre de la central.")

    try:
        with sqlite3.connect('vertimiento.db') as conn:
            df_db = pd.read_sql_query("SELECT rowid, * FROM reporte_vertimiento", conn)
        if not df_db.empty:
            # Mantener solo las 3 columnas operativas
            df_db = df_db[["rowid", "central", "fecha_inicio", "condicion"]]
            df_db.columns = ["ID", "CENTRAL", "FECHA INICIO VERTIMIENTO", "CONDICION ACTUAL"]
            df_edit = st.data_editor(df_db, hide_index=True, use_container_width=True,
                                     num_rows="dynamic", column_config={"ID": None})
            col1, _ = st.columns([1, 4])
            with col1:
                if st.button("💾 Guardar Cambios", key="vert_btn_save"):
                    with sqlite3.connect('vertimiento.db') as conn:
                        conn.execute("DELETE FROM reporte_vertimiento")
                        df_save = df_edit[["CENTRAL","FECHA INICIO VERTIMIENTO","CONDICION ACTUAL"]].copy()
                        df_save.columns = ["central","fecha_inicio","condicion"]
                        df_save.to_sql('reporte_vertimiento', conn, if_exists='append', index=False)
                    st.toast("¡Base de datos actualizada!"); st.rerun()
            st.markdown("---")
            if st.button("📥 Generar Plantilla", type="primary", key="vert_btn_gen"):
                if not os.path.exists(PLANTILLA_VERT):
                    st.error("⚠️ No se encontró CENTRALES_EN_VERTIMIENTO.xlsm en la carpeta de la app.")
                else:
                    try:
                        import copy as _copy
                        wb = openpyxl.load_workbook(PLANTILLA_VERT, keep_vba=True)
                        ws = wb["VERTIMIENTO-1"]
                        FILA_TMPL = 11  # Fila de referencia con formato completo

                        def _copiar_estilo(src, dst):
                            """Copia fuente, relleno, bordes, alineación y formato de número."""
                            dst.font         = _copy.copy(src.font)
                            dst.fill         = _copy.copy(src.fill)
                            dst.border       = _copy.copy(src.border)
                            dst.alignment    = _copy.copy(src.alignment)
                            dst.number_format = src.number_format

                        def _ajustar_formula(formula, fila_orig, fila_dest):
                            """Reemplaza referencias de fila en fórmulas Excel (A11 → A12, etc.)."""
                            if formula and str(formula).startswith('='):
                                return re.sub(
                                    r'([A-Z]+)' + str(fila_orig),
                                    lambda m: m.group(1) + str(fila_dest),
                                    str(formula)
                                )
                            return formula

                        alto_ref = ws.row_dimensions[FILA_TMPL].height or 29.25

                        for enum_i, (_, r) in enumerate(df_edit.iterrows()):
                            fila = FILA_TMPL + enum_i
                            # 1. Copiar formato y fórmulas de la fila plantilla a esta fila
                            for col_n in range(1, 9):
                                src_cell = ws.cell(FILA_TMPL, col_n)
                                dst_cell = ws.cell(fila, col_n)
                                _copiar_estilo(src_cell, dst_cell)
                                # Copiar fórmulas ajustando el número de fila
                                if src_cell.value and str(src_cell.value).startswith('='):
                                    dst_cell.value = _ajustar_formula(
                                        src_cell.value, FILA_TMPL, fila)
                            # 2. Altura de fila igual a la plantilla
                            ws.row_dimensions[fila].height = alto_ref
                            # 3. Escribir valores de la BD (sobreescriben el valor de plantilla)
                            ws.cell(fila, 1).value = r["CENTRAL"]
                            ws.cell(fila, 2).value = r["FECHA INICIO VERTIMIENTO"]
                            ws.cell(fila, 5).value = r["CONDICION ACTUAL"]

                        output = BytesIO()
                        wb.save(output)
                        st.download_button("⬇️ Descargar archivo", output.getvalue(),
                                           file_name="CENTRALES EN VERTIMIENTO.xlsm",
                                           mime="application/vnd.ms-excel.sheet.macroEnabled.12")
                    except Exception as e:
                        st.error(f"Error al generar la plantilla: {e}")
        else:
            st.info("ℹ️ No hay centrales registradas en la base de datos.")
    except Exception as e:
        st.error(f"Error al leer la base de datos: {e}")



# --- 8b. NOTIFICACIÓN ASEP ---

# Mapeo circuito -> area: cargado dinamicamente de CIRCUITOS.xlsx
_CIRC_DB_CACHE = None

def _cargar_db_circuitos():
    global _CIRC_DB_CACHE
    if _CIRC_DB_CACHE is not None:
        return _CIRC_DB_CACHE
    ruta = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'CIRCUITOS.xlsx')
    if not os.path.exists(ruta):
        ruta = os.path.join(os.getcwd(), 'CIRCUITOS.xlsx')
    db = {}
    if os.path.exists(ruta):
        try:
            wb_c = openpyxl.load_workbook(ruta, data_only=True)
            ws_c = wb_c.active
            inst_a = None
            agente_a = None
            for r in range(2, ws_c.max_row + 1):
                codigo = ws_c.cell(r, 3).value
                inst   = ws_c.cell(r, 4).value
                agente = ws_c.cell(r, 5).value
                if inst:   inst_a   = str(inst).strip().upper()
                if agente: agente_a = str(agente).strip().upper()
                if codigo and inst_a:
                    key = re.sub('[^A-Z0-9]', '', str(codigo).strip().upper())
                    if key not in db:
                        db[key] = {'instalacion': inst_a, 'agente': agente_a or ''}
            wb_c.close()
        except Exception:
            pass
    _CIRC_DB_CACHE = db
    return db

def _circuito_a_area(circ):
    db = _cargar_db_circuitos()
    c = re.sub('[^A-Z0-9]', '', circ.strip().upper())
    if c in db:
        return db[c]['instalacion']
    # Prefijo exacto: el codigo mas largo de la DB que sea prefijo de c
    matches = [k for k in db if c.startswith(k)]
    if matches:
        return db[max(matches, key=len)]['instalacion']
    # Fallback: extraer solo letras iniciales de c y buscar coincidencia parcial
    alpha = re.match('[A-Z]+', c)
    if alpha:
        pfx = alpha.group(0)
        for n in range(len(pfx), 1, -1):  # de largo a corto
            sub = pfx[:n]
            matches2 = [k for k in db if k.startswith(sub) and re.match('[A-Z]+', k) and re.match('[A-Z]+', k).group(0) == sub]
            if matches2:
                return db[max(matches2, key=len)]['instalacion']
    return circ.strip().upper()

def _circuito_a_agente(circ):
    db = _cargar_db_circuitos()
    c = re.sub('[^A-Z0-9]', '', circ.strip().upper())
    if c in db:
        return db[c]['agente']
    matches = [k for k in db if c.startswith(k)]
    if matches:
        return db[max(matches, key=len)]['agente']
    return ''

def _extraer_circuitos_bf(texto):
    tokens = re.split(r'[,/\s]+', texto.upper())
    result = []
    for t in tokens:
        t = t.strip().rstrip('.')
        if not t:
            continue
        if re.match(r'^[A-Z0-9]+-[A-Z0-9]+$', t) or re.match(r'^[A-Z]{2,4}[0-9]+$', t):
            result.append(t)
    return result

_ELABORADORES_ASEP = {
    'A': 'Ing. Joel Visuette',      'B': 'Ing. Ronny Pinto',
    'C': 'Ing. Juan Bethancourt',   'E': 'Ing. Jason Cortes',
    'F': 'Ing. Jorge Rosas',        'G': 'Ing. Gilberto Garrido',
    'H': 'Ing. Ricardo Pinzon',     'J': 'Ing. Rolando Rodriguez',
    'L': 'Ing. Leonel Contreras',   'M': 'Ing. Javier Rodriguez',
    'N': 'Ing. Michael Nipple',     'O': 'Ing. Eric Serrano',
    'Q': 'Ing. Fernando Lopez',     'S': 'Ing. Helder Franco',
    'T': 'Ing. Ernesto Tamayo',     'Z': 'Ing. David Gonzalez',
}

def _parsear_eventos_asep(df_bit):
    import unicodedata as _ud
    def _norm(t):
        return _ud.normalize('NFKD', str(t)).encode('ascii','ignore').decode().upper()

    def _hmin(h):
        try: return int(h[:2])*60+int(h[3:])
        except: return -1

    def _det_empresa(t):
        if 'EDEMET' in t or 'EDECHI' in t or 'NATURGY' in t: return 'EDEMET'
        if 'ENSA' in t: return 'ENSA'
        return 'OTRA'

    def _similar(a, b):
        a2 = re.sub('[^A-Z]', '', a.upper())
        b2 = re.sub('[^A-Z]', '', b.upper())
        if a2 == b2: return True
        if a2 in b2 or b2 in a2: return True
        if abs(len(a2)-len(b2)) <= 2:
            eq = sum(c1==c2 for c1,c2 in zip(a2,b2))
            if eq >= min(len(a2),len(b2)) - 2 and min(len(a2),len(b2)) > 4:
                return True
        return False

    def _add_area(d, area_str):
        partes = re.split(r',|\bY\b', area_str)
        for p in partes:
            p = p.strip().title()
            if not p: continue
            if not any(_similar(p, ex) for ex in d['areas']):
                d['areas'].append(p)

    def _extraer_areas(tnj):
        """Extrae el string de áreas afectadas en múltiples formatos posibles."""
        # Formato 1: AREAS AFECTADAS [texto]  o  AREAS AFECTADAS (texto)
        m = re.search(r'AREAS\s+AFECTADAS?\s*[:\-]?\s*[\[({]([^\]})]+)[\])}]', tnj)
        if m:
            return m.group(1)
        # Formato 2: AREAS AFECTADAS: texto libre hasta / | o fin de línea
        m = re.search(r'AREAS?\s+AFECTADAS?\s*[:\-]\s*([A-Z][A-Z,\s\-]+?)(?:\s*[/|]|$)', tnj)
        if m:
            return m.group(1)
        # Formato 3: AREA: texto (forma corta)
        m = re.search(r'\bAREA\s*[:\-]\s*([A-Z][A-Z,\s\-]+?)(?:\s*[/|]|$)', tnj)
        if m:
            return m.group(1)
        return None

    def _extraer_texto_circuitos(tnj):
        """Extrae el texto de lista de circuitos del 1er escalon en múltiples formatos."""
        # Formato 1: CIRCUITOS (X, Y, Z) o CIRCUITOS( X, Y, Z) o CIRCUITOS X, Y, Z
        m = re.search(r'CIRCUITOS\s*\(?\s*([^)]+?)\s*\)?[.\s]*(?:CARGA|INFORMA|CON UN TOTAL|MONTO)', tnj)
        if m: return m.group(1)
        m = re.search(r'CIRCUITOS\s*\(?\s*([A-Z0-9,\s/\.\-]+?)(?:\s*/\s*(?:2|SEGUNDO)|\s+MONTO|\s+INFORMA|$)', tnj)
        if m: return m.group(1)
        # Formato 2: CTOS: X, Y, Z  (abreviación usada por ENSA)
        m = re.search(r'CTOS\s*:\s*([A-Z0-9,\s\.\-]+?)(?:\s*\(\d|\s+Y\s+\d\s+ESCALON|\s+MONTO|\s+INFORMA|$)', tnj)
        if m: return m.group(1)
        # Formato 3: 1ER/PRIMER ESCALON= X, Y, Z/  (MW) ...
        # Termina en: /  (MW)  |  / 2DO  |  MONTO  |  INFORMA
        m = re.search(
            r'(?:1ER?|1|PRIMER)\s+ESCALON\s*[=:]\s*'
            r'([A-Z0-9,\s\.\-]+?)'
            r'(?:\s*/\s*(?:\(|\d|2DO|SEGUNDO)|\s+MONTO|\s+INFORMA|$)',
            tnj
        )
        if m: return m.group(1)
        # Formato 4 (fallback genérico): cualquier ESCALON= X, Y
        m = re.search(
            r'ESCALON[^=:]*[=:]\s*'
            r'([A-Z0-9,\s/\.\-]+?)'
            r'(?:\s*/\s*(?:CARGA|\(|\d|2DO)|\s+MONTO|\s+CON UN TOTAL|\s+INFORMO|$)',
            tnj
        )
        if m: return m.group(1)
        return None

    rows = []
    for _, row in df_bit.iterrows():
        rows.append(" ".join(str(x) for x in row.values if str(x) not in ('nan','')))

    # ── Paso 1: identificar todas las filas trigger con fecha/hora ────────────
    # Trigger = SPEAR ACTIVADO | BLOQUE DE CARGA ACTIVADO | ESQUEMA BF ACTIVADO | ESQUEMA BAJO VOLTAJE ACTIVADO
    triggers = []
    for idx, tj in enumerate(rows):
        tn = _norm(tj)
        is_trig = (
            ('ACTIVADO' in tn and 'SPEAR' in tn) or
            ('ACTIVADO' in tn and 'BLOQUE DE CARGA' in tn) or
            ('ACTIVADO' in tn and 'ESQUEMA BAJA FRECUENCIA' in tn) or
            ('ACTIVADO' in tn and 'ESQUEMA BAJO VOLTAJE' in tn)
        )
        if not is_trig:
            continue
        m = re.search(r'(\d{2}/\d{2}/\d{4})\s+(\d{2}:\d{2})', tj)
        if not m:
            continue
        triggers.append({'idx': idx, 'fecha': m.group(1), 'hora': m.group(2),
                         'hmin': _hmin(m.group(2)), 'texto': tn})

    if not triggers:
        return []

    # ── Paso 2: agrupar triggers del mismo evento (ventana 15 min mismo día) ─
    grupos = []
    usados = set()
    for ti, tr in enumerate(triggers):
        if ti in usados:
            continue
        grp = [tr]
        usados.add(ti)
        for tj2, tr2 in enumerate(triggers):
            if tj2 in usados: continue
            if tr2['fecha'] == tr['fecha'] and abs(tr2['hmin'] - tr['hmin']) <= 15:
                grp.append(tr2)
                usados.add(tj2)
        grupos.append(grp)

    # ── Paso 3: por cada grupo construir un evento ────────────────────────────
    eventos = []
    for grp in grupos:
        fecha_evento = grp[0]['fecha']
        hora_ini     = min(g['hora'] for g in grp)      # hora más temprana del grupo
        h_ini_min    = _hmin(hora_ini)

        # Ventana de análisis: desde 5 filas antes del primer trigger hasta +50 filas después del último
        idx_min = min(g['idx'] for g in grp)
        idx_max = max(g['idx'] for g in grp)
        j_start = max(0, idx_min - 5)
        j_end   = min(len(rows), idx_max + 200)

        empresas_data = {}
        hora_fin_norm = None
        contexto_textos = []

        for j in range(j_start, j_end):
            tj  = rows[j]
            tnj = _norm(tj)
            m_h = re.search(r'\d{2}/\d{2}/\d{4}\s+(\d{2}:\d{2})', tj)
            h_min_j = _hmin(m_h.group(1)) if m_h else h_ini_min
            if h_min_j > h_ini_min + 60:
                break
            contexto_textos.append(tj)

            # BLOQUE DE CARGA (SPEAR) — unificado: cualquier combinación de SPEAR o BLOQUE DE CARGA
            es_bloque_spear = (
                ('BLOQUE DE CARGA' in tnj or 'SPEAR' in tnj) and 'ACTIVADO' in tnj
            )
            if es_bloque_spear:
                area_str = _extraer_areas(tnj)

                # ── Caso multi-empresa en una línea ──────────────────────────
                # Ej: "NATURGY QUE DESLIGA 22 MW Y EN ENSA ... 4.5 MW. TOTAL = 26.5 MW"
                # Detectar pares (empresa, mw) explícitos en el texto
                _multi = re.findall(
                    r'(NATURGY|EDEMET|EDECHI|ENSA)\b[^0-9]{0,40}?(\d+(?:\.\d+)?)\s*MW',
                    tnj
                )
                if _multi:
                    for _emp_raw, _mw_raw in _multi:
                        _emp = 'EDEMET' if _emp_raw in ('NATURGY','EDEMET','EDECHI') else 'ENSA'
                        _d = empresas_data.setdefault(_emp, {'mw': 0, 'areas': []})
                        _d['mw'] += float(_mw_raw)
                        if area_str:
                            _add_area(_d, area_str)
                else:
                    # Caso empresa única — usar TOTALIZANDO o primer MW encontrado
                    emp = _det_empresa(tnj)
                    mw_m = (re.search(r'TOTALIZANDO\s+(\d+(?:\.\d+)?)\s*MW', tnj) or
                            re.search(r'TOTAL\s*=\s*(\d+(?:\.\d+)?)\s*MW', tnj) or
                            re.search(r'(\d+(?:\.\d+)?)\s*MW', tnj))
                    mw = float(mw_m.group(1)) if mw_m else 0
                    # Solo crear bucket si hay datos reales. Las líneas puras de trigger
                    # (ej: "SPEAR CONTINGENCIA GAT C6") no tienen empresa, MW ni áreas
                    # y generarían un bucket vacío de 'OTRA'.
                    if mw == 0 and not area_str:
                        pass  # nada que registrar
                    else:
                        d = empresas_data.setdefault(emp, {'mw': 0, 'areas': []})
                        d['mw'] += mw
                        if area_str:
                            _add_area(d, area_str)
                        # Circuitos con prefijo compartido: (CEB-9,10,11,...) → CEB-9, CEB-10, ...
                        _paren_m = re.search(r'\(([A-Z]+-\d+(?:,\d+)+)\)', tnj)
                        if _paren_m:
                            _parts = _paren_m.group(1).split(',')
                            _pfx = re.match(r'([A-Z]+-)', _parts[0])
                            if _pfx:
                                _circs_exp = [_parts[0]] + [_pfx.group(1) + n for n in _parts[1:]]
                                for _c in _circs_exp:
                                    _add_area(d, _circuito_a_area(_c))

            # ESQUEMA BAJA FRECUENCIA
            elif 'ESQUEMA BAJA FRECUENCIA' in tnj and 'ACTIVADO' in tnj:
                emp_texto = _det_empresa(tnj)
                mw_m = re.search(r'(\d+(?:\.\d+)?)\s*MW', tnj)
                mw = float(mw_m.group(1)) if mw_m else 0
                circ_txt = _extraer_texto_circuitos(tnj)
                circs = _extraer_circuitos_bf(circ_txt) if circ_txt else []
                area_str = _extraer_areas(tnj)
                # Asignar MW al agente detectado en el texto
                d = empresas_data.setdefault(emp_texto, {'mw': 0, 'areas': []})
                d['mw'] += mw
                if area_str:
                    _add_area(d, area_str)
                # Asignar cada circuito a su empresa según la DB (no depender del texto)
                for c in circs:
                    emp_circ = _circuito_a_agente(c) or emp_texto
                    area_circ = _circuito_a_area(c)
                    # Solo agregar si la función devuelve un área real (no el propio código)
                    if area_circ and area_circ != c.strip().upper():
                        dc = empresas_data.setdefault(emp_circ, {'mw': 0, 'areas': []})
                        _add_area(dc, area_circ)

            # ESQUEMA BAJO VOLTAJE
            elif 'ESQUEMA BAJO VOLTAJE' in tnj and 'ACTIVADO' in tnj:
                emp_texto = _det_empresa(tnj)
                mw_m = re.search(r'(\d+(?:\.\d+)?)\s*MW', tnj)
                mw = float(mw_m.group(1)) if mw_m else 0
                circ_txt = _extraer_texto_circuitos(tnj)
                circs = _extraer_circuitos_bf(circ_txt) if circ_txt else []
                area_str = _extraer_areas(tnj)
                d = empresas_data.setdefault(emp_texto, {'mw': 0, 'areas': []})
                d['mw'] += mw
                if area_str:
                    _add_area(d, area_str)
                for c in circs:
                    emp_circ = _circuito_a_agente(c) or emp_texto
                    area_circ = _circuito_a_area(c)
                    if area_circ and area_circ != c.strip().upper():
                        dc = empresas_data.setdefault(emp_circ, {'mw': 0, 'areas': []})
                        _add_area(dc, area_circ)

            # NORMALIZADO: ESQUEMA BF, BAJO VOLTAJE o SPEAR CONTINGENCIA → hora más tardía
            if 'NORMALIZADO' in tnj and m_h:
                if any(k in tnj for k in ['ESQUEMA BAJA FRECUENCIA','ESQUEMA BAJO VOLTAJE','SPEAR']):
                    h_n = m_h.group(1)
                    if hora_fin_norm is None or _hmin(h_n) > _hmin(hora_fin_norm):
                        hora_fin_norm = h_n

        if not empresas_data:
            continue

        tiempo_min = None
        if hora_fin_norm:
            diff = _hmin(hora_fin_norm) - h_ini_min
            if diff > 0:
                tiempo_min = diff

        eventos.append({
            'fecha':      fecha_evento,
            'hora_ini':   hora_ini,
            'hora_fin':   hora_fin_norm,
            'tiempo_min': tiempo_min,
            'empresas':   empresas_data,
            'contexto':   '\n'.join(contexto_textos[:50]),
        })
    return eventos

def _llamar_claude_causa(contexto):
    import json
    client = anthropic.Anthropic(api_key=_api_key_libranzas)
    resp = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=1000,
        messages=[{
            "role": "user",
            "content": (
                "Eres un ingeniero del Centro Nacional de Despacho de Panama. "
                "A partir del siguiente fragmento de bitacora electrica, redacta 3 versiones "
                "de la causa de interrupcion para un informe formal a la ASEP. "
                "Cada version debe ser concisa (1-3 oraciones), tecnica y en MAYUSCULAS. "
                "Responde UNICAMENTE con un JSON array de 3 strings, sin explicacion adicional, "
                "sin bloques de codigo markdown.\r\n\r\nBITACORA:\r\n" + contexto[:3000]
            )
        }]
    )
    text = resp.content[0].text.strip()
    text = re.sub(r'^```[a-z]*\n?', '', text)
    text = re.sub(r'\n?```$', '', text)
    return json.loads(text)


def _generar_reporte_asep(plantilla_path, num_evento, dia_hora, sectores,
                           causa, tiempo, elaborado_por, tipo_evento, idx, ev):
    """Genera el reporte ASEP rellenando la plantilla .xlsx con openpyxl."""
    import copy as _copy
    from openpyxl.styles import Font, Alignment, Border, Side

    if not os.path.exists(plantilla_path):
        st.error(f"No se encontró {plantilla_path}")
        return

    try:
        wb = openpyxl.load_workbook(plantilla_path)
        ws = wb.active

        # Helper: copiar estilo de celda origen a destino
        def _copy_style(src_cell, dst_cell):
            if src_cell.has_style:
                dst_cell.font      = _copy.copy(src_cell.font)
                dst_cell.fill      = _copy.copy(src_cell.fill)
                dst_cell.border    = _copy.copy(src_cell.border)
                dst_cell.alignment = _copy.copy(src_cell.alignment)
                dst_cell.number_format = src_cell.number_format

        # Helper: resolver celda real (top-left si es MergedCell)
        from openpyxl.cell.cell import MergedCell as _MergedCell
        def _real_cell(row1, col1):
            cell = ws.cell(row1, col1)
            if isinstance(cell, _MergedCell):
                for rng in ws.merged_cells.ranges:
                    if (rng.min_row <= row1 <= rng.max_row and
                            rng.min_col <= col1 <= rng.max_col):
                        return ws.cell(rng.min_row, rng.min_col)
            return cell

        # Helper: escribir valor preservando el estilo existente de la celda
        def _write(row1, col1, value, bold=False, wrap=False):
            """row1/col1 son 1-based (como openpyxl)."""
            cell = _real_cell(row1, col1)
            cell.value = value
            if bold or wrap:
                f = _copy.copy(cell.font) if cell.has_style else Font()
                if bold: f = Font(**{**f.__dict__, "bold": True})
                cell.font = f
            if wrap:
                a = _copy.copy(cell.alignment) if cell.has_style else Alignment()
                cell.alignment = Alignment(wrap_text=True,
                                           horizontal=a.horizontal,
                                           vertical=a.vertical)

        # \u2500\u2500 Nº de evento (fila 3, col H) \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
        _write(3, 8, int(num_evento))

        # -- Tipo de evento: casilla en la primera celda escribible de col>=B (sin pisar etiqueta)
        for _row_chk, _tname in [(17, 'GENERACION'), (18, 'TRANSMISION')]:
            _chk_val = '\u2611' if tipo_evento == _tname else ''
            _written = False
            for _col in range(2, 6):  # Busca en col B..E
                _cell = ws.cell(_row_chk, _col)
                if not isinstance(_cell, _MergedCell):
                    _cell.value = _chk_val
                    _written = True
                    break
                # MergedCell: buscar si su top-left esta en col>=2 (no A)
                for _rng in ws.merged_cells.ranges:
                    if (_rng.min_row <= _row_chk <= _rng.max_row and
                            _rng.min_col <= _col <= _rng.max_col and
                            _rng.min_col >= 2):
                        ws.cell(_rng.min_row, _rng.min_col).value = _chk_val
                        _written = True
                        break
                if _written:
                    break

        # \u2500\u2500 Día / hora (fila 21, col A) \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
        _write(21, 1, dia_hora, bold=True)

        # \u2500\u2500 Sectores (filas 24, 27, 30 \u2192 col A empresa, col B áreas+MW) \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
        _filas_sec = [24, 27, 30]
        for k, (empresa, datos) in enumerate(list(sectores.items())[:3]):
            _f = _filas_sec[k]
            _areas = (datos['areas'].upper() if isinstance(datos['areas'], str)
                      else ', '.join(datos['areas']).upper())
            _write(_f, 1, f'{empresa}:')
            _write(_f, 2, f'{_areas}. CARGA TOTAL = {datos["mw"]} MW.', wrap=True)

        # \u2500\u2500 Causa (fila 32, col A) \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
        _write(32, 1, causa.upper(), wrap=True)

        # \u2500\u2500 Tiempo de afectación (fila 38, col A) \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
        _write(38, 1, tiempo.upper())

        # \u2500\u2500 Elaborado por (fila 44, col A) \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
        _write(44, 1, elaborado_por)

        # \u2500\u2500 Guardar en BD y ofrecer descarga \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
        import json as _json
        with sqlite3.connect('eventos_asep.db') as conn:
            conn.execute(
                "INSERT OR REPLACE INTO eventos (numero, fecha, elaborado_por, datos_json) VALUES (?,?,?,?)",
                (num_evento, ev['fecha'], elaborado_por,
                 _json.dumps({'sectores': sectores, 'causa': causa, 'tiempo': tiempo,
                              'tipo_evento': tipo_evento}))
            )

        output = BytesIO()
        wb.save(output)
        fname = f"Evento N°{num_evento} de 2026.xlsx"
        st.download_button(f"Descargar {fname}", output.getvalue(),
                           file_name=fname,
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key=f"asep_dl_{idx}")
        st.success(f"\u2705 Evento N° {num_evento} generado \u2014 Tipo: {tipo_evento}")

    except Exception as _e:
        import traceback
        st.error(f"Error: {_e}")
        st.code(traceback.format_exc())


def vista_eventos_asep(archivo_bitacora=None):
    st.title("⚡ INTERRUPCIONES ASEP")

    PLANTILLA_ASEP = (lambda f: f if os.path.exists(f) else os.path.join(os.getcwd(), "Evento_N__de_2026.xlsx"))(os.path.join(os.path.dirname(os.path.abspath(__file__)), "Evento_N__de_2026.xlsx"))

    tab_nuevo, tab_hist = st.tabs(["Nuevo Evento", "Historial"])

    _ADMIN_USER = "Dsamuel25"
    _ADMIN_PASS = "Lavacalola2*"

    with tab_hist:
        try:
            with sqlite3.connect('eventos_asep.db') as conn:
                df_hist = pd.read_sql_query(
                    "SELECT numero, fecha, elaborado_por FROM eventos ORDER BY numero DESC", conn)
            if not df_hist.empty:
                st.dataframe(df_hist, hide_index=True, use_container_width=True)
            else:
                st.info("No hay eventos registrados.")
        except Exception as e:
            st.error(str(e))

        st.markdown("---")
        with st.expander("🗑️ Borrar registros (requiere credenciales)"):
            _u = st.text_input("Usuario", key="adm_user")
            _p = st.text_input("Contraseña", type="password", key="adm_pass")
            _borrar_todo = st.checkbox("Borrar TODOS los registros", key="adm_all")
            _num_borrar  = st.number_input("O borrar evento N°", min_value=0, step=1,
                                            key="adm_num",
                                            help="0 = no borrar por número individual")
            if st.button("Confirmar borrado", type="primary", key="adm_btn"):
                if _u == _ADMIN_USER and _p == _ADMIN_PASS:
                    with sqlite3.connect('eventos_asep.db') as conn:
                        if _borrar_todo:
                            conn.execute("DELETE FROM eventos")
                            st.success("✅ Todos los registros eliminados.")
                        elif _num_borrar > 0:
                            conn.execute("DELETE FROM eventos WHERE numero=?", (int(_num_borrar),))
                            st.success(f"✅ Evento N° {int(_num_borrar)} eliminado.")
                        else:
                            st.warning("Selecciona qué borrar.")
                    st.rerun()
                else:
                    st.error("Credenciales incorrectas.")

    with tab_nuevo:
        if archivo_bitacora is None:
            st.info("Sube la bitacora en el menu lateral para habilitar este reporte.")
            return

        # Invalidar caché automáticamente si la bitácora cambió
        _asep_file_id = f"{archivo_bitacora.name}_{archivo_bitacora.size}"
        if st.session_state.get('asep_file_id') != _asep_file_id:
            st.session_state.pop('asep_eventos', None)
            st.session_state.pop('asep_contexto_global', None)
            st.session_state['asep_file_id'] = _asep_file_id

        if st.button("ANALIZAR BITACORA", type="primary", key="asep_analizar"):
            df_bit = pd.read_excel(archivo_bitacora, header=None)
            eventos = _parsear_eventos_asep(df_bit)
            st.session_state['asep_eventos'] = eventos
            st.session_state['asep_contexto_global'] = '\n'.join(
                ' '.join(str(x) for x in row if str(x) != 'nan')
                for _, row in df_bit.iterrows()
            )

        eventos = st.session_state.get('asep_eventos', [])
        if not eventos:
            return

        st.success(f"{len(eventos)} evento(s) detectado(s)")

        for idx, ev in enumerate(eventos):
            with st.expander(f"Evento {ev['fecha']} {ev['hora_ini']}", expanded=(idx==0)):

                with sqlite3.connect('eventos_asep.db') as conn:
                    ultimo = conn.execute("SELECT MAX(numero) FROM eventos").fetchone()[0] or 0
                num_evento = st.number_input("Numero de Evento", value=int(ultimo)+1,
                                             min_value=1, step=1, key=f"asep_num_{idx}")
                elaborador = st.selectbox("Elaborado por", list(_ELABORADORES_ASEP.values()),
                                          key=f"asep_elab_{idx}")

                col1, col2 = st.columns(2)
                with col1:
                    dia_hora_txt = st.text_input("Dia / Hora", value=f"{ev['fecha']} {ev['hora_ini']}",
                                                 key=f"asep_dh_{idx}")
                with col2:
                    if ev['tiempo_min']:
                        h_t, m_t = divmod(ev['tiempo_min'], 60)
                        t_def = f"{ev['tiempo_min']} MINUTOS" if h_t == 0 else f"{h_t} HORA(S) {m_t} MINUTOS"
                    else:
                        t_def = ""
                    tiempo_txt = st.text_input("Tiempo de interrupcion", value=t_def,
                                               key=f"asep_tiempo_{idx}")

                st.markdown("**Sectores afectados:**")
                sectores_edit = {}
                for empresa, datos in ev['empresas'].items():
                    mw_total = float(round(datos['mw'], 2))
                    areas_def = ', '.join(datos['areas'])
                    c1, c2 = st.columns([1, 3])
                    with c1:
                        emp_ed = st.text_input("Agente", value=empresa, key=f"asep_emp_{idx}_{empresa}")
                        mw_ed  = st.number_input("Carga (MW)", value=mw_total, step=0.01,
                                                  key=f"asep_mw_{idx}_{empresa}")
                    with c2:
                        areas_ed = st.text_area("Areas afectadas", value=areas_def, height=90,
                                                key=f"asep_areas_{idx}_{empresa}")
                    sectores_edit[emp_ed] = {'mw': mw_ed, 'areas': areas_ed}

                st.markdown("**Causa de la interrupcion:**")
                if st.button("Sugerir causa", key=f"asep_ia_{idx}"):
                    with st.spinner("Analizando..."):
                        try:
                            sugs = _llamar_claude_causa(ev['contexto'])
                            st.session_state[f'asep_sugs_{idx}'] = sugs
                        except Exception as ex:
                            st.warning(f"Error IA: {ex}")

                sugs = st.session_state.get(f'asep_sugs_{idx}', [])
                for si, s in enumerate(sugs):
                    if st.button(f"Usar opcion {si+1}", key=f"asep_sug_{idx}_{si}"):
                        # Escribir directo en la key del text_area para que se refleje
                        st.session_state[f'asep_causa_{idx}'] = s
                        st.rerun()
                    st.caption(f"**{si+1}.** {s}")

                causa_txt = st.text_area("Causa", height=110, key=f"asep_causa_{idx}")

                # ── Tipo de evento ────────────────────────────────────────────
                tipo_evento = st.radio(
                    "Tipo de evento",
                    ["TRANSMISION", "GENERACION"],
                    index=0,
                    horizontal=True,
                    key=f"asep_tipo_{idx}",
                )

                if st.button("Generar Reporte", type="primary", key=f"asep_gen_{idx}"):
                    _generar_reporte_asep(
                        plantilla_path=PLANTILLA_ASEP,
                        num_evento=int(num_evento),
                        dia_hora=dia_hora_txt,
                        sectores=sectores_edit,
                        causa=causa_txt,
                        tiempo=tiempo_txt,
                        elaborado_por=elaborador,
                        tipo_evento=tipo_evento,
                        idx=idx,
                        ev=ev,
                    )






# --- 8. EJECUCIÓN PRINCIPAL ---

with st.sidebar:
    st.title("☰ Menú")
    opcion = st.radio("Seleccione Reporte:", [
        "Informe RHH",
        "Reportes Diarios EOR",
        "Indisponibilidad de Unidades",
        "Centrales en Vertimiento",
        "Interrupciones ASEP",
        "Validador de Libranzas",
    ])
    st.markdown("---")

    NO_BITACORA = {"Informe RHH", "Validador de Libranzas"}
    if opcion not in NO_BITACORA:
        _uploaded = st.file_uploader("📂⬆️ ADJUNTAR BITÁCORA", type=["xls", "xlsx"], key="u_global")
        if _uploaded is not None:
            st.session_state['_bitacora_bytes'] = _uploaded.getvalue()
            st.session_state['_bitacora_name']  = _uploaded.name
        if st.session_state.get('_bitacora_bytes'):
            class _BitWrapper:
                def __init__(self, data, name, size):
                    self._buf = BytesIO(data)
                    self.name = name
                    self.size = size
                def read(self, *a): return self._buf.read(*a)
                def seek(self, *a): return self._buf.seek(*a)
                def tell(self): return self._buf.tell()
                def getvalue(self): return self._buf.getvalue()
            bitacora_global = _BitWrapper(
                st.session_state['_bitacora_bytes'],
                st.session_state.get('_bitacora_name', 'bitacora.xls'),
                len(st.session_state['_bitacora_bytes'])
            )
            if _uploaded is None:
                st.caption(f"✅ {st.session_state.get('_bitacora_name','Bitácora cargada')}")
        else:
            bitacora_global = None
        st.markdown("---")
    else:
        bitacora_global = None

if opcion == "Informe RHH":
    vista_informe_tiempo()
    
elif opcion == "Validador de Libranzas":
    vista_libranzas()
elif opcion == "Interrupciones ASEP":
    vista_eventos_asep(bitacora_global)
elif bitacora_global is not None:
    if opcion == "Reportes Diarios EOR":
        tab_eor, tab_rtr = st.tabs(["📊 Reporte EOR", "⚡ Indisponibilidades RTR"])
        with tab_eor:
            vista_pan(bitacora_global)
        with tab_rtr:
            vista_rtr(bitacora_global)
    elif opcion == "Indisponibilidad de Unidades":
        vista_unidades(bitacora_global)
    elif opcion == "Centrales en Vertimiento":
        vista_vertimiento(bitacora_global)
else:
    st.info("👈 Sube la bitácora en el menú lateral para habilitar los reportes.")