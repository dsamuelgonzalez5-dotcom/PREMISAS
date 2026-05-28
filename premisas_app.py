import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, timedelta, date
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import re
import copy

st.markdown("""
<style>
.week-badge {background:#1F3864;color:white;padding:6px 18px;border-radius:6px;
             font-size:1.2rem;font-weight:bold;display:inline-block;}
.section-title {color:#1F3864;font-weight:bold;font-size:1rem;margin-bottom:4px;}
.tag-r  {background:#d4edda;color:#155724;padding:1px 6px;border-radius:3px;font-size:.8rem;font-weight:bold;}
.tag-i  {background:#fff3cd;color:#856404;padding:1px 6px;border-radius:3px;font-size:.8rem;font-weight:bold;}
.tag-ri {background:#cce5ff;color:#004085;padding:1px 6px;border-radius:3px;font-size:.8rem;font-weight:bold;}

/* Data editor: better contrast and auto row height */
div[data-testid="stDataFrame"] .ag-cell {
    color: #ffffff !important;
    font-weight: 500 !important;
    white-space: normal !important;
    line-height: 1.4 !important;
    padding-top: 6px !important;
    padding-bottom: 6px !important;
}
div[data-testid="stDataFrame"] .ag-cell-value {
    white-space: normal !important;
    overflow: visible !important;
    word-break: break-word !important;
}
div[data-testid="stDataFrame"] .ag-row {
    height: auto !important;
    min-height: 42px !important;
}
div[data-testid="stDataFrame"] .ag-header-cell-label {
    font-weight: 700 !important;
    color: #ffffff !important;
}
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════
# CONSTANTS
# ══════════════════════════════════════════════════════════════════════
DIAS_SEMANA = ["Sábado","Domingo","Lunes","Martes","Miércoles","Jueves","Viernes"]

# Equipment types that indicate INDISPONIBILIDAD (generation)
I_EQUIP_TYPES = [
    "GENERADOR", "UNIDAD DE GENERACION", "GRUPO GENERADOR",
    "INVERSOR ESTATICO", "SEDIMENTADOR", "PRESA", "CANAL DE TRASVASE",
]
# If PLANTA prefix + TRANSFORMADOR → also I
I_DESC_KW = [
    "mantenimiento", "falla", "reparacion", "reparación", "reemplazo",
    "cambio de rodete", "cambio", "inspeccion", "inspección",
    "revision", "revisión", "averia", "avería", "paro", "parada",
    "correctivo", "overhaul", "desmontaje", "trabajos"
]

# Equipment types that indicate RELEVANTE (transmission)
R_EQUIP_TYPES = [
    "LINEA DE CONEXION", "AUTOTRANSFORMADOR", "BANCO DE CAPACITORES",
    "BARRA", "PORTICO", "PÓRTICO", "TRANSFORMADOR DE POTENCIA"
]
R_VOLT_RE  = re.compile(r'\b(230|115|34\.5|34)-|\b(230|115|34\.5)\s*k[Vv]', re.I)
R_LINE_RE  = re.compile(r'\bLIN\s+(230|115|34)[-\s]', re.I)
R_DESC_KW  = [
    "reemplazo", "cambio de aisladores", "tendido", "opgw", "conductor",
    "prueba", "pruebas electricas", "pruebas eléctricas",
    "mantenimiento", "portico", "pórtico", "torre", "aislador",
    "herraje", "poda", "reemplazar", "interruptor"
]

# Plantilla column maps
PLANTILLA_LIB_COLS = [
    "Número","Tipo","Es repetitiva","Fecha Solicitud","Fecha Inicio",
    "Fecha Final","Duración Programada","Fecha Inicio Real","Fecha Final Real",
    "Duración Real","Descripción","Observaciones","Equipos",
    "Libranzas Vinculadas","Fecha Aprobación","Último Estado","R/I"
]
PLANTILLA_REL_COLS = [
    "Tipo","Fecha inicio","Hora inicio","Fecha final","Hora final",
    "Tipo de Equipos","Equipo","Subestación","Libranza",
    "Descripción del trabajo","Estado","Observaciones"
]
INDISP_COLS = [
    "Fecha inicio","Hora inicio","Fecha final","Hora final",
    "Unidad","Potencia (MW)","Libranza","Descripción"
]
PROY_COLS = [
    "Planta","Tecnología","MW","Fecha Planeamiento",
    "Sem Disp","Sem Prueba","Última Solicitud","Fecha finaliza libranza"
]

# ══════════════════════════════════════════════════════════════════════
# UTILITIES
# ══════════════════════════════════════════════════════════════════════
def parse_dt(val):
    if val is None: return None
    if isinstance(val, datetime): return val
    if isinstance(val, date): return datetime(val.year, val.month, val.day)
    s = str(val).strip()
    for fmt in ["%d/%m/%Y %H:%M","%d/%m/%Y","%Y-%m-%d %H:%M","%Y-%m-%d"]:
        try: return datetime.strptime(s, fmt)
        except: pass
    return None

def fmt_date(dt):
    if dt is None: return ""
    return dt.strftime("%d/%m/%Y")

def fmt_time(dt):
    if dt is None: return ""
    return dt.strftime("%H:%M")

def week_range(dates):
    valid = [d for d in dates if d]
    if not valid: return "—"
    s,e = valid[0], valid[-1]
    meses = ["","ene","feb","mar","abr","may","jun","jul","ago","sep","oct","nov","dic"]
    return f"{s.day} {meses[s.month]} – {e.day} {meses[e.month]} {e.year}"

# ══════════════════════════════════════════════════════════════════════
# CLASSIFICATION
# ══════════════════════════════════════════════════════════════════════
SCADA_KW = [
    "scada", "rtu", "comunicaciones", "fibra", "telecomunicaciones",
    "fibra óptica", "fibra optica", "red de comunicacion", "red de comunicación",
    "enlace", "radiocomunicacion", "radiocomunicación", "microondas",
]

DISTRIBUTION_AGENTS = ["ENSA-", "EDEMET-", "NATURGY-", "EDECHI-"]
DISTRIBUTION_KW = ["circuito"]

SPECIAL_PLANTS = {
    'FORTUNA','BAYANO','COSTA NORTE','CRISTOBAL',
    'ESTI','ESTÍ','GATÚN','GATUN','GENERADORA GATUN','GATUN II'
}

def build_lineas_set(ws_lineas):
    """Build set of equipment codes from Lineas sheet (col B)."""
    codes = set()
    for row in ws_lineas.iter_rows(min_row=1, values_only=True):
        if row[1] and str(row[1]).strip() not in ('-',''):
            for code in str(row[1]).split('\n'):
                code = code.strip()
                if code: codes.add(code.upper())
    return codes

def parse_unit_mw_lookup(ws_indisp):
    """Parse unit code → MW reference table from Indisponibilidades sheet.
    Returns: (unit_mw dict, plant_prefix dict)."""
    unit_mw = {}      # {'BAYG3': 67.55, ...}
    plant_prefix = {} # {'bayano': 'BAY', 'fortuna': 'FOR', ...}
    current_prefix = None
    current_units  = []

    for row in ws_indisp.iter_rows(min_row=100, max_row=ws_indisp.max_row, max_col=8, values_only=True):
        if row[5] is None: continue
        code = str(row[5]).strip()
        mw   = float(row[6] or 0)

        # Unit code pattern: 2–5 uppercase letters + G + digits
        if re.match(r'^[A-Z]{2,5}G\d+$', code):
            unit_mw[code] = mw
            current_units.append(code)
            m = re.match(r'^([A-Z]+)G\d+$', code)
            if m: current_prefix = m.group(1)
        else:
            # Total / plant name row
            plant_name = re.sub(r'^Total\s+', '', code).strip().lower()
            if current_prefix and plant_name:
                plant_prefix[plant_name] = current_prefix
            unit_mw[code] = mw
            current_units  = []
            current_prefix = None

    return unit_mw, plant_prefix

def get_potencia_for_libranza(row, unit_mw, plant_prefix):
    """Sum MW for all generation units in the libranza's Equipos field."""
    equipos = str(row.get("Equipos","") or "")
    total_mw = 0.0
    for line in equipos.split('\n'):
        line = line.strip()
        m = re.match(r'^PLANTA\s+(.+?)\s*->\s*(.+)', line, re.I)
        if not m: continue
        plant_name = m.group(1).strip().lower()
        rest = m.group(2)
        prefix = next(
            (v for k, v in plant_prefix.items()
             if k in plant_name or plant_name in k),
            None
        )
        if not prefix: continue
        for uid in re.findall(r'G\d+', rest, re.I):
            code = prefix + uid.upper()
            total_mw += unit_mw.get(code, 0.0)
    return total_mw

def extract_equip_codes(equipos_str):
    """Extract all equipment codes from Equipos field for Lineas set lookup."""
    codes = set()
    for line in str(equipos_str or "").split('\n'):
        line = line.strip()
        if '->' not in line: continue
        # Left side: location (may have line codes like "LIN 230-5A")
        loc = line.split('->')[0].strip()
        for tok in re.findall(r'\d+[-\w\.]+', loc):
            codes.add(tok.upper())
        # Right side: equipment codes after ':'
        right = line.split('->')[1]
        for part in right.split(':'):
            for tok in re.findall(r'[\w][-\w\.]*', part):
                if re.search(r'[A-Z]', tok): codes.add(tok.upper())
    return codes

def get_unit_name_and_mw(row, unit_mw, plant_prefix):
    """Return (unit_label, total_mw) for an I libranza."""
    equipos = str(row.get("Equipos","") or "")
    labels = []
    total_mw = 0.0
    for line in equipos.split('\n'):
        line = line.strip()
        m = re.match(r'^PLANTA\s+(.+?)\s*->\s*(.+)', line, re.I)
        if not m: continue
        plant_name = m.group(1).strip()
        rest = m.group(2)
        plant_lower = plant_name.lower()
        prefix = next(
            (v for k, v in plant_prefix.items()
             if k in plant_lower or plant_lower in k),
            None
        )
        unit_ids = re.findall(r'G\d+', rest, re.I)
        for uid in unit_ids:
            code = (prefix or plant_name[:3].upper()) + uid.upper()
            mw = unit_mw.get(code, 0.0)
            total_mw += mw
            labels.append(code)
    return (' '.join(labels) if labels else get_unit_from_equipos(equipos)), total_mw

def classify_libranza(row, lineas_codes=None, unit_mw=None, plant_prefix=None):
    """Return 'R', 'I', 'R-I' or '' for a libranza row."""
    numero = str(row.get("Número","") or "").upper()
    eq     = str(row.get("Equipos","") or "")
    eq_up  = eq.upper()
    desc   = str(row.get("Descripción","") or "").lower()
    is_i   = False
    is_r   = False

    # ── Exclude "libranza informativa" → never R ───────────────
    if "libranza informativa" in desc or "lib. informativa" in desc:
        is_r = False
    is_distrib = any(numero.startswith(a) for a in DISTRIBUTION_AGENTS)
    if is_distrib and any(k in desc or k in eq_up.lower() for k in DISTRIBUTION_KW):
        return ""

    # ── I: non-ETESA, non-SCADA ────────────────────────────────
    if not numero.startswith("ETESA-"):
        if not any(k in desc for k in SCADA_KW):
            is_i = True

    # ── R: ETESA equipment checked against Lineas sheet ────────
    if numero.startswith("ETESA-"):
        if lineas_codes:
            codes = extract_equip_codes(eq)
            # Exclude auxiliary generators from R
            codes -= {"GENERADOR AUXILIAR", "GEN AUX", "GENERADOR AUX"}
            if codes & lineas_codes:
                is_r = True
        if not is_r:
            if R_LINE_RE.search(eq_up): is_r = True
            if "BANCO DE CAPACITORES" in eq_up: is_r = True
            if "STATCOM" in eq_up or "SPEAR" in eq_up: is_r = True
            for t in R_EQUIP_TYPES:
                if t in eq_up and R_VOLT_RE.search(eq_up):
                    is_r = True; break

    # ── R: Special large plants → always R+I ───────────────────
    if is_i and not is_r:
        for line in eq.split('\n'):
            m = re.match(r'^PLANTA\s+(.+?)\s*->', line.strip(), re.I)
            if m:
                pname = m.group(1).strip().upper()
                if any(sp in pname or pname in sp for sp in SPECIAL_PLANTS):
                    is_r = True; break

    # ── R: Other agents >60 MW ─────────────────────────────────
    if is_i and not is_r and unit_mw and plant_prefix:
        if get_potencia_for_libranza(row, unit_mw, plant_prefix) > 60:
            is_r = True

    if is_r and is_i: return "R-I"
    if is_r: return "R"
    if is_i: return "I"
    return ""

# ══════════════════════════════════════════════════════════════════════
# EQUIPOS PARSER
# ══════════════════════════════════════════════════════════════════════
TIPO_MAP = {
    "LINEA DE CONEXION":"Línea","AUTOTRANSFORMADOR":"Transformador",
    "TRANSFORMADOR DE POTENCIA":"Transformador",
    "TRANSFORMADOR PUESTO A TIERRA":"Transformador","TRANSFORMADOR":"Transformador",
    "BANCO DE CAPACITORES":"Banco de Capacitores",
    "BANCO DE BATERÍAS":"Banco de Baterías","BANCO DE BATERIAS":"Banco de Baterías",
    "BARRA":"Barra","PORTICO":"Pórtico","PÓRTICO":"Pórtico",
    "INTERRUPTOR":"Interruptor","GENERADOR AUXILIAR":"Generador Auxiliar",
    "GENERADOR":"Generador","UNIDAD DE GENERACION":"Unidad de Generación",
    "GRUPO GENERADOR":"Grupo Generador","INVERSOR ESTATICO":"Inversor",
    "CUCHILLA MANUAL DE TIERRA":"Cuchilla","CUCHILLA MOTORIZADA":"Cuchilla",
    "CUCHILLA DE ATERRIZAJE":"Cuchilla","CUCHILLA MANUAL":"Cuchilla",
    "PRESA":"Presa","SEDIMENTADOR":"Sedimentador","CIRCUITO":"Circuito",
    "RELEVADOR":"Relevador","MEDIDOR":"Medidor","TABLERO DE CONTROL":"Tablero",
    "COMPUTADOR":"Computador","RTU":"RTU","SECCIONADOR":"Seccionador",
}

_ALL_TYPES_SORTED = sorted(TIPO_MAP.keys(), key=len, reverse=True)
_TYPE_RE = re.compile(
    r'(' + '|'.join(re.escape(t) for t in _ALL_TYPES_SORTED) + r'):\s*',
    re.IGNORECASE
)
_LOC_PREFIX_RE = re.compile(r'\s{2,}(?=(?:SE|LIN|PLANTA)\s+\S.*?->)')

def parse_equipo_entries(equipos_str):
    """Parse Equipos string into list of {location, raw_tipo, tipo, ids}."""
    results = []
    if not equipos_str: return results
    # Normalize: multiple entries on one line separated by 3+ spaces
    normalized = _LOC_PREFIX_RE.sub('\n', str(equipos_str))
    for line in normalized.split("\n"):
        line = line.strip()
        if not line or "->" not in line: continue
        m = re.match(r'^(.+?)\s*->\s*(.+)$', line)
        if not m: continue
        location = m.group(1).strip()
        rest = m.group(2).strip()
        matches = list(_TYPE_RE.finditer(rest))
        for i, tm in enumerate(matches):
            raw_tipo = tm.group(1).upper()
            start = tm.end()
            end = matches[i+1].start() if i+1 < len(matches) else len(rest)
            ids = rest[start:end].strip()
            tipo_friendly = TIPO_MAP.get(raw_tipo, raw_tipo.title())
            results.append({"location": location, "raw_tipo": raw_tipo,
                            "tipo": tipo_friendly, "ids": ids})
    return results

def get_relevante_equipo_info(row, lineas_lookup):
    """Extract Tipo de Equipos, ALL Equipos, Subestación for relevantes table."""
    entries = parse_equipo_entries(row.get("Equipos",""))
    R_PRIORITY = ["LINEA DE CONEXION","AUTOTRANSFORMADOR","TRANSFORMADOR DE POTENCIA",
                  "BANCO DE CAPACITORES","BARRA","PORTICO","PÓRTICO","INTERRUPTOR"]

    # Collect ALL R-type entries
    r_entries = []
    for priority in R_PRIORITY:
        for e in entries:
            if priority in e["raw_tipo"] and e not in r_entries:
                r_entries.append(e)
    if not r_entries and entries:
        r_entries = entries[:1]
    if not r_entries:
        return {"tipo":"—","equipo":"—","sust":"—"}

    # Tipo: from first R entry
    tipo = r_entries[0]["tipo"]

    # All equipo IDs combined
    all_ids = "  ".join(
        re.sub(r'\(RTR\)', '', e["ids"]).strip()
        for e in r_entries
    ).strip()

    # Subestación: from first line entry lookup, else location
    sust = ""
    for e in r_entries:
        if "LINEA DE CONEXION" in e["raw_tipo"]:
            for token in re.sub(r'\(RTR\)', '', e["ids"]).split():
                if token.strip() in lineas_lookup:
                    sust = lineas_lookup[token.strip()]
                    break
        if sust: break
    if not sust:
        loc = r_entries[0]["location"]
        loc = re.sub(r'^(SE|LIN|PLANTA)\s+', '', loc, flags=re.I).strip()
        sust = loc.title()

    return {"tipo": tipo, "equipo": all_ids, "sust": sust}

def get_unit_from_equipos(equipos_str):
    """Extract generation unit description for indisponibilidades."""
    entries = parse_equipo_entries(equipos_str)
    GEN_TYPES = ["GENERADOR","UNIDAD DE GENERACION","GRUPO GENERADOR","INVERSOR ESTATICO"]
    for e in entries:
        if any(t in e["raw_tipo"] for t in GEN_TYPES):
            loc = re.sub(r'^PLANTA\s+', '', e["location"], flags=re.I).strip()
            return f"{e['ids']}  ({loc})"
    if entries:
        e = entries[0]
        return f"{e['ids']}  ({e['location']})"
    return equipos_str[:60] if equipos_str else ""

# ══════════════════════════════════════════════════════════════════════
# LOADERS
# ══════════════════════════════════════════════════════════════════════
@st.cache_data(show_spinner=False)
def load_plantilla(file_bytes):
    """Load template workbook and extract calendar, line lookup, proyectos."""
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    out = {"wb_bytes": file_bytes}

    # ── Week calendar from Datos ───────────────────────────────────
    ws = wb["Datos"]
    weeks = {}
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=8, values_only=True):
        label = str(row[0] or "").strip()
        if label.startswith("Semana"):
            try:
                num = int(label.split()[1])
                if num in weeks:
                    continue  # keep first occurrence (current year), skip duplicates
                dates = [parse_dt(row[i]) for i in range(1,8)]
                weeks[num] = dates  # [Sab,Dom,Lun,Mar,Mie,Jue,Vie]
            except: pass
    out["weeks"] = weeks

    # ── Current week = plantilla week + 1 ────────────────────────────
    ws_i = wb["Indisponibilidades"]
    title = str(ws_i.cell(1,1).value or "")
    m = re.search(r"Semana\s+(\d+)", title)
    plantilla_week = int(m.group(1)) if m else 1
    out["current_week"] = plantilla_week + 1  # generate NEXT week

    # ── Lineas lookup: code → description ─────────────────────────
    lineas = {}
    ws_l = wb["Lineas"]
    for row in ws_l.iter_rows(min_row=1, values_only=True):
        if row[1] and row[2]:
            for code in str(row[1]).split("\n"):
                code = code.strip()
                if code:
                    lineas[code] = str(row[2]).strip()
    out["lineas_lookup"] = lineas

    # ── Proyectos de Generación ────────────────────────────────────
    ws_p = wb["Proyectos de Generacion"]
    proy_rows = []
    hdr_found = False
    for row in ws_p.iter_rows(min_row=13, values_only=True):
        if not hdr_found:
            if row[0] == "Planta": hdr_found = True
            continue
        if not any(c is not None for c in row[:8]): continue
        r = list(row) + [None]*9
        fecha_plan = r[3]
        fecha_fin  = r[7]
        proy_rows.append({
            "Planta":     str(r[0] or ""),
            "Tecnología": str(r[1] or ""),
            "MW":         r[2],
            "Fecha Planeamiento": fmt_date(parse_dt(fecha_plan)) if fecha_plan else str(r[3] or ""),
            "Sem Disp":   str(r[4] or ""),
            "Sem Prueba": str(r[5] or ""),
            "Última Solicitud": str(r[6] or ""),
            "Fecha finaliza libranza": fmt_date(parse_dt(fecha_fin)) if isinstance(fecha_fin, datetime) else str(r[7] or ""),
            "_fecha_fin_raw": fecha_fin,
        })
    out["proyectos"] = pd.DataFrame(proy_rows) if proy_rows else pd.DataFrame(columns=PROY_COLS+["_fecha_fin_raw"])

    # ── Unit MW reference table ────────────────────────────────────
    unit_mw, plant_prefix = parse_unit_mw_lookup(wb["Indisponibilidades"])
    out["unit_mw"]       = unit_mw
    out["plant_prefix"]  = plant_prefix

    # ── Lineas set for R classification ──────────────────────────
    out["lineas_codes"] = build_lineas_set(wb["Lineas"])

    # ── Existing Indisponibilidades (flat unique list) ─────────────
    out["indisp_existing"] = parse_existing_indisp_flat(wb["Indisponibilidades"])

    # ── Previous week's Libranzas Relevantes (carry forward) ──────
    out["relevantes_anteriores"] = parse_relevantes_anteriores(wb["Libranzas Relevantes"])

    return out

def parse_relevantes_anteriores(ws):
    """Parse previous week's Libranzas Relevantes sheet → list of dicts."""
    rows = []
    hdr_found = False
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not hdr_found:
            if row[0] == "Tipo": hdr_found = True
            continue
        if not any(c is not None for c in row[:9]): continue
        r = list(row) + [None]*12
        rows.append({
            "Tipo":               str(r[0] or ""),
            "Fecha inicio":       str(r[1] or ""),
            "Hora inicio":        str(r[2] or ""),
            "Fecha final":        str(r[3] or ""),
            "Hora final":         str(r[4] or ""),
            "Tipo de Equipos":    str(r[5] or ""),
            "Equipo":             str(r[6] or ""),
            "Subestación":        str(r[7] or ""),
            "Libranza":           str(r[8] or ""),
            "Descripción del trabajo": str(r[9] or ""),
            "Estado":             str(r[10] or ""),
            "Observaciones":      str(r[11] or ""),
            "_sort_dt":           parse_dt(str(r[1] or "")) or datetime.max,
        })
    return rows

def parse_existing_indisp_flat(ws):
    """Parse plantilla's Indisponibilidades → flat unique list of entries (one per libranza)."""
    seen = set()
    entries = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not any(c is not None for c in row[:9]): continue
        if row[1] and "INDISPONIBILIDAD TOTAL" in str(row[1]): continue
        if isinstance(row[0], datetime): continue  # day header
        if row[7] is None: continue
        lib = str(row[7]).strip()
        if not lib or lib in seen: continue
        seen.add(lib)
        entries.append({
            "Fecha inicio":  str(row[1] or ""),
            "Hora inicio":   str(row[2] or ""),
            "Fecha final":   str(row[3] or ""),
            "Hora final":    str(row[4] or ""),
            "Unidad":        str(row[5] or ""),
            "Potencia (MW)": float(row[6] or 0),
            "Libranza":      lib,
            "Descripción":   str(row[8] or ""),
        })
    return entries

def parse_existing_indisp(ws):  # kept for backward compat
    return parse_existing_indisp_flat(ws)

@st.cache_data(show_spinner=False)
def load_source_libranzas(file_bytes):
    """Load libranzas source file using header names, not positional indices."""
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = []
    headers = {}
    hdr_found = False

    for row in ws.iter_rows(min_row=1, values_only=True):
        if not hdr_found:
            if row[0] == "Número":
                hdr_found = True
                headers = {str(c).strip(): i for i, c in enumerate(row) if c is not None}
            continue
        if not any(c is not None for c in row): continue

        def g(col, default=""):
            idx = headers.get(col)
            if idx is None: return default
            val = row[idx] if idx < len(row) else None
            return str(val) if val is not None else default

        rows.append({
            "Número":            row[0],
            "Tipo":              g("Tipo"),
            "Es repetitiva":     g("Es repetitiva"),
            "Agente":            g("Agente"),
            "Fecha Solicitud":   g("Fecha Solicitud"),
            "Solicitante":       g("Solicitante"),
            "Fecha Inicio":      g("Fecha Inicio"),
            "Fecha Final":       g("Fecha Final"),
            "Duración Programada": g("Duración Programada"),
            "Fecha Inicio Real": g("Fecha Inicio Real"),
            "Fecha Final Real":  g("Fecha Final Real"),
            "Duración Real":     g("Duración Real"),
            "Descripción":       g("Descripción"),
            "Observaciones":     g("Observaciones"),
            "Equipos":           g("Equipos"),
            "Libranzas Vinculadas": g("Libranzas Vinculadas"),
            "Responsable de Campo": g("Responsable de Campo"),
            "Cargo Responsable Campo": g("Cargo Responsable Campo"),
            "Fecha Aprobación":  g("Fecha Aprobación"),
            "Último Estado":     g("Último Estado"),
        })
    return pd.DataFrame(rows) if rows else pd.DataFrame()

@st.cache_data(show_spinner=False)
def load_indisp_file(file_bytes):
    """Parse INDISPONIBILIDAD DE UNIDADES file.
    Cols: B=Unidad, C=Fecha salida, D=Causa, E=Libranza, F=Fecha entrada, G=Potencia(MW)
    Returns flat DataFrame with INDISP_COLS + ['status']."""
    MESES_ES = {'ene':'01','feb':'02','mar':'03','abr':'04','may':'05','jun':'06',
                'jul':'07','ago':'08','sep':'09','oct':'10','nov':'11','dic':'12'}

    def parse_es_date(val):
        if val is None: return None
        s = str(val).strip()
        # Try standard formats first
        for fmt in ["%Y-%m-%d %H:%M:%S","%Y-%m-%d %H:%M","%Y-%m-%d"]:
            try: return datetime.strptime(s[:len(fmt)+2].strip(), fmt)
            except: pass
        # Spanish: "26-mar-2026 09:30"
        m = re.match(r'(\d{1,2})-([a-záéíóú]+)-(\d{4})(?:\s+(\d{1,2}:\d{2}))?', s, re.I)
        if m:
            mes = MESES_ES.get(m.group(2).lower()[:3], '01')
            hora = m.group(4) or '00:00'
            try: return datetime.strptime(f"{m.group(1):0>2}/{mes}/{m.group(3)} {hora}", "%d/%m/%Y %H:%M")
            except: pass
        return None

    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=True)
        ws = wb.active
        rows = []
        hdr_found = False
        for row in ws.iter_rows(min_row=1, values_only=True):
            # Find header row (col B = "UNIDAD")
            if not hdr_found:
                if str(row[1] or '').strip().upper() == 'UNIDAD':
                    hdr_found = True
                continue
            b = row[1]; c = row[2]; d = row[3]; e = row[4]; f = row[5]; g = row[6]
            if b is None or str(b).strip() in ('','▼','▲'): continue
            if 'TOTAL' in str(b or '').upper() or 'ZONA' in str(b or '').upper(): continue
            fi = parse_es_date(c)
            ff = parse_es_date(f)
            if fi is None: continue
            rows.append({
                "Fecha inicio":  fmt_date(fi),
                "Hora inicio":   fmt_time(fi),
                "Fecha final":   fmt_date(ff) if ff else "SIN FECHA",
                "Hora final":    fmt_time(ff) if ff else "",
                "Unidad":        str(b).strip(),
                "Potencia (MW)": float(g or 0),
                "Libranza":      str(e or "").strip(),
                "Descripción":   str(d or "").strip(),
                "status":        "vieja",
            })
        return pd.DataFrame(rows) if rows else pd.DataFrame(columns=INDISP_COLS+["status"])
    except Exception as e:
        return pd.DataFrame(columns=INDISP_COLS+["status"])

# ══════════════════════════════════════════════════════════════════════
# WEEK LOOKUP
# ══════════════════════════════════════════════════════════════════════
def find_week_for_date(target, weeks):
    """Find week number using Sábado/Viernes only (avoids bad Domingo year)."""
    if target is None: return None
    dt = parse_dt(target)
    if dt is None: return None
    target_d = dt.date() if isinstance(dt, datetime) else dt
    for wn, dates in weeks.items():
        sab = dates[0]   # Sábado — always correct year
        vie = dates[6]   # Viernes — always correct year
        if sab is None or vie is None: continue
        sab_d = sab.date() if isinstance(sab, datetime) else sab
        vie_d = vie.date() if isinstance(vie, datetime) else vie
        if sab_d <= target_d <= vie_d:
            return wn
    return None

# ══════════════════════════════════════════════════════════════════════
# PROCESSORS
# ══════════════════════════════════════════════════════════════════════
def process_nuevas(df_raw, lineas_codes=None, unit_mw=None, plant_prefix=None):
    """Filter and classify libranzas nuevas."""
    if df_raw.empty: return pd.DataFrame(columns=PLANTILLA_LIB_COLS)
    # Exclude ALL Cancelado + ALL Emergencia
    mask = (df_raw["Último Estado"].str.strip() != "Cancelado") & \
           (df_raw["Tipo"].str.strip() != "Emergencia")
    df = df_raw[mask].copy()

    out_rows = []
    for _, row in df.iterrows():
        ri = classify_libranza(row, lineas_codes, unit_mw, plant_prefix)
        out_rows.append({
            "Número":           row["Número"],
            "Tipo":             row["Tipo"],
            "Es repetitiva":    row["Es repetitiva"],
            "Fecha Solicitud":  row["Fecha Solicitud"],
            "Fecha Inicio":     row["Fecha Inicio"],
            "Fecha Final":      row["Fecha Final"],
            "Duración Programada": row["Duración Programada"],
            "Fecha Inicio Real":   row["Fecha Inicio Real"],
            "Fecha Final Real":    row["Fecha Final Real"],
            "Duración Real":       row["Duración Real"],
            "Descripción":         row["Descripción"],
            "Observaciones":       row["Observaciones"],
            "Equipos":             row["Equipos"],
            "Libranzas Vinculadas": row.get("Libranzas Vinculadas",""),
            "Fecha Aprobación":    row["Fecha Aprobación"],
            "Último Estado":       row["Último Estado"],
            "R/I":                 ri,
        })
    return pd.DataFrame(out_rows)

def process_viejas(df_raw, lineas_codes=None, unit_mw=None, plant_prefix=None):
    """Filter libranzas viejas: keep only Aprobado and Recibido, classify R/I."""
    if df_raw.empty: return pd.DataFrame(columns=PLANTILLA_LIB_COLS)
    keep = ["Aprobado","Recibido"]
    df = df_raw[df_raw["Último Estado"].str.strip().isin(keep)].copy()
    out_rows = []
    for _, row in df.iterrows():
        ri = classify_libranza(row, lineas_codes, unit_mw, plant_prefix)
        out_rows.append({
            "Número":           row["Número"],
            "Tipo":             row["Tipo"],
            "Es repetitiva":    row["Es repetitiva"],
            "Fecha Solicitud":  row["Fecha Solicitud"],
            "Fecha Inicio":     row["Fecha Inicio"],
            "Fecha Final":      row["Fecha Final"],
            "Duración Programada": row["Duración Programada"],
            "Fecha Inicio Real":   row["Fecha Inicio Real"],
            "Fecha Final Real":    row["Fecha Final Real"],
            "Duración Real":       row["Duración Real"],
            "Descripción":         row["Descripción"],
            "Observaciones":       row["Observaciones"],
            "Equipos":             row["Equipos"],
            "Libranzas Vinculadas": row.get("Libranzas Vinculadas",""),
            "Fecha Aprobación":    row["Fecha Aprobación"],
            "Último Estado":       row["Último Estado"],
            "R/I":                 ri,
        })
    return pd.DataFrame(out_rows)

def build_relevantes(df_nuevas, df_viejas, lineas_lookup, relevantes_anteriores=None, week_start=None):
    """Build relevantes: viejas first (blue), then nuevas (no color), sorted by Fecha inicio.
    Excludes: libranza informativa, secondary linked libranzas sharing de-energization."""
    rows      = []
    full_data = {}   # libranza_num → full source row
    seen      = set()
    viejas_nums = set(str(r) for r in df_viejas["Número"].dropna()) \
                  if df_viejas is not None and not df_viejas.empty else set()
    DEENERG_KW = ["desenergiz","de-energiz","desconex","apertura","seccionamiento"]

    def add_relevantes_from(df, classify=False, source_status="nueva"):
        if df is None or df.empty: return
        for _, row in df.iterrows():
            desc = str(row.get("Descripción","") or "").lower()
            if "libranza informativa" in desc or "lib. informativa" in desc:
                continue
            ri = row.get("R/I","") if not classify else classify_libranza(row)
            if "R" not in str(ri): continue
            eq_up = str(row.get("Equipos","") or "").upper()
            if "GENERADOR AUXILIAR" in eq_up and not any(
                t in eq_up for t in ["LINEA","AUTOTRANSFORMADOR","BANCO DE CAPACITORES"]):
                continue
            num = str(row.get("Número",""))
            if num in seen: continue
            seen.add(num)
            full_data[num] = row
            fi = parse_dt(row.get("Fecha Inicio",""))
            ff = parse_dt(row.get("Fecha Final",""))
            info = get_relevante_equipo_info(row, lineas_lookup)
            rows.append({
                "Tipo":               row.get("Es repetitiva",""),
                "Fecha inicio":       fmt_date(fi),
                "Hora inicio":        fmt_time(fi),
                "Fecha final":        fmt_date(ff),
                "Hora final":         fmt_time(ff),
                "Tipo de Equipos":    info["tipo"],
                "Equipo":             info["equipo"],
                "Subestación":        info["sust"],
                "Libranza":           num,
                "Descripción del trabajo": row.get("Descripción",""),
                "Estado":             row.get("Último Estado",""),
                "Observaciones":      row.get("Observaciones",""),
                "_sort_dt":           fi or datetime.max,
                "_status":            source_status,
            })

    add_relevantes_from(df_viejas, classify=False, source_status="vieja")
    add_relevantes_from(df_nuevas, classify=False, source_status="nueva")

    if relevantes_anteriores:
        for r in relevantes_anteriores:
            lib = r.get("Libranza","")
            if not lib or lib in seen: continue
            ff = parse_dt(r.get("Fecha final",""))
            if week_start and ff and ff.date() < week_start: continue
            seen.add(lib)
            status = "vieja" if lib in viejas_nums else "nueva"
            rows.append({k: r[k] for k in PLANTILLA_REL_COLS} | {
                "_sort_dt": r.get("_sort_dt", datetime.max),
                "_status":  status,
            })

    # Filter secondary linked libranzas (vinculada is base → remove secondary)
    rel_nums  = {r["Libranza"] for r in rows}
    to_remove = set()
    for r in rows:
        num     = r["Libranza"]
        src     = full_data.get(num)
        if src is None: continue
        vinc    = str(src.get("Libranzas Vinculadas","") or "")
        if not vinc or vinc in ("None","nan","No hay libranzas vinculadas",""): continue
        for vin in re.split(r"[,;\n\s]+", vinc):
            vin = vin.strip()
            if not vin or vin not in rel_nums: continue
            vin_src = full_data.get(vin)
            if vin_src is None: continue
            fi_curr = parse_dt(src.get("Fecha Inicio",""))
            ff_curr = parse_dt(src.get("Fecha Final",""))
            fi_vin  = parse_dt(vin_src.get("Fecha Inicio",""))
            ff_vin  = parse_dt(vin_src.get("Fecha Final",""))
            if fi_curr is None or fi_vin is None: continue
            # Vinculada (base) started before current AND current dates within base range
            if fi_vin <= fi_curr:
                dates_within = (ff_vin is None or ff_curr is None or
                                fi_curr >= fi_vin and (ff_curr <= ff_vin or abs((ff_curr - ff_vin).total_seconds()) < 3600))
                eq_c = extract_equip_codes(str(src.get("Equipos","") or ""))
                eq_v = extract_equip_codes(str(vin_src.get("Equipos","") or ""))
                if dates_within and eq_c & eq_v:
                    to_remove.add(num)
    rows = [r for r in rows if r["Libranza"] not in to_remove]

    if not rows: return pd.DataFrame(columns=PLANTILLA_REL_COLS + ["_status"])
    df_out = pd.DataFrame(rows)
    df_out["_is_nueva"] = (df_out["_status"] == "nueva").astype(int)
    df_out = df_out.sort_values(["_is_nueva","_sort_dt"]).drop(columns=["_is_nueva","_sort_dt"])
    return df_out.reset_index(drop=True)

def build_indisponibilidades(indisp_existing, df_viejas, df_nuevas,
                             unit_mw, plant_prefix, weeks, current_week,
                             indisp_file_df=None):
    """Build flat indisponibilidades DataFrame.
    Priority: external file > plantilla existing > nuevas-I
    status: 'vieja' (blue) | 'nueva' (white)
    """
    week_dates = weeks.get(current_week, [None]*7)
    sab = week_dates[0]
    fixed_dates = []
    for i, d in enumerate(week_dates):
        if d is None: fixed_dates.append(None)
        elif sab and abs((d - sab).days) > 7:
            fixed_dates.append(sab + timedelta(days=i))
        else: fixed_dates.append(d)
    valid_dts = [d for d in fixed_dates if d]
    if not valid_dts: return pd.DataFrame(columns=INDISP_COLS + ["status"])

    week_start = valid_dts[0].date()
    week_end   = valid_dts[-1].date()
    result = []
    seen   = set()

    # ── 1. External indisponibilidades file (primary source) ───────
    if indisp_file_df is not None and not indisp_file_df.empty:
        for _, r in indisp_file_df.iterrows():
            fi = parse_dt(r.get("Fecha inicio",""))
            ff_str = str(r.get("Fecha final",""))
            ff = parse_dt(ff_str) if ff_str not in ("SIN FECHA","") else None
            lib = str(r.get("Libranza","")).strip()
            if fi is None: continue
            fi_d = fi.date()
            ff_d = ff.date() if ff else date(2099, 12, 31)
            if fi_d <= week_end and ff_d >= week_start:
                entry = {k: r.get(k,"") for k in INDISP_COLS}
                entry["status"] = "vieja"
                result.append(entry)
                if lib: seen.add(lib)
    else:
        # ── 2. Fallback: plantilla existing filtered by viejas ──────
        viejas_nums = set(str(r) for r in df_viejas["Número"].dropna()) \
                      if df_viejas is not None and not df_viejas.empty else set()
        for entry in (indisp_existing or []):
            lib = str(entry.get("Libranza",""))
            if not lib or lib in seen: continue
            if lib in viejas_nums:
                result.append({**{k: entry.get(k,"") for k in INDISP_COLS}, "status": "vieja"})
            seen.add(lib)

    # ── 3. New I entries from libranzas_nuevas ─────────────────────
    GEN_TYPES = ["GENERADOR","UNIDAD DE GENERACION","GRUPO GENERADOR","INVERSOR ESTATICO"]
    if df_nuevas is not None and not df_nuevas.empty:
        df_i = df_nuevas[df_nuevas["R/I"].str.contains("I", na=False)]
        for _, row in df_i.iterrows():
            eq = str(row.get("Equipos","") or "").upper()
            if not any(t in eq for t in GEN_TYPES): continue
            num = str(row.get("Número",""))
            if num in seen: continue
            seen.add(num)
            fi = parse_dt(row.get("Fecha Inicio",""))
            ff = parse_dt(row.get("Fecha Final",""))
            if fi is None or ff is None: continue
            unit_label, mw = get_unit_name_and_mw(row, unit_mw or {}, plant_prefix or {})
            result.append({
                "Fecha inicio": fmt_date(fi), "Hora inicio": fmt_time(fi),
                "Fecha final":  fmt_date(ff), "Hora final":  fmt_time(ff),
                "Unidad": unit_label, "Potencia (MW)": mw,
                "Libranza": num, "Descripción": str(row.get("Descripción","")),
                "status": "nueva"
            })

    if not result: return pd.DataFrame(columns=INDISP_COLS + ["status"])
    df = pd.DataFrame(result)
    df["_ff"] = df["Fecha final"].apply(parse_dt)
    df_v = df[df["status"]=="vieja"].copy()
    df_n = df[df["status"]=="nueva"].sort_values("_ff")
    return pd.concat([df_v, df_n], ignore_index=True).drop(columns=["_ff"])


def update_proyectos(df_proy, current_week, weeks):
    """Update Sem Disp (from fecha_fin lookup) and Sem Prueba (replace week number)."""
    if df_proy.empty: return df_proy
    df = df_proy.copy()
    prev_week = current_week - 1
    for idx, row in df.iterrows():
        raw = row.get("_fecha_fin_raw")
        fecha_str = row.get("Fecha finaliza libranza","")
        dt = parse_dt(raw) if (raw is not None and str(raw) not in ["None","nan",""]) \
             else parse_dt(fecha_str)
        if dt is not None:
            wn = find_week_for_date(dt, weeks)
            if wn is not None:
                df.at[idx,"Sem Disp"] = f"Semana {wn}"
        sem_p = str(row.get("Sem Prueba","") or "")
        if sem_p and str(prev_week) in sem_p:
            df.at[idx,"Sem Prueba"] = re.sub(
                rf'\bSemana\s+{prev_week}\b', f'Semana {current_week}', sem_p)
    return df

def detect_proyectos_from_libranzas(df_viejas, df_nuevas, df_proy_existing):
    """Detect libranzas with 'prueba de generacion' not yet in proyectos, add them."""
    PRUEBA_KW = ["prueba de generacion", "prueba de generación", "pruebas de generacion",
                 "pruebas de generación", "prueba de puesta en servicio"]
    existing_libs = set(str(r).strip() for r in df_proy_existing["Última Solicitud"].dropna()) \
                    if not df_proy_existing.empty else set()

    new_rows = []
    for df in [df_viejas, df_nuevas]:
        if df is None or df.empty: continue
        for _, row in df.iterrows():
            num = str(row.get("Número",""))
            desc = str(row.get("Descripción","") or "").lower()
            if not any(kw in desc for kw in PRUEBA_KW): continue
            if num in existing_libs: continue  # already in proyectos
            # Try to extract plant name from Equipos
            equipos = str(row.get("Equipos","") or "")
            m = re.search(r'(?:PLANTA|SE)\s+([A-ZÁÉÍÓÚÑ][A-Za-záéíóúñ\s]+?)\s*->', equipos)
            planta = m.group(1).strip().title() if m else num
            ff = row.get("Fecha Final","")
            new_rows.append({
                "Planta":     planta,
                "Tecnología": "",
                "MW":         None,
                "Fecha Planeamiento": "",
                "Sem Disp":   "",
                "Sem Prueba": "",
                "Última Solicitud": num,
                "Fecha finaliza libranza": ff,
                "_fecha_fin_raw": parse_dt(str(ff)),
            })
            existing_libs.add(num)

    if not new_rows: return df_proy_existing
    df_new = pd.DataFrame(new_rows)
    return pd.concat([df_proy_existing, df_new], ignore_index=True)
    """Update Sem Disp (from fecha_fin lookup) and Sem Prueba (replace week number)."""
    if df_proy.empty: return df_proy
    df = df_proy.copy()
    prev_week = current_week - 1

    for idx, row in df.iterrows():
        # ── Sem Disp: look up fecha_fin in calendar ──────────────────
        raw = row.get("_fecha_fin_raw")
        fecha_str = row.get("Fecha finaliza libranza","")
        dt = parse_dt(raw) if (raw is not None and str(raw) not in ["None","nan",""]) \
             else parse_dt(fecha_str)
        if dt is not None:
            wn = find_week_for_date(dt, weeks)
            if wn is not None:
                df.at[idx,"Sem Disp"] = f"Semana {wn}"

        # ── Sem Prueba: replace previous week number with current ─────
        sem_p = str(row.get("Sem Prueba","") or "")
        if sem_p and str(prev_week) in sem_p:
            df.at[idx,"Sem Prueba"] = re.sub(
                rf'\bSemana\s+{prev_week}\b',
                f'Semana {current_week}',
                sem_p
            )
    return df

# ══════════════════════════════════════════════════════════════════════
# EXPORT
# ══════════════════════════════════════════════════════════════════════
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF", size=9)
TITLE_FONT = Font(bold=True, color="1F3864", size=11)
DATA_FONT  = Font(size=9)
THIN  = Side(border_style="thin", color="CCCCCC")
BORD  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP  = Alignment(vertical="top", wrap_text=True)
CTR   = Alignment(horizontal="center", vertical="center")

def _hdr(ws, headers, row=1):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row, c, h)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = CTR; cell.border = BORD

def _write(ws, df, start_row=3):
    for ri, (_, r) in enumerate(df.iterrows(), start_row):
        for ci, col in enumerate(df.columns, 1):
            cell = ws.cell(ri, ci, r[col])
            cell.font = DATA_FONT; cell.border = BORD; cell.alignment = WRAP

def _clear_rows(ws, min_row):
    """Clear data rows safely, skipping merged cells."""
    for row in ws.iter_rows(min_row=min_row, max_row=ws.max_row):
        for cell in row:
            try:
                cell.value = None
            except AttributeError:
                pass  # skip merged cells

def export_premisas(state):
    """Generate updated plantilla Excel - memory optimized."""
    import gc
    gc.collect()

    wb_template = load_workbook(BytesIO(state["prem_plantilla_bytes"]),
                                keep_links=False)
    gc.collect()

    # ── Remove unwanted sheets ────────────────────────────────────────
    for sheet_name in ["Lineas", "Datos", "LIBRANZAS NUEVAS", "LIBRANZAS VIEJAS"]:
        if sheet_name in wb_template.sheetnames:
            del wb_template[sheet_name]
    gc.collect()

    # ── LIBRANZAS RELEVANTES ─────────────────────────────────────────
    ws = wb_template["Libranzas Relevantes"]
    for merge in list(ws.merged_cells.ranges): ws.unmerge_cells(str(merge))
    _clear_rows(ws, 3)
    _hdr(ws, PLANTILLA_REL_COLS, row=2)
    df_r = state.get("prem_df_relevantes")
    if df_r is not None and not df_r.empty:
        BLUE_FILL_R  = PatternFill("solid", fgColor="BDD7EE")
        GREEN_FILL   = PatternFill("solid", fgColor="C6EFCE")
        PINK_FILL    = PatternFill("solid", fgColor="FFB6C1")
        WHITE_FILL_R = PatternFill("solid", fgColor="FFFFFF")
        for ri, (_, r) in enumerate(df_r.iterrows(), 3):
            status   = str(r.get("_status","")) if "_status" in r.index else ""
            is_vieja = status == "vieja"
            tipo_val = str(r.get("Tipo","")).strip().lower()
            for ci, col in enumerate(PLANTILLA_REL_COLS, 1):
                val = r.get(col,"") if col in r.index else ""
                try:
                    cell = ws.cell(ri, ci, val)
                    cell.font = DATA_FONT; cell.border = BORD
                    cell.alignment = WRAP; cell.fill = WHITE_FILL_R
                    if ci == 1:
                        cell.fill = GREEN_FILL if "continua" in tipo_val else (PINK_FILL if "repetitiva" in tipo_val else WHITE_FILL_R)
                    elif ci == 9 and is_vieja:
                        cell.fill = BLUE_FILL_R
                        cell.font = Font(size=9, bold=True)
                except: pass
    del df_r; gc.collect()

    # ── INDISPONIBILIDADES ───────────────────────────────────────────
    BLUE_FILL  = PatternFill("solid", fgColor="BDD7EE")
    TOTAL_FONT = Font(bold=True, size=9)

    ws = wb_template["Indisponibilidades"]
    cur_week = state.get("prem_current_week", 1)
    for merge in list(ws.merged_cells.ranges): ws.unmerge_cells(str(merge))
    try: ws.cell(1, 1).value = f"Indisponibilidades de Generación Intersemanales - Semana {cur_week}"
    except: pass
    try: ws.cell(2, 1).value = f"Semana {cur_week}"
    except: pass
    _clear_rows(ws, 3)

    indisp_df  = state.get("prem_indisp_data")
    weeks      = state.get("prem_weeks", {})
    week_dates = weeks.get(cur_week, [None]*7)

    # Fix Domingo dates
    sab = week_dates[0] if week_dates else None
    fixed_dates = []
    for i, d in enumerate(week_dates):
        if d is None: fixed_dates.append(None)
        elif sab and abs((d-sab).days) > 7: fixed_dates.append(sab + timedelta(days=i))
        else: fixed_dates.append(d)

    ri = 3
    weekly_total = 0.0

    for i, dia in enumerate(DIAS_SEMANA):
        dt = fixed_dates[i] if i < len(fixed_dates) else None
        if dt is None: continue
        day_date = dt.date()

        # Filter indisp_df for this day
        if indisp_df is None or (hasattr(indisp_df,'empty') and indisp_df.empty):
            day_rows = []
        else:
            day_rows = []
            for _, row_data in indisp_df.iterrows():
                fi = parse_dt(row_data.get("Fecha inicio",""))
                ff = parse_dt(row_data.get("Fecha final",""))
                if fi is None or ff is None: continue
                if fi.date() <= day_date <= ff.date():
                    day_rows.append(row_data)

        if not day_rows: continue

        # Sort: viejas first, then nuevas; within each group by Fecha inicio ascending
        def sort_key(r):
            is_nueva = 1 if str(r.get("status","")) == "nueva" else 0
            fi = parse_dt(r.get("Fecha inicio","")) or datetime.max
            return (is_nueva, fi)
        day_rows = sorted(day_rows, key=sort_key)

        WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
        NO_BORDER  = Border()
        DATA_BORD  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

        day_start_ri = ri   # first data row for this day
        day_total    = 0.0

        for row_data in day_rows:
            status   = str(row_data.get("status","")) if hasattr(row_data, 'get') else ""
            is_vieja = status == "vieja"

            for ci, col in enumerate(INDISP_COLS, 2):
                val = row_data.get(col, "") if hasattr(row_data,'get') else ""
                try:
                    cell = ws.cell(ri, ci, val)
                    cell.font   = DATA_FONT
                    cell.border = DATA_BORD
                    cell.fill   = WHITE_FILL
                    # Blue ONLY on Libranza col (ci=8) for viejas
                    if is_vieja and ci == 8:
                        cell.fill = BLUE_FILL
                        cell.font = Font(size=9, bold=True)
                except: pass

            mw = float(row_data.get("Potencia (MW)", 0) or 0) if hasattr(row_data,'get') else 0
            day_total   += mw
            weekly_total += mw
            ri += 1

        # Merge col A for all data rows of this day
        day_end_ri = ri - 1
        try:
            if day_end_ri > day_start_ri:
                ws.merge_cells(f"A{day_start_ri}:A{day_end_ri}")
            cell_a = ws.cell(day_start_ri, 1)
            cell_a.value     = f"{dia.lower()} {day_date.day}"
            cell_a.font      = Font(size=9, bold=True)
            cell_a.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
            cell_a.border    = DATA_BORD
            cell_a.fill      = WHITE_FILL
        except: pass

        # Daily total: merge B:F, value + MW in G, NO borders
        try:
            ws.merge_cells(f"B{ri}:F{ri}")
            cl = ws.cell(ri, 2)
            cl.value     = "INDISPONIBILIDAD TOTAL EN HORAS PUNTA (MW):"
            cl.font      = TOTAL_FONT
            cl.fill      = WHITE_FILL
            cl.border    = NO_BORDER
            cl.alignment = Alignment(vertical="center")
            cm = ws.cell(ri, 7)
            cm.value  = round(day_total, 4)
            cm.font   = TOTAL_FONT
            cm.fill   = WHITE_FILL
            cm.border = NO_BORDER
            for ci in [1, 3, 4, 5, 6, 8, 9]:
                try:
                    c = ws.cell(ri, ci)
                    c.value = None; c.fill = WHITE_FILL; c.border = NO_BORDER
                except: pass
        except: pass
        ri += 2

    # Weekly grand total
    try:
        WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
        ws.merge_cells(f"B{ri}:F{ri}")
        cl = ws.cell(ri, 2)
        cl.value = "INDISPONIBILIDAD TOTAL EN HORAS PUNTA (MW):"
        cl.font  = Font(bold=True, size=9, color="FF0000")
        cl.fill  = WHITE_FILL; cl.border = Border()
        cm = ws.cell(ri, 7)
        cm.value = round(weekly_total, 4)
        cm.font  = Font(bold=True, size=9, color="FF0000")
        cm.fill  = WHITE_FILL; cm.border = Border()
    except: pass

    # Clear everything below the grand total (remove old borders/fills)
    last_written = ri
    WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
    for row in ws.iter_rows(min_row=last_written+1, max_row=min(last_written+200, ws.max_row)):
        for cell in row:
            try:
                cell.value  = None
                cell.fill   = WHITE_FILL
                cell.border = Border()
            except: pass

    del indisp_df; gc.collect()

    # ── PROYECTOS DE GENERACIÓN ──────────────────────────────────────
    ws = wb_template["Proyectos de Generacion"]
    df_p = state.get("prem_df_proyectos", pd.DataFrame())
    if not df_p.empty:
        # Find data start row
        data_start = None
        for ridx, row in enumerate(ws.iter_rows(min_row=13, values_only=True), 13):
            if row[0] == "Planta":
                data_start = ridx + 1
                break
        if data_start:
            # Clear existing data
            _clear_rows(ws, data_start)
            for ri2, (_, r) in enumerate(df_p.iterrows(), data_start):
                cols_to_write = [c for c in PROY_COLS if c in df_p.columns]
                for ci, col in enumerate(cols_to_write, 1):
                    ws.cell(ri2, ci).value = r.get(col,"")
                    ws.cell(ri2, ci).font = DATA_FONT

    gc.collect()
    buf = BytesIO()
    wb_template.save(buf)
    wb_template.close()
    del wb_template; gc.collect()
    buf.seek(0)
    return buf.read()

# ══════════════════════════════════════════════════════════════════════
# MAIN APP
# ══════════════════════════════════════════════════════════════════════
TEMP_EXPORT = "/tmp/premisas_export.xlsx"

def export_init(plantilla_bytes):
    import gc, os; gc.collect()
    wb = load_workbook(BytesIO(plantilla_bytes), keep_links=False)
    for s in ["Lineas","Datos","LIBRANZAS NUEVAS","LIBRANZAS VIEJAS"]:
        if s in wb.sheetnames: del wb[s]
    wb.save(TEMP_EXPORT); wb.close(); del wb; gc.collect()

def export_write_indisp(indisp_data, weeks, current_week):
    import gc; gc.collect()
    WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
    BLUE_FILL  = PatternFill("solid", fgColor="BDD7EE")
    TOTAL_FONT = Font(bold=True, size=9)
    NO_BORDER  = Border()
    DATA_BORD  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    wb = load_workbook(TEMP_EXPORT, keep_links=False)
    ws = wb["Indisponibilidades"]
    for merge in list(ws.merged_cells.ranges): ws.unmerge_cells(str(merge))
    try: ws.cell(1,1).value = f"Indisponibilidades de Generación Intersemanales - Semana {current_week}"
    except: pass
    try: ws.cell(2,1).value = f"Semana {current_week}"
    except: pass
    _clear_rows(ws, 3)
    week_dates = weeks.get(current_week, [None]*7)
    sab = week_dates[0]
    fixed_dates = []
    for i, d in enumerate(week_dates):
        if d is None: fixed_dates.append(None)
        elif sab and abs((d-sab).days)>7: fixed_dates.append(sab + timedelta(days=i))
        else: fixed_dates.append(d)
    ri = 3; weekly_total = 0.0
    for i, dia in enumerate(DIAS_SEMANA):
        dt = fixed_dates[i] if i < len(fixed_dates) else None
        if dt is None: continue
        day_date = dt.date()
        day_rows = []
        if indisp_data is not None and not (hasattr(indisp_data,"empty") and indisp_data.empty):
            for _, row_data in indisp_data.iterrows():
                fi = parse_dt(row_data.get("Fecha inicio",""))
                ff = parse_dt(row_data.get("Fecha final",""))
                if fi and ff and fi.date() <= day_date <= ff.date():
                    day_rows.append(row_data)
        if not day_rows: continue
        day_rows = sorted(day_rows, key=lambda r: (0 if str(r.get("status",""))=="vieja" else 1, parse_dt(r.get("Fecha inicio","")) or datetime.max))
        day_start_ri = ri; day_total = 0.0
        for row_data in day_rows:
            is_vieja = str(row_data.get("status",""))=="vieja"
            for ci, col in enumerate(INDISP_COLS, 2):
                val = row_data.get(col,"") if hasattr(row_data,"get") else ""
                try:
                    cell = ws.cell(ri, ci, val); cell.font=DATA_FONT; cell.border=DATA_BORD; cell.fill=WHITE_FILL
                    if is_vieja and ci==8: cell.fill=BLUE_FILL; cell.font=Font(size=9,bold=True)
                except: pass
            mw = float(row_data.get("Potencia (MW)",0) or 0) if hasattr(row_data,"get") else 0
            day_total+=mw; weekly_total+=mw; ri+=1
        day_end_ri = ri-1
        try:
            if day_end_ri>day_start_ri: ws.merge_cells(f"A{day_start_ri}:A{day_end_ri}")
            ca=ws.cell(day_start_ri,1); ca.value=f"{dia.lower()} {day_date.day}"
            ca.font=Font(size=9,bold=True); ca.alignment=Alignment(vertical="center",horizontal="center",wrap_text=True)
            ca.border=DATA_BORD; ca.fill=WHITE_FILL
        except: pass
        try:
            ws.merge_cells(f"B{ri}:F{ri}")
            cl=ws.cell(ri,2); cl.value="INDISPONIBILIDAD TOTAL EN HORAS PUNTA (MW):"; cl.font=TOTAL_FONT; cl.fill=WHITE_FILL; cl.border=NO_BORDER
            cm=ws.cell(ri,7); cm.value=round(day_total,4); cm.font=TOTAL_FONT; cm.fill=WHITE_FILL; cm.border=NO_BORDER
            for ci in [1,3,4,5,6,8,9]:
                try: c=ws.cell(ri,ci); c.value=None; c.fill=WHITE_FILL; c.border=NO_BORDER
                except: pass
        except: pass
        ri+=2
    try:
        ws.merge_cells(f"B{ri}:F{ri}")
        cl=ws.cell(ri,2); cl.value="INDISPONIBILIDAD TOTAL EN HORAS PUNTA (MW):"; cl.font=Font(bold=True,size=9,color="FF0000"); cl.fill=WHITE_FILL; cl.border=Border()
        cm=ws.cell(ri,7); cm.value=round(weekly_total,4); cm.font=Font(bold=True,size=9,color="FF0000"); cm.fill=WHITE_FILL; cm.border=Border()
    except: pass
    for row in ws.iter_rows(min_row=ri+1, max_row=min(ri+200, ws.max_row)):
        for cell in row:
            try: cell.value=None; cell.fill=WHITE_FILL; cell.border=NO_BORDER
            except: pass
    wb.save(TEMP_EXPORT); wb.close(); del wb; gc.collect()

def export_write_relevantes(df_relevantes):
    import gc; gc.collect()
    wb = load_workbook(TEMP_EXPORT, keep_links=False)
    ws = wb["Libranzas Relevantes"]
    for merge in list(ws.merged_cells.ranges): ws.unmerge_cells(str(merge))
    _clear_rows(ws, 3); _hdr(ws, PLANTILLA_REL_COLS, row=2)
    if df_relevantes is not None and not df_relevantes.empty:
        BLUE_FILL_R=PatternFill("solid",fgColor="BDD7EE"); GREEN_FILL=PatternFill("solid",fgColor="C6EFCE")
        PINK_FILL=PatternFill("solid",fgColor="FFB6C1"); WHITE_FILL_R=PatternFill("solid",fgColor="FFFFFF")
        for ri,(_, r) in enumerate(df_relevantes.iterrows(), 3):
            is_vieja=str(r.get("_status",""))=="vieja" if "_status" in r.index else False
            tipo_val=str(r.get("Tipo","")).strip().lower()
            for ci, col in enumerate(PLANTILLA_REL_COLS, 1):
                val=r.get(col,"") if col in r.index else ""
                try:
                    cell=ws.cell(ri,ci,val); cell.font=DATA_FONT; cell.border=BORD; cell.alignment=WRAP; cell.fill=WHITE_FILL_R
                    if ci==1: cell.fill=GREEN_FILL if "continua" in tipo_val else (PINK_FILL if "repetitiva" in tipo_val else WHITE_FILL_R)
                    elif ci==9 and is_vieja: cell.fill=BLUE_FILL_R; cell.font=Font(size=9,bold=True)
                except: pass
    wb.save(TEMP_EXPORT); wb.close(); del wb; gc.collect()

def export_write_proyectos(df_proyectos):
    import gc; gc.collect()
    wb = load_workbook(TEMP_EXPORT, keep_links=False)
    ws = wb["Proyectos de Generacion"]
    for merge in list(ws.merged_cells.ranges): ws.unmerge_cells(str(merge))
    start_row = 14
    for i, row in enumerate(ws.iter_rows(min_row=13, max_row=20, values_only=True), 13):
        if row[0]=="Planta": start_row=i+1; break
    _clear_rows(ws, start_row)
    if df_proyectos is not None and not df_proyectos.empty:
        for ri2,(_, r) in enumerate(df_proyectos.iterrows(), start_row):
            for ci, col in enumerate(PROY_COLS, 1):
                try: ws.cell(ri2, ci, r.get(col,"")).font=DATA_FONT
                except: pass
    wb.save(TEMP_EXPORT); wb.close(); del wb; gc.collect()

def export_get_bytes():
    import os
    if not os.path.exists(TEMP_EXPORT): return None
    with open(TEMP_EXPORT,"rb") as f: return f.read()


def vista_premisas():
    st.markdown("## ⚡ Módulo de Premisas")

    # ── Session state init ────────────────────────────────────────────
    KEYS = ["prem_plantilla_bytes","prem_current_week","prem_weeks",
            "prem_lineas_lookup","prem_unit_mw","prem_plant_prefix",
            "prem_lineas_codes","prem_indisp_existing","prem_rel_ant",
            "prem_week_start","prem_df_nuevas","prem_df_viejas",
            "prem_indisp_data","prem_df_relevantes","prem_df_proyectos",
            "prem_export_bytes","prem_step"]
    for k in KEYS:
        if k not in st.session_state:
            st.session_state[k] = None
    if st.session_state.prem_step is None:
        st.session_state.prem_step = 0

    step = st.session_state.prem_step

    # ══════════════════════════════════════════════════════════════════
    # SIDEBAR
    # ══════════════════════════════════════════════════════════════════
    with st.sidebar:
        st.title("📂 Archivos")
        st.divider()

        f_plantilla = st.file_uploader("1. Plantilla semana anterior", type="xlsx", key="prem_up_plantilla")
        f_nuevas    = st.file_uploader("2. libranzas_nuevas.xlsx",     type="xlsx", key="prem_up_nuevas")
        f_viejas    = st.file_uploader("3. libranzas_viejas.xlsx",     type="xlsx", key="prem_up_viejas")
        f_indisp    = st.file_uploader("4. Indisponibilidades (opcional)", type="xlsx", key="prem_up_indisp")

        st.divider()

        # ── PASO 1: Cargar plantilla ──────────────────────────────────
        st.markdown("**Paso 1 — Cargar plantilla**")
        btn1 = st.button("📂 Cargar plantilla", use_container_width=True,
                         disabled=not f_plantilla,
                         key="prem_btn1")
        if btn1 and f_plantilla:
            with st.spinner("Leyendo plantilla..."):
                try:
                    pb = f_plantilla.read()
                    pl = load_plantilla(pb)
                    week_dates = pl["weeks"].get(pl["current_week"], [None]*7)
                    sab = week_dates[0]
                    fixed = [sab + timedelta(days=i) if (d and sab and abs((d-sab).days)>7) else d
                             for i,d in enumerate(week_dates)]
                    st.session_state.update({
                        "prem_plantilla_bytes":  pb,
                        "prem_current_week":     pl["current_week"],
                        "prem_weeks":            pl["weeks"],
                        "prem_lineas_lookup":    pl["lineas_lookup"],
                        "prem_unit_mw":          pl.get("unit_mw",{}),
                        "prem_plant_prefix":     pl.get("plant_prefix",{}),
                        "prem_lineas_codes":     pl.get("lineas_codes",set()),
                        "prem_indisp_existing":  pl.get("indisp_existing",[]),
                        "prem_rel_ant":          pl.get("relevantes_anteriores",[]),
                        "prem_week_start":       next((d.date() for d in fixed if d), None),
                        "prem_df_proyectos":     pl.get("proyectos", None),
                        "prem_df_nuevas":        None,
                        "prem_df_viejas":        None,
                        "prem_indisp_data":      None,
                        "prem_df_relevantes":    None,
                        "prem_export_bytes":     None,
                        "prem_step":             1,
                    })
                    st.success(f"✅ Semana {pl['current_week']} cargada")
                    st.rerun()
                except Exception as e:
                    st.error(f"Error: {e}")

        # ── PASO 2: Filtrar libranzas ─────────────────────────────────
        st.markdown("**Paso 2 — Filtrar libranzas**")
        btn2 = st.button("🔍 Filtrar libranzas", use_container_width=True,
                         disabled=(step < 1 or not f_nuevas or not f_viejas),
                         key="prem_btn2")
        if btn2 and f_nuevas and f_viejas and step >= 1:
            with st.spinner("Filtrando..."):
                try:
                    lc = st.session_state.prem_lineas_codes
                    um = st.session_state.prem_unit_mw
                    pp = st.session_state.prem_plant_prefix
                    df_n = process_nuevas(load_source_libranzas(f_nuevas.read()), lc, um, pp)
                    df_v = process_viejas(load_source_libranzas(f_viejas.read()), lc, um, pp)
                    st.session_state.prem_df_nuevas   = df_n
                    st.session_state.prem_df_viejas   = df_v
                    st.session_state.prem_indisp_data = None
                    st.session_state.prem_df_relevantes = None
                    st.session_state.prem_export_bytes  = None
                    st.session_state.prem_step = 2
                    n_r = df_n["R/I"].str.contains("R", na=False).sum()
                    n_i = df_n["R/I"].str.contains("I", na=False).sum()
                    st.success(f"✅ {len(df_n)} nuevas | {len(df_v)} viejas | R:{n_r} I:{n_i}")
                    st.rerun()
                except Exception as e:
                    st.error(f"Error: {e}")

        # ── PASO 3: Indisponibilidades ────────────────────────────────
        st.markdown("**Paso 3 — Indisponibilidades**")
        btn3 = st.button("⚠️ Generar Indisponibilidades", use_container_width=True,
                         disabled=(step < 2), key="prem_btn3")
        if btn3 and step >= 2:
            with st.spinner("Generando indisponibilidades..."):
                try:
                    indisp_file_df = load_indisp_file(f_indisp.read()) if f_indisp else None
                    indisp = build_indisponibilidades(
                        st.session_state.prem_indisp_existing,
                        st.session_state.prem_df_viejas,
                        st.session_state.prem_df_nuevas,
                        st.session_state.prem_unit_mw,
                        st.session_state.prem_plant_prefix,
                        st.session_state.prem_weeks,
                        st.session_state.prem_current_week,
                        indisp_file_df
                    )
                    st.session_state.prem_indisp_data = indisp
                    st.session_state.prem_export_bytes = None
                    st.session_state.prem_step = max(st.session_state.prem_step, 3)
                    n = len(indisp) if indisp is not None and not (hasattr(indisp,"empty") and indisp.empty) else 0
                    st.success(f"✅ {n} indisponibilidades")
                    st.rerun()
                except Exception as e:
                    st.error(f"Error: {e}")

        # ── PASO 4: Relevantes ────────────────────────────────────────
        st.markdown("**Paso 4 — Relevantes**")
        btn4 = st.button("🔑 Generar Relevantes", use_container_width=True,
                         disabled=(step < 2), key="prem_btn4")
        if btn4 and step >= 2:
            with st.spinner("Generando relevantes..."):
                try:
                    df_rel = build_relevantes(
                        st.session_state.prem_df_nuevas,
                        st.session_state.prem_df_viejas,
                        st.session_state.prem_lineas_lookup,
                        st.session_state.prem_rel_ant,
                        st.session_state.prem_week_start
                    )
                    st.session_state.prem_df_relevantes = df_rel
                    st.session_state.prem_export_bytes  = None
                    st.session_state.prem_step = max(st.session_state.prem_step, 4)
                    st.success(f"✅ {len(df_rel)} relevantes")
                    st.rerun()
                except Exception as e:
                    st.error(f"Error: {e}")

        # ── PASO 5: Proyectos ─────────────────────────────────────────
        st.markdown("**Paso 5 — Proyectos**")
        btn5 = st.button("🏗️ Actualizar Proyectos", use_container_width=True,
                         disabled=(step < 2), key="prem_btn5")
        if btn5 and step >= 2:
            with st.spinner("Actualizando proyectos..."):
                try:
                    df_p = update_proyectos(
                        st.session_state.prem_df_proyectos,
                        st.session_state.prem_current_week,
                        st.session_state.prem_weeks
                    )
                    df_p = detect_proyectos_from_libranzas(
                        st.session_state.prem_df_viejas,
                        st.session_state.prem_df_nuevas,
                        df_p
                    )
                    st.session_state.prem_df_proyectos = df_p
                    st.session_state.prem_export_bytes  = None
                    st.session_state.prem_step = max(st.session_state.prem_step, 5)
                    st.success(f"✅ {len(df_p)} proyectos")
                    st.rerun()
                except Exception as e:
                    st.error(f"Error: {e}")

        # ── Info semana ───────────────────────────────────────────────
        if st.session_state.prem_current_week:
            st.divider()
            cw = st.session_state.prem_current_week
            wdates = st.session_state.prem_weeks.get(cw,[]) if st.session_state.prem_weeks else []
            st.markdown(f'<div class="week-badge">Semana {cw}</div>', unsafe_allow_html=True)
            st.caption(week_range(wdates))
            c1,c2 = st.columns(2)
            df_n = st.session_state.prem_df_nuevas
            df_v = st.session_state.prem_df_viejas
            c1.metric("Nuevas", len(df_n) if df_n is not None else 0)
            c2.metric("Viejas", len(df_v) if df_v is not None else 0)

        # ── Exportación por pasos ─────────────────────────────────────
        st.divider()
        st.markdown("**Exportación**")
        can_init = st.session_state.prem_plantilla_bytes is not None

        if st.button("📂 1. Inicializar exportación", use_container_width=True,
                     disabled=not can_init, key="prem_exp_init"):
            with st.spinner("Inicializando..."):
                try:
                    export_init(st.session_state.prem_plantilla_bytes)
                    st.session_state.prem_exp_step = 1
                    st.success("✅ Listo"); st.rerun()
                except Exception as e: st.error(str(e))

        exp_step = st.session_state.get("prem_exp_step", 0)

        if st.button("⚠️ 2. Escribir Indisponibilidades", use_container_width=True,
                     disabled=(exp_step < 1), key="prem_exp_indisp"):
            with st.spinner("Escribiendo indisponibilidades..."):
                try:
                    export_write_indisp(st.session_state.prem_indisp_data,
                                        st.session_state.prem_weeks,
                                        st.session_state.prem_current_week)
                    st.session_state.prem_exp_step = max(exp_step, 2)
                    st.success("✅ Listo"); st.rerun()
                except Exception as e: st.error(str(e))

        if st.button("🔑 3. Escribir Relevantes", use_container_width=True,
                     disabled=(exp_step < 1), key="prem_exp_rel"):
            with st.spinner("Escribiendo relevantes..."):
                try:
                    export_write_relevantes(st.session_state.prem_df_relevantes)
                    st.session_state.prem_exp_step = max(exp_step, 3)
                    st.success("✅ Listo"); st.rerun()
                except Exception as e: st.error(str(e))

        if st.button("🏗️ 4. Escribir Proyectos", use_container_width=True,
                     disabled=(exp_step < 1), key="prem_exp_proy"):
            with st.spinner("Escribiendo proyectos..."):
                try:
                    export_write_proyectos(st.session_state.prem_df_proyectos)
                    st.session_state.prem_exp_step = max(exp_step, 4)
                    st.success("✅ Listo"); st.rerun()
                except Exception as e: st.error(str(e))

        if exp_step >= 1:
            cw = st.session_state.prem_current_week or 0
            data = export_get_bytes()
            if data:
                st.download_button(
                    "📥 Descargar plantilla",
                    data=data,
                    file_name=f"Premisas_SEM_{cw:02d}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

    # ══════════════════════════════════════════════════════════════════
    # MAIN CONTENT
    # ══════════════════════════════════════════════════════════════════
    if st.session_state.prem_df_nuevas is None and st.session_state.prem_step == 0:
        st.info("👈 Cargue los archivos en el panel lateral y siga los pasos para comenzar.")
        return
    elif st.session_state.prem_df_nuevas is None:
        st.info("👈 Continúe con el Paso 2 — Filtrar libranzas.")
        return

    cw    = st.session_state.prem_current_week
    weeks = st.session_state.prem_weeks

# ── Standalone execution ──────────────────────────────────────────────
try:
    st.set_page_config(
        page_title="Premisas CND", page_icon="⚡",
        layout="wide", initial_sidebar_state="expanded"
    )
except Exception:
    pass  # already set by parent app when imported
vista_premisas()
